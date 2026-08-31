"""问卷数据标注模块

提供两个独立于报告生成流程之外的标注功能：
1. AI 作答识别 (ai_detect)：判断受访者是否使用 AI 填写主观题
2. 回答质量打标 (quality)：为每道主观题和每位受访者整体打 无效/普通/优秀 标签

两个功能可独立使用，也可组合（先 AI 识别，用户确认后再打标）。
"""

import io
import json
import re
from typing import Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from app.core.config import (
    ANNOTATE_QUALITY_EXCELLENT_AVG_THRESHOLD,
    ANNOTATE_QUALITY_EXCELLENT_MAX_INVALID_RATIO,
    ANNOTATE_QUALITY_INVALID_AVG_THRESHOLD,
    ANNOTATE_QUALITY_INVALID_HARD_RATIO,
    ANNOTATE_QUALITY_LOW_EFFORT_MIN_ANSWERS,
    ANNOTATE_QUALITY_LOW_EFFORT_MIN_SIGNALS,
    ANNOTATE_QUALITY_LOW_EFFORT_MIN_STRUCTURED_ITEMS,
    ANNOTATE_QUALITY_SHORT_TEXT_MAX_CHARS,
    ANNOTATE_QUALITY_SHORT_TEXT_MAX_WORDS,
)

QUALITY_LABELS = {"无效反馈", "普通反馈", "优秀反馈", "N/A"}

# 标注列样式
_YELLOW_FILL  = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
_GRAY_FILL    = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
_HEADER_FILL  = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
_BOLD_FONT    = Font(bold=True)

# ============================================================
# 列检测
# ============================================================

def detect_id_column(headers: list[str], rows: list[list]) -> int:
    """检测玩家 ID 列。
    优先匹配常见 ID 关键词，找不到时默认第二列（index 1），与用户 prompt 约定一致。
    """
    for i, h in enumerate(headers):
        hl = h.lower().strip()
        if any(kw in hl for kw in [
            "discordid", "discord_id", "discord", "whatsapp",
            "mlbbid", "mlbb_id", "player_id", "playerid",
            "玩家id", "用户id", "respondent",
        ]):
            return i
    return min(1, len(headers) - 1)


def detect_open_text_columns(rows: list[list], headers: list[str]) -> list[int]:
    """检测主观题（开放题）列，逻辑与 server.py 的 _heuristic_type 保持一致。"""
    if len(rows) <= 1:
        return []
    body = rows[1:]
    result = []
    for i, header in enumerate(headers):
        h = header.lower().strip()
        # 跳过时间戳等
        if any(kw in h for kw in ["时间", "timestamp", "submit", "提交", "date", "日期"]):
            continue
        vals = [str(r[i]) if i < len(r) else "" for r in body]
        non_empty = [v.strip() for v in vals if v.strip()]
        if not non_empty:
            continue
        total = len(non_empty)
        # 纯数字列 → 量表
        nums = sum(1 for v in non_empty if _try_float(v))
        if nums / total > 0.85:
            continue
        # 多选题（分隔符）
        delim_count = sum(1 for v in non_empty if any(d in v for d in [",", "，", ";", "；", "、", "|"]))
        if delim_count / total > 0.25:
            continue
        # 唯一值少 → 选择题
        unique_vals = set(non_empty)
        if len(unique_vals) <= 8 or len(unique_vals) / total < 0.25:
            continue
        # 长文本 → 开放题
        avg_len = sum(len(v) for v in non_empty) / total
        if avg_len > 25:
            result.append(i)
    return result


def _try_float(s: str) -> bool:
    try:
        float(s)
        return True
    except ValueError:
        return False


# ============================================================
# 格式化工具
# ============================================================

def _rows_to_md_table(
    batch_rows: list[list],
    headers: list[str],
    col_indexes: list[int],
    max_cell_chars: int | None = None,
) -> str:
    """把选定列格式化为 Markdown 表格。"""
    sel_headers = [headers[i] if i < len(headers) else f"列{i}" for i in col_indexes]

    def esc(s: str) -> str:
        text = str(s).replace("|", "\\|").replace("\n", " ").strip()
        if max_cell_chars and len(text) > max_cell_chars:
            return text[:max_cell_chars].rstrip() + "…（已截断）"
        return text

    lines = ["| " + " | ".join(esc(h) for h in sel_headers) + " |"]
    lines.append("| " + " | ".join(["---"] * len(sel_headers)) + " |")
    for row in batch_rows:
        cells = [str(row[i]) if i < len(row) else "" for i in col_indexes]
        lines.append("| " + " | ".join(esc(c) for c in cells) + " |")
    return "\n".join(lines)


# ============================================================
# Query 构建
# ============================================================

_AI_DETECT_INPUT_TMPL = """\
任务模式：AI 内容生成识别
批次：{batch_num}
玩家数量：{total}
{background_block}输入数据如下。第一列是玩家唯一 ID，其余列是该玩家的全部主观题回答：

{table}
"""

_QUALITY_INPUT_TMPL = """\
任务模式：逐题反馈质量打标
批次：{batch_num}
玩家数量：{total}
需要逐题返回的列：{col_desc}

输入数据如下。第一列是玩家唯一 ID，其余列是需要独立判断的主观题回答：

{table}
"""


def build_ai_detect_query(
    batch_rows: list[list],
    headers: list[str],
    open_text_cols: list[int],
    id_col: int,
    batch_num: int | str = 1,
    background: str = "",
) -> str:
    """构建 AI 作答识别模型查询。"""
    cols = [id_col] + [c for c in open_text_cols if c != id_col]
    table = _rows_to_md_table(batch_rows, headers, cols)
    bg_block = f"调研背景参考：{background.strip()}\n" if background.strip() else ""
    return _AI_DETECT_INPUT_TMPL.format(
        background_block=bg_block,
        batch_num=batch_num,
        total=len(batch_rows),
        table=table,
    )


def build_quality_label_query(
    batch_rows: list[list],
    headers: list[str],
    open_text_cols: list[int],
    id_col: int,
    batch_num: int | str = 1,
    include_translations: bool = True,
) -> str:
    """构建回答质量识别模型查询。"""
    cols = [id_col] + [c for c in open_text_cols if c != id_col]
    table = _rows_to_md_table(batch_rows, headers, cols)
    col_desc = "、".join(
        f"「{headers[c] if c < len(headers) else f'列{c}'}」(col_{c})"
        for c in open_text_cols
    )
    return _QUALITY_INPUT_TMPL.format(
        batch_num=batch_num,
        total=len(batch_rows),
        col_desc=col_desc,
        table=table,
    )


def build_translation_repair_query(items: list[dict]) -> str:
    """构建只修复缺失中文翻译的紧凑查询。"""
    payload = [
        {
            "id": str(item.get("id", "")),
            "key": str(item.get("key", "")),
            "text": str(item.get("text", "")),
        }
        for item in items
    ]
    return json.dumps(payload, ensure_ascii=False)


# ============================================================
# 结果解析
# ============================================================

def _repair_json_quotes(text: str) -> str:
    """将 JSON 字符串值内部的裸双引号转义为 \\\"。
    判断依据：当前 " 不是被 \\ 转义的，且下一个非空白字符不是 JSON 结构符（: , } ] \\n \\r），
    则认为它是值内部的裸引号而非字符串结束符。
    """
    out: list[str] = []
    in_str = False
    i = 0
    while i < len(text):
        c = text[i]
        if in_str:
            if c == '\\' and i + 1 < len(text):
                out.append(c)
                out.append(text[i + 1])
                i += 2
                continue
            elif c == '"':
                j = i + 1
                while j < len(text) and text[j] in ' \t':
                    j += 1
                nxt = text[j] if j < len(text) else ''
                if nxt in ':,}]\n\r':
                    out.append('"')
                    in_str = False
                else:
                    out.append('\\"')
            else:
                out.append(c)
        else:
            out.append(c)
            if c == '"':
                in_str = True
        i += 1
    return ''.join(out)


def _extract_json_array(text: str) -> Optional[list]:
    """从 LLM 输出中提取 JSON 数组，兼容 ```json 围栏，容忍字符串内裸双引号。"""
    def _try_parse(raw: str) -> Optional[list]:
        try:
            result = json.loads(raw)
            return result if isinstance(result, list) else None
        except json.JSONDecodeError:
            pass
        try:
            result = json.loads(_repair_json_quotes(raw))
            return result if isinstance(result, list) else None
        except json.JSONDecodeError:
            pass
        return None

    m = re.search(r"```(?:json)?\s*(\[.*?\])\s*```", text, re.DOTALL)
    if m:
        result = _try_parse(m.group(1))
        if result is not None:
            return result
    # 尝试裸数组（贪婪）
    m = re.search(r"\[.*\]", text, re.DOTALL)
    if m:
        result = _try_parse(m.group(0))
        if result is not None:
            return result
    return None


def parse_ai_detect_result(llm_output: str) -> tuple[list[dict], str]:
    """解析 AI 检测结果。
    Returns: (results, error_msg) — results 为空列表表示解析失败。
    每条结果：{id, ai_prob, polish_prob, reason, evidence, counter_evidence, translations}
    """
    arr = _extract_json_array(llm_output)
    if arr is None:
        return [], "无法从输出中提取 JSON 数组"
    results = []
    for item in arr:
        if not isinstance(item, dict):
            continue
        row_id = str(item.get("id", "")).strip()
        ai_prob = _strict_probability(item.get("ai_prob"))
        polish_prob = _strict_probability(item.get("polish_prob"))
        reason = str(item.get("reason", "")).strip()
        translations = item.get("translations") or {}
        if not row_id or ai_prob is None or polish_prob is None or not reason:
            continue
        if not isinstance(translations, dict):
            continue
        results.append({
            "id": row_id,
            "ai_prob": ai_prob,
            "polish_prob": polish_prob,
            "reason": reason,
            "evidence": str(item.get("evidence", "")).strip(),
            "counter_evidence": str(item.get("counter_evidence", "")).strip(),
            "translations": dict(translations),
        })
    return (results, "") if results else ([], "JSON 数组内没有符合 AI schema 的结果")


def parse_quality_result(llm_output: str) -> tuple[list[dict], str]:
    """解析质量打标结果。
    Returns: (results, error_msg)
    每条结果：{id, q_labels, q_reasons, q_evidence, translations}
    """
    arr = _extract_json_array(llm_output)
    if arr is None:
        return [], "无法从输出中提取 JSON 数组"
    results = []
    for item in arr:
        if not isinstance(item, dict):
            continue
        row_id = str(item.get("id", "")).strip()
        q_labels = item.get("q_labels") or {}
        q_reasons = item.get("q_reasons") or {}
        q_evidence = item.get("q_evidence") or {}
        translations = item.get("translations") or {}
        if not row_id or not all(isinstance(value, dict) for value in (
            q_labels, q_reasons, q_evidence, translations,
        )):
            continue
        results.append({
            "id": row_id,
            "q_labels": dict(q_labels),
            "q_reasons": dict(q_reasons),
            "q_evidence": dict(q_evidence),
            "translations": dict(translations),
        })
    return (results, "") if results else ([], "JSON 数组内没有符合质量 schema 的结果")


def parse_translation_repair_result(llm_output: str) -> tuple[list[dict], str]:
    """解析逐单元格中文翻译修复结果。"""
    arr = _extract_json_array(llm_output)
    if arr is None:
        return [], "无法从输出中提取翻译 JSON 数组"
    results = []
    for item in arr:
        if not isinstance(item, dict):
            continue
        row_id = str(item.get("id", "")).strip()
        key = str(item.get("key", "")).strip()
        translation = str(item.get("translation", "")).strip()
        if row_id and key.startswith("col_") and translation:
            results.append({"id": row_id, "key": key, "translation": translation})
    return (results, "") if results else ([], "翻译 JSON 数组内没有有效结果")


def _strict_probability(val) -> int | None:
    try:
        parsed = int(val)
    except (TypeError, ValueError):
        return None
    return parsed if 0 <= parsed <= 100 else None


_QUALITY_SCORE = {"无效反馈": 0, "普通反馈": 1, "优秀反馈": 2}
_SCORE_HEADER_RE = re.compile(r"(?:\brate\b|\brating\b|\bscore\b|评分|打分|分数|量表)", re.IGNORECASE)
_RANK_HEADER_RE = re.compile(r"(?:\brank(?:ing)?\b|\border\b|排序|排行|名次|优先级)", re.IGNORECASE)


def _header_text(headers: list, headers_zh: list, index: int) -> str:
    original = str(headers[index]).strip() if index < len(headers) else ""
    translated = str(headers_zh[index]).strip() if index < len(headers_zh) else ""
    return f"{original} {translated}".strip()


def _strict_number(value: object) -> float | None:
    text = str(value or "").strip()
    if not re.fullmatch(r"[-+]?\d+(?:\.\d+)?", text):
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _header_preview(headers: list, headers_zh: list, indexes: list[int]) -> str:
    names = []
    for index in indexes[:3]:
        name = _header_text(headers, headers_zh, index) or f"列{index}"
        names.append(name[:40])
    suffix = "等" if len(indexes) > 3 else ""
    return "、".join(names) + suffix


def _mechanical_ranking_evidence(
    row: list,
    headers: list,
    headers_zh: list,
    candidate_cols: list[int],
) -> str:
    """只识别明确的 A→B→C 或 1→2→3 展示顺序，不推断语义排序。"""
    minimum = ANNOTATE_QUALITY_LOW_EFFORT_MIN_STRUCTURED_ITEMS
    numeric_values: list[int] = []
    numeric_cols: list[int] = []
    for col in candidate_cols:
        value = _strict_number(row[col] if col < len(row) else "")
        if value is None or not value.is_integer():
            continue
        numeric_values.append(int(value))
        numeric_cols.append(col)
    if len(numeric_values) >= minimum and numeric_values == list(range(1, len(numeric_values) + 1)):
        return (
            f"{len(numeric_values)}个排序项按问卷列顺序填写为"
            f"{'→'.join(str(value) for value in numeric_values)}（"
            f"{_header_preview(headers, headers_zh, numeric_cols)}）"
        )

    for col in candidate_cols:
        raw = str(row[col] if col < len(row) else "").strip()
        parts = [
            part.strip()
            for part in re.split(r"[,，;；、>|\n]+", raw)
            if part.strip()
        ]
        if len(parts) < minimum:
            continue
        letters: list[str] = []
        for part in parts:
            match = re.search(
                r"(?:\b(?:track|song|music|option|item)\s*)?\b([A-Z])\b",
                part,
                re.IGNORECASE,
            )
            if not match:
                letters = []
                break
            letters.append(match.group(1).upper())
        expected = [chr(ord("A") + index) for index in range(len(letters))]
        if letters and letters == expected:
            return (
                f"排序题按展示字母顺序填写为{'→'.join(letters)}（"
                f"{_header_preview(headers, headers_zh, [col])}）"
            )
    return ""


def detect_low_effort_signals(
    row: list,
    headers: list,
    open_text_cols: list[int],
    id_col: int,
    q_labels: dict[str, str],
    *,
    headers_zh: list | None = None,
) -> dict:
    """返回可审计的整卷低投入组合信号；单项信号绝不触发硬门槛。"""
    translated_headers = headers_zh or []
    excluded = set(open_text_cols) | {id_col}
    score_cols: list[int] = []
    score_values: list[float] = []
    rank_cols: list[int] = []
    for col in range(max(len(headers), len(row))):
        if col in excluded:
            continue
        header = _header_text(headers, translated_headers, col)
        if _RANK_HEADER_RE.search(header):
            rank_cols.append(col)
        if _SCORE_HEADER_RE.search(header):
            value = _strict_number(row[col] if col < len(row) else "")
            if value is not None:
                score_cols.append(col)
                score_values.append(value)

    signals: list[dict[str, str]] = []
    minimum_structured = ANNOTATE_QUALITY_LOW_EFFORT_MIN_STRUCTURED_ITEMS
    if len(score_values) >= minimum_structured and len(set(score_values)) == 1:
        value = score_values[0]
        display_value = str(int(value)) if value.is_integer() else str(value)
        signals.append({
            "code": "uniform_scores",
            "kind": "structured",
            "evidence": (
                f"{len(score_values)}个明确评分项全部为{display_value}（"
                f"{_header_preview(headers, translated_headers, score_cols)}）"
            ),
        })

    ranking_evidence = _mechanical_ranking_evidence(
        row, headers, translated_headers, rank_cols,
    )
    if ranking_evidence:
        signals.append({
            "code": "mechanical_ranking",
            "kind": "structured",
            "evidence": ranking_evidence,
        })

    answers = []
    for col in open_text_cols:
        if str(q_labels.get(f"col_{col}", "N/A")) == "N/A":
            continue
        text = str(row[col] if col < len(row) else "").strip()
        if text:
            answers.append(text)
    if len(answers) >= ANNOTATE_QUALITY_LOW_EFFORT_MIN_ANSWERS:
        short_flags = []
        for text in answers:
            compact_length = len(re.sub(r"[^\w\u4e00-\u9fff]", "", text, flags=re.UNICODE))
            word_count = len(re.findall(r"[A-Za-z0-9]+", text))
            short_flags.append(
                compact_length <= ANNOTATE_QUALITY_SHORT_TEXT_MAX_CHARS
                or (word_count > 0 and word_count <= ANNOTATE_QUALITY_SHORT_TEXT_MAX_WORDS)
            )
        if all(short_flags):
            signals.append({
                "code": "pervasively_short_text",
                "kind": "text",
                "evidence": f"{len(answers)}道非N/A主观回答全部为极短表达",
            })

        normalized = [re.sub(r"[\W_]", "", text.lower(), flags=re.UNICODE) for text in answers]
        normalized = [text for text in normalized if text]
        if normalized:
            top_count = max(normalized.count(text) for text in set(normalized))
            if top_count >= 2 and top_count * 2 >= len(normalized):
                signals.append({
                    "code": "repeated_simple_text",
                    "kind": "text",
                    "evidence": f"{top_count}/{len(normalized)}道主观回答为规范化后完全重复的简单描述",
                })

    kinds = {signal["kind"] for signal in signals}
    triggered = (
        len(signals) >= ANNOTATE_QUALITY_LOW_EFFORT_MIN_SIGNALS
        and {"structured", "text"}.issubset(kinds)
    )
    return {"triggered": triggered, "signals": signals}


def calculate_overall_quality(
    q_labels: dict[str, str],
    open_text_cols: list[int],
    *,
    low_effort: dict | None = None,
) -> tuple[str, str]:
    """按非 N/A 题目加权计算整体质量，并返回可复核的计数与硬门槛说明。"""
    labels = [str(q_labels.get(f"col_{col}", "N/A")) for col in open_text_cols]
    assessed = [label for label in labels if label != "N/A"]
    if not assessed:
        return "N/A", "非N/A题目0道：无效0、普通0、优秀0；无可评估回答，整体为N/A"

    invalid_count = assessed.count("无效反馈")
    ordinary_count = assessed.count("普通反馈")
    excellent_count = assessed.count("优秀反馈")
    total = len(assessed)
    invalid_ratio = invalid_count / total
    weighted_total = sum(_QUALITY_SCORE.get(label, 0) for label in assessed)
    average = weighted_total / total
    majority_gate = invalid_ratio > ANNOTATE_QUALITY_INVALID_HARD_RATIO
    low_effort_result = low_effort or {"triggered": False, "signals": []}
    low_effort_gate = bool(low_effort_result.get("triggered"))

    if majority_gate or low_effort_gate:
        overall = "无效反馈"
    elif average < ANNOTATE_QUALITY_INVALID_AVG_THRESHOLD:
        overall = "无效反馈"
    elif (
        average >= ANNOTATE_QUALITY_EXCELLENT_AVG_THRESHOLD
        and invalid_ratio <= ANNOTATE_QUALITY_EXCELLENT_MAX_INVALID_RATIO
    ):
        overall = "优秀反馈"
    else:
        overall = "普通反馈"

    hard_reasons = []
    if majority_gate:
        hard_reasons.append(
            f"无效比例超过{ANNOTATE_QUALITY_INVALID_HARD_RATIO:.0%}"
        )
    if low_effort_gate:
        hard_reasons.append("整份答卷出现多项跨类型低投入信号")
    hard_text = f"已触发（{'；'.join(hard_reasons)}）" if hard_reasons else "未触发"

    signal_evidence = [
        str(signal.get("evidence", "")).strip()
        for signal in low_effort_result.get("signals", [])
        if str(signal.get("evidence", "")).strip()
    ]
    if signal_evidence:
        low_effort_text = (
            ("已触发" if low_effort_gate else "未触发")
            + f"（发现{len(signal_evidence)}项：{'；'.join(signal_evidence)}）"
        )
    else:
        low_effort_text = "未触发（未发现可审计的组合信号）"

    reason = (
        f"非N/A题目{total}道：无效{invalid_count}、普通{ordinary_count}、"
        f"优秀{excellent_count}；无效比例{invalid_ratio:.2%}；"
        f"加权总分{weighted_total}分、平均分{average:.2f}"
        f"（无效=0、普通=1、优秀=2）；整体硬门槛：{hard_text}；"
        f"低投入组合信号：{low_effort_text}；整体判为{overall}"
    )
    return overall, reason


# ============================================================
# Excel 生成
# ============================================================


def generate_annotated_excel(
    rows: list[list],
    headers: list[str],
    ai_results: list[dict],
    confirmed_ai_ids: set[str],
    quality_results: list[dict],
    open_text_cols: list[int],
    id_col: int,
    tasks: dict,
) -> bytes:
    """保留原表并追加可复核的 AI、质量、原文证据和中文翻译列。"""
    do_ai = bool(tasks.get("ai_detect"))
    do_quality = bool(tasks.get("quality"))
    id_to_ai = {str(result.get("id", "")): result for result in ai_results}
    id_to_quality = {str(result.get("id", "")): result for result in quality_results}
    open_text_set = set(open_text_cols)

    prefix_headers: list[str] = []
    if do_ai:
        prefix_headers.extend([
            "AI作答标签", "AI内容生成概率", "AI润色概率", "AI判断原因",
            "AI原文证据", "AI反向证据",
        ])
    if do_quality:
        prefix_headers.extend(["整体反馈质量", "整体质量原因"])

    col_spec: list[dict] = []
    for col_idx, header in enumerate(headers):
        if do_quality and col_idx in open_text_set:
            col_spec.extend([
                {"type": "quality_label", "header": f"[{header}]质量标注", "index": col_idx},
                {"type": "quality_reason", "header": f"[{header}]质量原因", "index": col_idx},
                {"type": "quality_evidence", "header": f"[{header}]原文证据", "index": col_idx},
            ])
        col_spec.append({"type": "original", "header": header, "index": col_idx})
        if col_idx in open_text_set and (do_ai or do_quality):
            col_spec.append({
                "type": "translation", "header": f"[{header}]中文翻译", "index": col_idx,
            })

    full_headers = prefix_headers + [spec["header"] for spec in col_spec]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "标注结果"
    ws.append(full_headers)

    annotation_suffixes = ("质量标注", "质量原因", "原文证据", "中文翻译")
    for cell in ws[1]:
        value = str(cell.value or "")
        if value in prefix_headers or value.endswith(annotation_suffixes):
            cell.fill = _HEADER_FILL
        cell.font = _BOLD_FONT
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    for row_data in rows[1:]:
        row_id = str(row_data[id_col]).strip() if id_col < len(row_data) else ""
        ai_info = id_to_ai.get(row_id, {})
        quality_info = id_to_quality.get(row_id, {})
        is_ai = row_id in confirmed_ai_ids
        translations: dict[str, str] = {}
        translations.update(ai_info.get("translations") or {})
        translations.update(quality_info.get("translations") or {})

        output_row: list = []
        if do_ai:
            output_row.extend([
                "高概率AI作答" if is_ai else "非高概率AI作答",
                ai_info.get("ai_prob", ""),
                ai_info.get("polish_prob", ""),
                ai_info.get("reason", ""),
                ai_info.get("evidence", ""),
                ai_info.get("counter_evidence", ""),
            ])
        if do_quality:
            output_row.extend([
                "高概率AI作答" if is_ai else quality_info.get("overall", ""),
                "已确认高概率AI作答，不进入质量打标" if is_ai else quality_info.get("overall_reason", ""),
            ])

        for spec in col_spec:
            col_idx = spec["index"]
            key = f"col_{col_idx}"
            spec_type = spec["type"]
            if spec_type == "quality_label":
                value = "-" if is_ai else (quality_info.get("q_labels") or {}).get(key, "")
            elif spec_type == "quality_reason":
                value = "-" if is_ai else (quality_info.get("q_reasons") or {}).get(key, "")
            elif spec_type == "quality_evidence":
                value = "-" if is_ai else (quality_info.get("q_evidence") or {}).get(key, "")
            elif spec_type == "translation":
                value = translations.get(key, "")
            else:
                value = row_data[col_idx] if col_idx < len(row_data) else ""
            output_row.append(value)

        ws.append(output_row)
        row_num = ws.max_row
        for cell in ws[row_num]:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        for column in range(1, len(prefix_headers) + 1):
            ws.cell(row=row_num, column=column).fill = _YELLOW_FILL
        for offset, spec in enumerate(col_spec, len(prefix_headers) + 1):
            if spec["type"] != "original":
                ws.cell(row=row_num, column=offset).fill = _YELLOW_FILL
        if is_ai:
            for column in range(1, len(full_headers) + 1):
                ws.cell(row=row_num, column=column).fill = _GRAY_FILL

    compact_headers = {"AI作答标签", "AI内容生成概率", "AI润色概率", "整体反馈质量"}
    wide_headers = {"AI判断原因", "AI原文证据", "AI反向证据", "整体质量原因"}
    for col_num, header in enumerate(full_headers, 1):
        letter = openpyxl.utils.get_column_letter(col_num)
        if header in compact_headers or header.endswith("质量标注"):
            width = 16
        elif header in wide_headers or header.endswith(("质量原因", "原文证据", "中文翻译")):
            width = 32
        else:
            width = 28
        ws.column_dimensions[letter].width = width

    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
