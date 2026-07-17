"""帖子评论舆情分析 · 预处理与统计（纯函数，无 Dify / 网络依赖）。

负责：文件解析、规则清洗、分层抽样、批次切分、本地统计聚合。
Dify 并发调用与 SSE 编排在 server.py 中完成，本模块只提供可单元测试的纯逻辑。

清洗规则、抽样比例、批次大小等阈值集中在本文件顶部，方便后续调参。
"""
import csv
import hashlib
import heapq
import io
import json
import random
import re
import unicodedata
from zipfile import BadZipFile

import openpyxl
from openpyxl.utils.exceptions import InvalidFileException

# ── 可调参数 ──────────────────────────────────────────────

# 评论内容列名匹配关键词（大小写不敏感，子串匹配）
COMMENT_COL_KEYS = ("message", "评论", "内容", "留言", "content", "text", "comment")
COMMENT_COL_EXCLUDE_KEYS = ("comment_id", "commentid", "评论id", "评论_id")

MIN_LEN = 15            # 规则①长度过短：strip 后长度 < 此值即丢弃
KB_MASH_RATIO = 0.6     # 规则②键盘乱打：单个高频字符占比 > 此值
KB_MASH_MINLEN = 20     # 规则②键盘乱打：仅对长度 > 此值的评论生效

SAMPLE_MAX_N = 5000     # 初始分层抽样目标；分析时先跑这部分
SAMPLE_POOL_MAX_N = 15000  # 动态补样池上限；相关评论不足时最多补到这里
RELATED_TARGET_N = 1000    # 进入后续分析的目标相关评论数
LONG_CANDIDATE_MAX_N = 1500  # 玩家原文精选候选：清洗后字数最多的评论数
# 长度分层边界与抽取比例（短 15~30 / 中 31~80 / 长 >80）
STRATA = (
    ("short", 15, 30, 0.50),
    ("mid", 31, 80, 0.35),
    ("long", 81, None, 0.15),
)

BATCH_SIZE = 250        # 主题提取批大小（5000 → 20 批）
RELEVANCE_BATCH_SIZE = 40  # 相关性判断要求逐条返回 idx；保持小批量降低漏返风险
CLASSIFY_BATCH_SIZE = 40  # 分类要求逐条返回；配合 idx 回填与自动拆分平衡速度和稳定性
QUOTE_SELECT_BATCH_SIZE = 75  # 玩家原文精选候选批大小
QUOTE_SELECT_BATCH_KEEP_N = 5  # 每批最多保留候选
QUOTE_SELECT_FINAL_POOL_N = 150  # 进入最终精选的 batch 候选上限，防止 final 输入过大
QUOTE_SELECT_FINAL_N = 50  # 最终最多展示评论原文数

OTHER_THEME_PCT = 3.0   # 占比低于此值（%）的主题合并入「其他」

# Excel 文件有时会因为整列格式、复制粘贴残留等原因，把 used range 撑到几十万
# 甚至 1048576 行。评论分析只需要抽样，不能在上传阶段无上限扫描空行。
COMMENT_MAX_SCAN_ROWS = 500_000
EXCEL_MAX_SCAN_ROWS = COMMENT_MAX_SCAN_ROWS
EXCEL_STOP_AFTER_BLANK_ROWS = 2_000
COLUMN_SAMPLE_ROWS = 500
PREPROCESS_PROGRESS_ROWS = 5_000

# 规则⑤纯索要奖励：整条评论只在索要钻石/皮肤/英雄/奖励等，无其他意见。
# 不使用复杂整句正则：社媒长评论会触发灾难性回溯，导致上传阶段卡死。
_REWARD_FILLER_WORDS = {
    "please", "pls", "plz", "sir", "admin", "hi", "hello", "hey", "free",
    "me", "my", "us", "the", "a", "some", "more", "give", "gift", "send",
    "want", "need", "gimme", "i", "we", "can", "you", "to", "for", "get",
    "got", "have", "has", "had", "one", "new", "chou", "epet",
    "pahingi", "minta", "bagi", "kasih", "gratis", "tolong", "dong",
    "hahaha", "haha", "hehe", "lol",
}
_REWARD_TARGET_WORDS = {
    "skin", "skins", "diamond", "diamonds", "hero", "heroes", "reward",
    "rewards", "gift", "gifts", "coin", "coins", "ticket", "tickets",
}
_REWARD_REQUEST_WORDS = {
    "please", "pls", "plz", "free", "give", "gift", "send", "want", "need",
    "gimme", "pahingi", "minta", "bagi", "kasih", "gratis", "tolong",
}
_REWARD_VALUE_WORDS = {
    "price", "expensive", "cheap", "cost", "worth", "buy", "purchase",
    "obtain", "obtained", "obtainable", "exchange", "redeem", "event",
    "task", "quest", "mission", "draw", "gacha", "token", "tokens",
    "hard", "difficult", "easy", "fair", "unfair", "why", "how", "when",
    "available", "release", "released", "path", "rule", "rules",
}
_REWARD_TOKEN_RE = re.compile(r"[a-z]+", re.I)
_REWARD_CN_RE = re.compile(
    r"^[\s，。！？、~*\-]*"
    r"(?:求|请|给我|送我|想要|要|跪求|跪求送|免费)+"
    r"(?:个|点|些)?"
    r"(?:免费)?"
    r"(?:钻石|皮肤|英雄|奖励|礼包|金币|点券)+"
    r"[\s，。！？、~*\-]*$"
)
_REWARD_CN_LOOSE_RE = re.compile(
    r"^(?=.{1,40}$)(?=.*(?:皮肤|钻石|英雄|奖励|礼包|金币|点券))"
    r"(?=.*(?:求|请|给我|送我|想要|跪求|免费|白嫖|送|来个|给点))"
    r"(?!.*(?:价格|太贵|获取|兑换|活动|规则|任务|难|容易|值得|什么时候|上线|怎么买|怎么拿|路径)).*$"
)


# ── 宽松 JSON 解析 ────────────────────────────────────────

def loads_loose(raw: str):
    """容错解析 LLM 返回的 JSON（dict 或 list 均可）。返回 (obj, err)。

    server.py 现有的 _json_loads_loose 只接受 JSON 对象，而本功能的
    extract / merge / classify 节点输出的是 JSON 数组，需单独处理。
    依次尝试：原文 → markdown 代码块内容 → [..] 边界 → {..} 边界。
    """
    raw = (raw or "").strip()
    if not raw:
        return None, "empty"
    candidates = [raw]
    fenced = re.search(r"```(?:json)?\s*(.*?)```", raw, re.S | re.I)
    if fenced:
        candidates.insert(0, fenced.group(1).strip())
    ls, le = raw.find("["), raw.rfind("]")
    if 0 <= ls < le:
        candidates.append(raw[ls:le + 1])
    bs, be = raw.find("{"), raw.rfind("}")
    if 0 <= bs < be:
        candidates.append(raw[bs:be + 1])

    last_err = ""
    for text in candidates:
        try:
            return json.loads(text), ""
        except Exception as e:  # noqa: BLE001 - 容错解析，任何异常都继续尝试
            last_err = str(e)
    return None, (last_err[:180] or "invalid json")


# ── 文件解析 ──────────────────────────────────────────────

def _decode_csv(content: bytes) -> str:
    """多编码兼容地解码 CSV（utf-8-sig 优先以兼容 BOM）。"""
    for enc in ("utf-8-sig", "utf-8", "gbk", "gb2312"):
        try:
            return content.decode(enc)
        except UnicodeDecodeError:
            continue
    raise ValueError("无法解析 CSV 文件，请确认文件编码为 UTF-8 或 GBK")


def _locate_comment_column(header: list[str]) -> int:
    """根据列名关键词定位评论列；匹配不到时回退到最“文本化”的列。

    返回列索引。回退策略：选择表头里第一个非空、且不像 ID/时间/数字列名的列。
    """
    keyword = _keyword_comment_column(header)
    if keyword:
        return keyword[0]
    # 回退：第一个非空表头列
    for idx, name in enumerate(header):
        if str(name).strip():
            return idx
    return 0


def _keyword_comment_column(header: list[str]) -> tuple[int, str] | None:
    normalized = [str(name).strip().lower() for name in header]
    for key in COMMENT_COL_KEYS:
        for idx, low in enumerate(normalized):
            if not low or any(ex in low for ex in COMMENT_COL_EXCLUDE_KEYS):
                continue
            if low == key:
                return idx, key
    for key in COMMENT_COL_KEYS:
        for idx, low in enumerate(normalized):
            if not low or any(ex in low for ex in COMMENT_COL_EXCLUDE_KEYS):
                continue
            if key in low:
                return idx, key
    return None


_ID_OR_TIME_HEADER_KEYS = (
    "id", "uid", "user", "name", "昵称", "用户", "时间", "date", "time",
    "created", "updated", "url", "link", "点赞", "like", "score",
)
_ID_LIKE_RE = re.compile(r"^[A-Za-z0-9_\-:.]{6,}$")
_TIME_LIKE_RE = re.compile(r"^\d{4}[-/]\d{1,2}[-/]\d{1,2}|^\d{10,13}$")


def _is_id_or_time_like(value: str) -> bool:
    text = str(value or "").strip()
    if not text:
        return False
    if _TIME_LIKE_RE.search(text):
        return True
    if _ID_LIKE_RE.match(text) and not any(ch.isspace() for ch in text) and len(text) <= 40:
        return True
    return False


def _comment_column_score(header_name: str, values: list[str]) -> float:
    nonempty = [str(v or "").strip() for v in values if str(v or "").strip()]
    if not nonempty:
        return -1.0
    sample_n = len(nonempty)
    lengths = [len(v) for v in nonempty]
    avg_len = sum(lengths) / sample_n
    long_ratio = sum(1 for n in lengths if n >= MIN_LEN) / sample_n
    letter_ratio = sum(1 for v in nonempty if has_any_letter(v)) / sample_n
    unique_ratio = len({v.lower() for v in nonempty}) / sample_n
    id_like_ratio = sum(1 for v in nonempty if _is_id_or_time_like(v)) / sample_n
    header_low = str(header_name or "").strip().lower()
    header_penalty = 18.0 if any(k in header_low for k in _ID_OR_TIME_HEADER_KEYS) else 0.0
    return (
        min(avg_len, 120) * 0.35
        + long_ratio * 45
        + letter_ratio * 35
        + unique_ratio * 10
        - id_like_ratio * 55
        - header_penalty
    )


def _detect_comment_column(header: list[str], sample_rows: list[list]) -> tuple[int, str, dict]:
    keyword = _keyword_comment_column(header)
    if keyword:
        idx, key = keyword
        return idx, str(header[idx]).strip() if idx < len(header) else "", {
            "method": "header_keyword",
            "keyword": key,
            "confidence": "high",
        }

    width = max([len(header), *(len(r) for r in sample_rows)] or [0])
    best_idx = 0
    best_score = -1.0
    scores: list[dict] = []
    for idx in range(width):
        values = [str(r[idx]) if idx < len(r) and r[idx] is not None else "" for r in sample_rows]
        name = str(header[idx]).strip() if idx < len(header) else ""
        score = _comment_column_score(name, values)
        scores.append({"index": idx + 1, "name": name, "score": round(score, 2)})
        if score > best_score:
            best_score = score
            best_idx = idx

    col_name = str(header[best_idx]).strip() if best_idx < len(header) else ""
    confidence = "medium" if best_score >= 35 else "low"
    return best_idx, col_name, {
        "method": "content_score",
        "confidence": confidence,
        "score": round(best_score, 2),
        "candidates": sorted(scores, key=lambda x: x["score"], reverse=True)[:3],
    }


def _parse_csv_comments(content: bytes) -> tuple[list[str], str]:
    """解析 CSV，并且只抽取评论列，避免大文件上传时构造整张表。"""
    text = _decode_csv(content)
    try:
        reader = csv.reader(io.StringIO(text))
        header = next(reader, None)
        if not header:
            raise ValueError("文件为空")
        col_idx = _locate_comment_column(header)
        col_name = str(header[col_idx]).strip() if col_idx < len(header) else ""

        comments: list[str] = []
        for row in reader:
            if col_idx >= len(row):
                continue
            cell = str(row[col_idx]).strip()
            if cell:
                comments.append(cell)
    except csv.Error as e:
        raise ValueError(f"无法解析 CSV 文件：{e}") from e

    if not comments:
        raise ValueError("文件只有表头或评论列为空")
    return comments, col_name


def _parse_excel_comments(content: bytes) -> tuple[list[str], str]:
    """解析 Excel 第一张表，并且只读取评论列。

    原实现会把整张工作表所有列都读入内存。评论数据常带大量无关字段，
    大文件上传时会让 /upload 看起来卡死。这里先读表头定位列，再用
    min_col/max_col 只流式读取该列。
    """
    print("[comment-parse] excel load_workbook start", flush=True)
    try:
        wb = openpyxl.load_workbook(
            io.BytesIO(content),
            read_only=True,
            data_only=True,
            keep_links=False,
        )
    except (InvalidFileException, BadZipFile, OSError) as e:
        raise ValueError("无法解析 Excel 文件；如为 .xls 旧格式，请另存为 .xlsx 或 CSV 后再上传") from e
    print("[comment-parse] excel load_workbook done", flush=True)

    try:
        ws = wb.active
        print(f"[comment-parse] active sheet={ws.title!r} max_row={ws.max_row} max_col={ws.max_column}", flush=True)
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            raise ValueError("文件为空")
        header = [("" if c is None else str(c)) for c in header_row]
        col_idx = _locate_comment_column(header)
        col_name = str(header[col_idx]).strip() if col_idx < len(header) else ""
        print(f"[comment-parse] comment column index={col_idx + 1} name={col_name!r}", flush=True)

        comments: list[str] = []
        col_no = col_idx + 1
        max_row = ws.max_row or EXCEL_MAX_SCAN_ROWS
        scan_to = min(max_row, EXCEL_MAX_SCAN_ROWS)
        blank_streak = 0
        for row_no, (cell,) in enumerate(
            ws.iter_rows(min_row=2, max_row=scan_to, min_col=col_no, max_col=col_no, values_only=True),
            start=2,
        ):
            text = ("" if cell is None else str(cell)).strip()
            if text:
                comments.append(text)
                blank_streak = 0
            else:
                blank_streak += 1
                if comments and blank_streak >= EXCEL_STOP_AFTER_BLANK_ROWS:
                    print(
                        f"[comment-parse] stop after {blank_streak} blank rows at row={row_no}",
                        flush=True,
                    )
                    break
            if row_no % 5000 == 0:
                print(f"[comment-parse] scanned row={row_no} comments={len(comments)}", flush=True)
        if max_row > EXCEL_MAX_SCAN_ROWS:
            print(
                f"[comment-parse] sheet max_row={max_row} exceeds scan cap={EXCEL_MAX_SCAN_ROWS}; stopped at cap",
                flush=True,
            )
    finally:
        wb.close()

    if not comments:
        raise ValueError("文件只有表头或评论列为空")
    return comments, col_name


def parse_comment_file(file_bytes: bytes, filename: str) -> tuple[list[str], str]:
    """解析上传的评论文件，返回 (评论文本列表, 命中的列名)。

    仅支持 .csv / .xlsx。自动定位评论列，提取文本并去首尾空白，
    跳过空单元格。不在此处做清洗（清洗交给 filter_comments）。
    """
    name_low = (filename or "").lower()
    if name_low.endswith(".csv"):
        return _parse_csv_comments(file_bytes)
    elif name_low.endswith(".xlsx"):
        return _parse_excel_comments(file_bytes)
    else:
        raise ValueError("仅支持 CSV / Excel（.csv / .xlsx）文件")


# ── 规则清洗 ──────────────────────────────────────────────

def has_any_letter(text: str) -> bool:
    """是否包含任意语言的字母字符。

    依据 Unicode category 首字母为 'L'：
    - Lo 覆盖泰语/阿拉伯语/CJK 等非拉丁字母
    - Ll/Lu/Lt/Lm 覆盖拉丁、西里尔等大小写字母
    数字(N*)、标点(P*)、符号(S*)、空白(Z*) 均不计入，从而避免误删
    泰语、阿拉伯语、越南语、西里尔语等多语种评论。
    """
    return any(unicodedata.category(c).startswith("L") for c in text)


def _is_keyboard_mash(text: str) -> bool:
    """键盘乱打：长度 > KB_MASH_MINLEN 且单个高频字符占比 > KB_MASH_RATIO。"""
    stripped = text.strip()
    if len(stripped) <= KB_MASH_MINLEN:
        return False
    counts: dict[str, int] = {}
    for c in stripped:
        if c.isspace():
            continue
        counts[c] = counts.get(c, 0) + 1
    non_space_len = sum(counts.values())
    if non_space_len == 0:
        return False
    top = max(counts.values())
    return top / non_space_len > KB_MASH_RATIO


def _is_reward_only(text: str) -> bool:
    """整条评论只在索要奖励（钻石/皮肤/英雄等），无其他实质意见。"""
    stripped = text.strip()
    if not stripped:
        return False
    if _REWARD_CN_RE.match(stripped):
        return True
    if _REWARD_CN_LOOSE_RE.match(stripped):
        return True

    tokens = [t.lower() for t in _REWARD_TOKEN_RE.findall(stripped)]
    if not tokens or len(tokens) > 24:
        return False
    has_target = any(t in _REWARD_TARGET_WORDS for t in tokens)
    if not has_target:
        return False
    if any(t in _REWARD_VALUE_WORDS for t in tokens):
        return False
    if not any(t in _REWARD_REQUEST_WORDS for t in tokens):
        return False
    allowed = _REWARD_FILLER_WORDS | _REWARD_TARGET_WORDS
    unknown = [t for t in tokens if t not in allowed]
    return len(unknown) <= 2


def filter_comments(comments: list[str]) -> tuple[list[str], dict[str, int]]:
    """规则清洗。任一规则命中即丢弃；返回 (保留列表, 各规则丢弃计数)。

    规则（按命中即丢弃）：
    - too_short：strip 后长度 < MIN_LEN
    - keyboard_mash：单字符高频占比超阈值（"aaaaa..."、"asdfgh..."）
    - no_letter：不含任意语言字母字符（"123123"、"??!!!"）
    - duplicate：大小写不敏感完全去重，保留第一条
    - reward_only：整条只索要奖励

    去重在最后做，保证"保留第一条"指向通过前几条规则的首条。
    """
    stats = {
        "too_short": 0,
        "keyboard_mash": 0,
        "no_letter": 0,
        "reward_only": 0,
        "duplicate": 0,
    }
    seen: set[str] = set()
    kept: list[str] = []
    for idx, raw in enumerate(comments, start=1):
        text = raw.strip()
        if idx % 5000 == 0:
            print(f"[comment-filter] processed={idx} kept={len(kept)}", flush=True)
        if len(text) < MIN_LEN:
            stats["too_short"] += 1
            continue
        if _is_keyboard_mash(text):
            stats["keyboard_mash"] += 1
            continue
        if not has_any_letter(text):
            stats["no_letter"] += 1
            continue
        if _is_reward_only(text):
            stats["reward_only"] += 1
            continue
        key = text.lower()
        if key in seen:
            stats["duplicate"] += 1
            continue
        seen.add(key)
        kept.append(text)
    print(f"[comment-filter] complete processed={len(comments)} kept={len(kept)}", flush=True)
    return kept, stats


class StreamingCommentSampler:
    """边清洗边按长度分层做 reservoir sample，避免保存全量评论。"""

    def __init__(self, max_n: int = SAMPLE_MAX_N, seed: int | None = 42):
        self.max_n = max_n
        self.rng = random.Random(seed)
        self.stats = {
            "too_short": 0,
            "keyboard_mash": 0,
            "no_letter": 0,
            "reward_only": 0,
            "duplicate": 0,
        }
        self.seen_hashes: set[str] = set()
        self.valid_count = 0
        self._long_counter = 0
        self._long_heap: list[tuple[int, int, str]] = []
        self.buckets: dict[str, list[str]] = {name: [] for name, *_ in STRATA}
        self.bucket_seen: dict[str, int] = {name: 0 for name, *_ in STRATA}

    def add(self, raw: str) -> bool:
        text = str(raw or "").strip()
        if len(text) < MIN_LEN:
            self.stats["too_short"] += 1
            return False
        if _is_keyboard_mash(text):
            self.stats["keyboard_mash"] += 1
            return False
        if not has_any_letter(text):
            self.stats["no_letter"] += 1
            return False
        if _is_reward_only(text):
            self.stats["reward_only"] += 1
            return False

        key = hashlib.sha1(text.lower().encode("utf-8", errors="ignore")).hexdigest()
        if key in self.seen_hashes:
            self.stats["duplicate"] += 1
            return False
        self.seen_hashes.add(key)

        self.valid_count += 1
        self._add_long_candidate(text)
        name = _stratum_of(text)
        self.bucket_seen[name] += 1
        pool = self.buckets[name]
        # 每个分层最多保留 max_n 条候选，最终再按 STRATA 比例合并到 max_n。
        if len(pool) < self.max_n:
            pool.append(text)
        else:
            j = self.rng.randrange(self.bucket_seen[name])
            if j < self.max_n:
                pool[j] = text
        return True

    def _add_long_candidate(self, text: str) -> None:
        self._long_counter += 1
        item = (len(text), self._long_counter, text)
        if len(self._long_heap) < LONG_CANDIDATE_MAX_N:
            heapq.heappush(self._long_heap, item)
        elif item[0] > self._long_heap[0][0]:
            heapq.heapreplace(self._long_heap, item)

    def sample(self) -> list[str]:
        candidates: list[str] = []
        for name, *_ in STRATA:
            candidates.extend(self.buckets[name])
        return stratified_sample(candidates, self.max_n)

    def long_candidates(self) -> list[str]:
        return [
            text for _length, _counter, text in sorted(
                self._long_heap,
                key=lambda item: (item[0], item[1]),
                reverse=True,
            )
        ]


def _preprocess_progress(
    message: str,
    scan_rows: int,
    nonempty_count: int,
    sampler: StreamingCommentSampler,
    *,
    capped: bool = False,
    warning: str = "",
) -> dict:
    return {
        "kind": "progress",
        "message": message,
        "scan_rows": scan_rows,
        "nonempty_count": nonempty_count,
        "valid_count": sampler.valid_count,
        "sample_count": min(sampler.valid_count, sampler.max_n),
        "scan_capped": capped,
        "warning": warning,
    }


def _preprocess_done(
    filename: str,
    col_idx: int,
    col_name: str,
    detection: dict,
    scan_rows: int,
    nonempty_count: int,
    sampler: StreamingCommentSampler,
    *,
    capped: bool,
    warning: str = "",
) -> dict:
    sample_pool = sampler.sample()
    if not sample_pool:
        raise ValueError("清洗后没有有效评论，请检查文件内容")
    initial_sample = sample_pool[:SAMPLE_MAX_N]
    long_candidates = sampler.long_candidates()
    return {
        "kind": "done",
        "result": {
            "filename": filename,
            "comment_column": col_name,
            "comment_column_index": col_idx + 1,
            "comment_column_detection": detection,
            "scan_rows": scan_rows,
            "nonempty_count": nonempty_count,
            "valid_count": sampler.valid_count,
            "sample_count": len(initial_sample),
            "sample_pool_count": len(sample_pool),
            "sample_pool_max_n": sampler.max_n,
            "initial_sample_max_n": SAMPLE_MAX_N,
            "scan_capped": capped,
            "warning": warning,
            "filter_stats": sampler.stats,
            "sample": sample_pool,
            "long_candidates": long_candidates,
            "long_candidate_count": len(long_candidates),
            "long_candidate_max_n": LONG_CANDIDATE_MAX_N,
            "preview_comments": initial_sample[:5],
        },
    }


def _decode_csv_sample(raw: bytes) -> str:
    for enc in ("utf-8-sig", "utf-8", "gbk", "gb2312"):
        try:
            raw.decode(enc)
            return enc
        except UnicodeDecodeError:
            continue
    raise ValueError("无法解析 CSV 文件，请确认文件编码为 UTF-8 或 GBK")


def _normalize_row(row) -> list[str]:
    return ["" if c is None else str(c) for c in row]


def _feed_preprocess_row(
    row: list,
    col_idx: int,
    sampler: StreamingCommentSampler,
) -> bool:
    if col_idx >= len(row):
        return False
    text = str(row[col_idx] or "").strip()
    if not text:
        return False
    sampler.add(text)
    return True


def _preprocess_csv_file(path: str, filename: str, max_scan_rows: int):
    with open(path, "rb") as bf:
        enc = _decode_csv_sample(bf.read(65536))
    yield {"kind": "progress", "message": "正在读取 CSV 表头与样本行…"}

    sampler = StreamingCommentSampler(max_n=SAMPLE_POOL_MAX_N)
    scan_rows = 0
    nonempty_count = 0
    capped = False
    warning = ""
    # 社媒导出的 CSV 偶尔会在整体 UTF-8 文件中混入少量脏字节。
    # 这里用 replacement character 跳过坏字节，避免单条脏评论导致整批预处理失败。
    with open(path, "r", encoding=enc, errors="replace", newline="") as f:
        reader = csv.reader(f)
        header = next(reader, None)
        if not header:
            raise ValueError("文件为空")
        sample_rows: list[list[str]] = []
        sample_limit = min(COLUMN_SAMPLE_ROWS, max_scan_rows)
        for row in reader:
            sample_rows.append(_normalize_row(row))
            if len(sample_rows) >= sample_limit:
                break
        col_idx, col_name, detection = _detect_comment_column(_normalize_row(header), sample_rows)
        yield {
            "kind": "progress",
            "message": f"已识别评论列：{col_name or ('第 ' + str(col_idx + 1) + ' 列')}",
            "column_index": col_idx + 1,
            "column_name": col_name,
            "column_detection": detection,
        }

        for row in sample_rows:
            scan_rows += 1
            if _feed_preprocess_row(row, col_idx, sampler):
                nonempty_count += 1

        for row in reader:
            if scan_rows >= max_scan_rows:
                capped = True
                warning = f"文件超过 {max_scan_rows} 行，本次仅扫描前 {max_scan_rows} 行评论"
                break
            scan_rows += 1
            if _feed_preprocess_row(_normalize_row(row), col_idx, sampler):
                nonempty_count += 1
            if scan_rows % PREPROCESS_PROGRESS_ROWS == 0:
                yield _preprocess_progress(
                    f"已扫描 {scan_rows} 行，保留有效评论 {sampler.valid_count} 条",
                    scan_rows,
                    nonempty_count,
                    sampler,
                    capped=capped,
                    warning=warning,
                )

    yield _preprocess_done(
        filename, col_idx, col_name, detection, scan_rows, nonempty_count, sampler,
        capped=capped, warning=warning,
    )


def _preprocess_excel_file(path: str, filename: str, max_scan_rows: int):
    yield {"kind": "progress", "message": "正在打开 Excel 文件…"}
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_links=False)
    except (InvalidFileException, BadZipFile, OSError) as e:
        raise ValueError("无法解析 Excel 文件；如为 .xls 旧格式，请另存为 .xlsx 或 CSV 后再上传") from e

    sampler = StreamingCommentSampler(max_n=SAMPLE_POOL_MAX_N)
    scan_rows = 0
    nonempty_count = 0
    capped = False
    warning = ""
    try:
        ws = wb.active
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            raise ValueError("文件为空")
        header = _normalize_row(header_row)
        row_iter = ws.iter_rows(min_row=2, values_only=True)
        sample_rows: list[list[str]] = []
        sample_limit = min(COLUMN_SAMPLE_ROWS, max_scan_rows)
        for row in row_iter:
            sample_rows.append(_normalize_row(row))
            if len(sample_rows) >= sample_limit:
                break
        col_idx, col_name, detection = _detect_comment_column(header, sample_rows)
        yield {
            "kind": "progress",
            "message": f"已识别评论列：{col_name or ('第 ' + str(col_idx + 1) + ' 列')}",
            "column_index": col_idx + 1,
            "column_name": col_name,
            "column_detection": detection,
        }

        blank_streak = 0
        for row in sample_rows:
            scan_rows += 1
            text = str(row[col_idx] if col_idx < len(row) else "").strip()
            if text:
                blank_streak = 0
                nonempty_count += 1
                sampler.add(text)
            else:
                blank_streak += 1

        for row in row_iter:
            if scan_rows >= max_scan_rows:
                capped = True
                warning = f"文件超过 {max_scan_rows} 行，本次仅扫描前 {max_scan_rows} 行评论"
                break
            scan_rows += 1
            row_values = _normalize_row(row)
            text = str(row_values[col_idx] if col_idx < len(row_values) else "").strip()
            if text:
                blank_streak = 0
                nonempty_count += 1
                sampler.add(text)
            else:
                blank_streak += 1
                if nonempty_count and blank_streak >= EXCEL_STOP_AFTER_BLANK_ROWS:
                    warning = f"连续 {EXCEL_STOP_AFTER_BLANK_ROWS} 个空行，已提前停止扫描"
                    break
            if scan_rows % PREPROCESS_PROGRESS_ROWS == 0:
                yield _preprocess_progress(
                    f"已扫描 {scan_rows} 行，保留有效评论 {sampler.valid_count} 条",
                    scan_rows,
                    nonempty_count,
                    sampler,
                    capped=capped,
                    warning=warning,
                )
    finally:
        wb.close()

    yield _preprocess_done(
        filename, col_idx, col_name, detection, scan_rows, nonempty_count, sampler,
        capped=capped, warning=warning,
    )


def preprocess_comment_file(path: str, filename: str, max_scan_rows: int = COMMENT_MAX_SCAN_ROWS):
    """流式预处理评论文件，yield progress/done 事件供 SSE 转发。"""
    name_low = (filename or "").lower()
    if name_low.endswith(".csv"):
        yield from _preprocess_csv_file(path, filename, max_scan_rows)
    elif name_low.endswith(".xlsx"):
        yield from _preprocess_excel_file(path, filename, max_scan_rows)
    else:
        raise ValueError("仅支持 CSV / Excel（.csv / .xlsx）文件")


# ── 分层随机抽样 ──────────────────────────────────────────

def _stratum_of(text: str) -> str:
    n = len(text)
    for name, lo, hi, _ratio in STRATA:
        if n >= lo and (hi is None or n <= hi):
            return name
    # 长度 < 最小边界的（理论上已被 MIN_LEN 过滤）归入 short
    return STRATA[0][0]


def stratified_sample(
    valid_comments: list[str],
    max_n: int = SAMPLE_MAX_N,
    seed: int | None = 42,
) -> list[str]:
    """按长度分层随机抽样至多 max_n 条。

    - 有效评论数 <= max_n：全量返回（不抽样，保持原顺序）。
    - 否则按 STRATA 比例分配名额，桶内随机抽取；某桶不足时把剩余名额
      按比例重分配给其他桶，尽量取满 max_n。
    """
    if len(valid_comments) <= max_n:
        return list(valid_comments)

    rng = random.Random(seed)
    buckets: dict[str, list[str]] = {name: [] for name, *_ in STRATA}
    for text in valid_comments:
        buckets[_stratum_of(text)].append(text)

    # 初始名额
    targets: dict[str, int] = {
        name: int(round(max_n * ratio)) for name, _lo, _hi, ratio in STRATA
    }
    # 限制在桶容量内，记录缺额
    picked: dict[str, int] = {}
    deficit = 0
    for name, *_ in STRATA:
        want = targets[name]
        have = len(buckets[name])
        take = min(want, have)
        picked[name] = take
        deficit += want - take

    # 把缺额分配给仍有余量的桶
    if deficit > 0:
        for name, *_ in STRATA:
            if deficit <= 0:
                break
            spare = len(buckets[name]) - picked[name]
            if spare > 0:
                add = min(spare, deficit)
                picked[name] += add
                deficit -= add

    result: list[str] = []
    for name, *_ in STRATA:
        pool = buckets[name]
        k = picked[name]
        if k >= len(pool):
            result.extend(pool)
        else:
            result.extend(rng.sample(pool, k))
    # 防御：四舍五入可能略超 max_n
    if len(result) > max_n:
        result = rng.sample(result, max_n)
    return result


def make_batches(comments: list[str], batch_size: int = BATCH_SIZE) -> list[list[str]]:
    """把评论切分为若干批，供并发调用 Dify。"""
    return [comments[i:i + batch_size] for i in range(0, len(comments), batch_size)]


# ── 本地统计聚合 ──────────────────────────────────────────

def format_quote(translation: str, original: str) -> str:
    """代表性引用格式化为「中文翻译」（原文: ...）。

    翻译缺失时退化为只显示原文，避免出现空书名号。
    """
    original = (original or "").strip()
    translation = (translation or "").strip()
    if translation and translation != original:
        return f"「{translation}」（原文: {original}）"
    return f"「{original}」"


def aggregate(
    final_themes: list[dict],
    classifications: list[dict],
    other_pct: float = OTHER_THEME_PCT,
) -> dict:
    """对分类结果做本地统计：占比、情感分布、代表引用、低频合并入「其他」。

    多标签：一条评论可命中多个主题（theme_ids），分别计入各主题。

    入参：
    - final_themes: 合并去重后的主题列表，每项至少含 id / name / description。
    - classifications: 每条评论一条记录（顺序与原评论一致），形如：
        {
          "theme_ids": [str, ...],  # 命中的主题列表；不在 final_themes 中的归入 other
          "sentiment": "positive" | "negative" | "neutral" | "mixed",  # 评论整体情感
          "original": str,          # 原文评论
          "translation": str,       # 中文翻译（可空）
          "is_quote_candidate": bool,
        }
      兼容旧单标签字段 "theme_id"（若无 theme_ids 则按单主题处理）。

    返回：
    {
      "total_classified": int,      # 评论总数
      "sentiment_overall": {"positive_pct", "neutral_pct", "negative_pct"},  # 按评论计，和≈100%
      "themes": [ {id,name,description,count,percentage,
                   positive_count/pct, negative_count/pct, neutral_count/pct,
                   quotes:[...]} ],   # 占比 = 提及该主题的评论数/总评论数，多标签下各主题之和可能 >100%
      "other_themes": [ {name,count,percentage} ],  # 占比 < other_pct
    }
    """
    theme_ids = {t["id"] for t in final_themes}
    counts: dict[str, dict] = {
        t["id"]: {"total": 0, "positive": 0, "negative": 0, "neutral": 0, "mixed": 0}
        for t in final_themes
    }
    counts["other"] = {"total": 0, "positive": 0, "negative": 0, "neutral": 0, "mixed": 0}
    # 每主题、每情感的候选引用池
    quotes_pool: dict[str, dict[str, list[tuple[str, str]]]] = {
        t["id"]: {"positive": [], "negative": [], "neutral": []} for t in final_themes
    }

    overall = {"positive": 0, "negative": 0, "neutral": 0, "mixed": 0}
    comment_count = 0

    for item in classifications:
        comment_count += 1
        sentiment = item.get("sentiment", "neutral")
        if sentiment not in ("positive", "negative", "neutral", "mixed"):
            sentiment = "neutral"
        overall[sentiment] += 1  # 整体情感按"每条评论"只计一次

        # 多标签：取 theme_ids 列表；兼容旧单标签 theme_id
        tids = item.get("theme_ids")
        if not isinstance(tids, list) or not tids:
            single = item.get("theme_id")
            tids = [single] if single else ["other"]

        seen_in_comment: set[str] = set()
        for raw_tid in tids:
            tid = str(raw_tid) if raw_tid else "other"
            if tid not in theme_ids:
                tid = "other"
            if tid in seen_in_comment:
                continue  # 同一评论对同一主题不重复计数
            seen_in_comment.add(tid)

            counts[tid]["total"] += 1
            counts[tid][sentiment] += 1

            if (
                tid != "other"
                and item.get("is_quote_candidate")
                and item.get("original")
            ):
                bucket = sentiment if sentiment in ("positive", "negative") else "neutral"
                pool = quotes_pool[tid][bucket]
                if len(pool) < 5:
                    pool.append((item.get("translation", ""), item.get("original", "")))

    total_classified = comment_count

    # 整体情感分布（按评论计，mixed 归入中性侧，和≈100%）
    overall_base = comment_count or 1
    sentiment_overall = {
        "positive_pct": round(overall["positive"] / overall_base * 100, 1),
        "negative_pct": round(overall["negative"] / overall_base * 100, 1),
        "neutral_pct": round(
            (overall["neutral"] + overall["mixed"]) / overall_base * 100, 1
        ),
    }

    themes_out: list[dict] = []
    other_themes_out: list[dict] = []
    # 主题占比基数 = 评论总数（多标签下各主题占比之和可能 >100%）
    base = comment_count or 1

    for t in final_themes:
        tid = t["id"]
        c = counts[tid]
        cnt = c["total"]
        pct = round(cnt / base * 100, 1)
        denom = cnt or 1

        if pct < other_pct:
            other_themes_out.append({"name": t["name"], "count": cnt, "percentage": pct})
            continue

        pool = quotes_pool[tid]
        # 每种情感各取 1 条代表，凑 2~3 条
        quotes: list[str] = []
        for bucket in ("positive", "negative", "neutral"):
            for translation, original in pool[bucket][:1]:
                quotes.append(format_quote(translation, original))
        # 不足 2 条时再补
        if len(quotes) < 2:
            for bucket in ("positive", "negative", "neutral"):
                for translation, original in pool[bucket][1:]:
                    quotes.append(format_quote(translation, original))
                    if len(quotes) >= 3:
                        break
                if len(quotes) >= 3:
                    break
        quotes = quotes[:3]

        themes_out.append({
            "id": tid,
            "name": t["name"],
            "description": t.get("description", ""),
            "count": cnt,
            "percentage": pct,
            "positive_count": c["positive"],
            "positive_pct": round(c["positive"] / denom * 100, 1),
            "negative_count": c["negative"],
            "negative_pct": round(c["negative"] / denom * 100, 1),
            "neutral_count": c["neutral"] + c["mixed"],
            "neutral_pct": round((c["neutral"] + c["mixed"]) / denom * 100, 1),
            "quotes": quotes,
        })

    themes_out.sort(key=lambda x: x["count"], reverse=True)
    other_themes_out.sort(key=lambda x: x["count"], reverse=True)

    return {
        "total_classified": total_classified,
        "sentiment_overall": sentiment_overall,
        "themes": themes_out,
        "other_themes": other_themes_out,
    }
