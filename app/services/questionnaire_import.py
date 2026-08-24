"""定性问卷来源适配：解析倍市得原问卷并与可读回答数据确定性对齐。"""

from __future__ import annotations

import io
import json
import re
from dataclasses import dataclass, field

import openpyxl


_QUESTION_RE = re.compile(r"^Q(\d+)\[([^\]]+)\]$")
_CODE_QUESTION_RE = re.compile(r"^Q(\d+)\.(.*)$", re.DOTALL)
_CODE_OPTION_RE = re.compile(r"^(\d+)\.(.*)$", re.DOTALL)
_CODED_DATA_HEADER_RE = re.compile(
    r"^Q(\d+)(?:__(?:\d+|open))?$",
    re.IGNORECASE,
)
_CONTACT_RE = re.compile(r"手机号|手机号码|联系电话|联系方式", re.IGNORECASE)
_WHATSAPP_CONTACT_RE = re.compile(
    r"^\s*(?:"
    r"(?:请(?:填写|输入|留下|提供)\s*(?:您的?)?)|"
    r"(?:(?:please\s+)?(?:provide|enter|leave|fill\s+in)\s+(?:your\s+)?)|"
    r"(?:your\s+)"
    r")?"
    r"whatsapp"
    r"(?:\s*(?:账号|帐号|号码|联系方式|id|number|account|contact))?"
    r"\s*[？?:：]?\s*$",
    re.IGNORECASE,
)
_JSON_FENCE_RE = re.compile(r"```(?:json)?\s*(.*?)```", re.IGNORECASE | re.DOTALL)
_CHINESE_RE = re.compile(r"[\u4e00-\u9fff]")
_LATIN_RE = re.compile(r"[A-Za-z]")
_LATIN_WORD_RE = re.compile(r"[A-Za-z]+")

_BESTED_ROLE_MAP = {
    "描述题": "ignore",
    "单选题": "single_choice",
    "多选题": "multi_choice",
    "矩阵单选题": "matrix_single",
    "矩阵多选题": "matrix_multi",
    "矩阵打分题": "matrix_scale",
    "矩阵量表题": "matrix_scale",
    "量表题": "scale",
    "打分题": "scale",
    "填空题": "open_text",
}


@dataclass
class _BestedQuestion:
    qid: int
    source_type: str
    role: str
    title: str
    options: list[str] = field(default_factory=list)
    option_codes: list[int] = field(default_factory=list)
    rows: list[str] = field(default_factory=list)
    row_codes: list[int] = field(default_factory=list)


@dataclass
class _BestedCodeQuestion:
    qid: int
    title: str
    options: list[str] = field(default_factory=list)
    option_codes: list[int] = field(default_factory=list)
    rows: list[str] = field(default_factory=list)
    row_codes: list[int] = field(default_factory=list)


def _cell_text(value) -> str:
    return "" if value is None else str(value).strip()


def _load_workbook(content: bytes):
    try:
        return openpyxl.load_workbook(
            io.BytesIO(content),
            read_only=True,
            data_only=True,
        )
    except Exception as exc:
        raise ValueError("无法读取 Excel 内容，请确认文件来自倍市得且未损坏") from exc


def _worksheet_rows(ws) -> list[list[str]]:
    rows: list[list[str]] = []
    for row in ws.iter_rows(values_only=True):
        rows.append([_cell_text(value) for value in row])
    while rows and not any(cell for cell in rows[-1]):
        rows.pop()
    return rows


def _questionnaire_sheet(workbook):
    if "问卷内容" in workbook.sheetnames:
        return workbook["问卷内容"]
    return workbook[workbook.sheetnames[0]]


def _parse_bested_questionnaire(content: bytes) -> tuple[list[_BestedQuestion], str]:
    workbook = _load_workbook(content)
    try:
        rows = _worksheet_rows(_questionnaire_sheet(workbook))
    finally:
        workbook.close()
    if len(rows) <= 1:
        raise ValueError("调研问卷为空或缺少题目")

    questions: list[_BestedQuestion] = []
    current: _BestedQuestion | None = None
    section = ""
    for row in rows[1:]:
        first = row[0] if row else ""
        second = row[1] if len(row) > 1 else ""
        match = _QUESTION_RE.match(first)
        if match:
            raw_type = match.group(2).strip()
            role = _BESTED_ROLE_MAP.get(raw_type)
            if not role:
                raise ValueError(f"暂不支持 Q{match.group(1)} 的题型「{raw_type}」")
            if not second:
                raise ValueError(f"Q{match.group(1)} 缺少题干")
            current = _BestedQuestion(
                qid=int(match.group(1)),
                source_type=raw_type,
                role=role,
                title=second,
            )
            questions.append(current)
            section = ""
            continue
        if not current:
            continue
        if first == "选项":
            section = "options"
            continue
        if first == "矩阵行":
            section = "rows"
            continue
        if first.isdigit() and second:
            if section == "options":
                current.option_codes.append(int(first))
                current.options.append(second)
            elif section == "rows":
                current.row_codes.append(int(first))
                current.rows.append(second)

    if not questions:
        raise ValueError("未识别到 Q号[题型] 格式的题目")
    seen: set[int] = set()
    for question in questions:
        if question.qid in seen:
            raise ValueError(f"原问卷中 Q{question.qid} 重复")
        seen.add(question.qid)
        if question.role in {"single_choice", "multi_choice", "matrix_single", "matrix_multi"} \
                and not question.options:
            raise ValueError(f"Q{question.qid} 缺少选项")
        if question.role.startswith("matrix_") and not question.rows:
            raise ValueError(f"Q{question.qid} 缺少矩阵行")
        if len(question.option_codes) != len(set(question.option_codes)):
            raise ValueError(f"原问卷中 Q{question.qid} 存在重复选项编码")
        if len(question.row_codes) != len(set(question.row_codes)):
            raise ValueError(f"原问卷中 Q{question.qid} 存在重复矩阵行编码")

    questionnaire_text = "\n".join(
        " | ".join(cell for cell in row if cell)
        for row in rows
        if any(row)
    )
    return questions, questionnaire_text


def _parse_response_workbook(
    content: bytes,
) -> tuple[list[list[str]], list[_BestedCodeQuestion]]:
    workbook = _load_workbook(content)
    try:
        if "data" not in workbook.sheetnames or "code" not in workbook.sheetnames:
            raise ValueError("回答文件必须同时包含 data 和 code 工作表")
        data_rows = _worksheet_rows(workbook["data"])
        code_rows = _worksheet_rows(workbook["code"])
    finally:
        workbook.close()
    if len(data_rows) <= 1:
        raise ValueError("data 工作表为空或只有表头")

    code_questions: list[_BestedCodeQuestion] = []
    current: _BestedCodeQuestion | None = None
    section = ""
    for row in code_rows:
        value = row[1] if len(row) > 1 else ""
        match = _CODE_QUESTION_RE.match(value)
        if match:
            current = _BestedCodeQuestion(
                qid=int(match.group(1)),
                title=match.group(2).strip(),
            )
            code_questions.append(current)
            section = ""
            continue
        if not current:
            continue
        marker = value.casefold()
        if marker in {"option", "选项"}:
            section = "options"
            continue
        if marker in {"subquestion", "sub question", "sub_question", "矩阵行"}:
            section = "rows"
            continue
        option_match = _CODE_OPTION_RE.match(value)
        if section == "options" and option_match:
            current.option_codes.append(int(option_match.group(1)))
            current.options.append(option_match.group(2).strip())
            continue
        if section == "rows" and option_match:
            current.row_codes.append(int(option_match.group(1)))
            current.rows.append(option_match.group(2).strip())
            continue
        if section in {"options", "rows"} and value:
            section = ""
    if not code_questions:
        raise ValueError("code 工作表中未识别到 Q号.题干")
    qids = [question.qid for question in code_questions]
    if len(qids) != len(set(qids)):
        raise ValueError("code 工作表中存在重复题号")
    for question in code_questions:
        if len(question.option_codes) != len(set(question.option_codes)):
            raise ValueError(f"code 工作表中 Q{question.qid} 存在重复选项编码")
        if len(question.row_codes) != len(set(question.row_codes)):
            raise ValueError(f"code 工作表中 Q{question.qid} 存在重复矩阵行编码")
    return data_rows, code_questions


def _norm(value: str) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip().casefold()


def _is_contact_field(title: str) -> bool:
    text = str(title or "").strip()
    return bool(_CONTACT_RE.search(text) or _WHATSAPP_CONTACT_RE.fullmatch(text))


def _is_chinese_dominant(text: str) -> bool:
    chinese_units = len(_CHINESE_RE.findall(str(text or "")))
    latin_units = len(_LATIN_WORD_RE.findall(str(text or "")))
    return chinese_units > 0 and chinese_units >= latin_units


def _find_exact_header(
    headers: list[str],
    title: str,
    used: set[int],
    cursor: int,
) -> int | None:
    wanted = _norm(title)
    for index in range(cursor, len(headers)):
        if index not in used and _norm(headers[index]) == wanted:
            return index
    matches = [
        index for index, header in enumerate(headers)
        if index not in used and _norm(header) == wanted
    ]
    return matches[0] if len(matches) == 1 else None


def _group_headers(
    headers: list[str],
    title: str,
    expected_labels: list[str],
    used: set[int],
) -> list[int]:
    prefix = f"{title}__"
    candidates = [
        index for index, header in enumerate(headers)
        if index not in used and header.startswith(prefix)
    ]
    suffix_to_index = {
        _norm(headers[index][len(prefix):]): index
        for index in candidates
    }
    indexes: list[int] = []
    missing: list[str] = []
    for label in expected_labels:
        index = suffix_to_index.get(_norm(label))
        if index is None:
            missing.append(label)
        else:
            indexes.append(index)
    if missing or len(indexes) != len(candidates):
        detail = "、".join(missing[:3]) if missing else "存在原问卷之外的拆分列"
        raise ValueError(f"题目「{title}」与回答列无法完整匹配：{detail}")
    return indexes


def _row_value(row: list[str], index: int) -> str:
    return row[index] if index < len(row) else ""


def _uses_coded_data_headers(
    headers: list[str],
    code_questions: list[_BestedCodeQuestion],
) -> bool:
    code_qids = {question.qid for question in code_questions}
    for header in headers:
        match = _CODED_DATA_HEADER_RE.match(str(header or "").strip())
        if match and int(match.group(1)) in code_qids:
            return True
    return False


def _validate_code_label_codes(
    question: _BestedQuestion,
    code_question: _BestedCodeQuestion,
) -> None:
    if not question.role.startswith("matrix_") and code_question.row_codes:
        raise ValueError(
            f"回答文件 Q{code_question.qid} 的矩阵行结构与原问卷题型不一致"
        )
    has_open_sentinel = (
        len(code_question.options) == 1
        and _norm(code_question.options[0]) in {"open", "开放", "填空"}
    )
    if question.role == "open_text" and code_question.options:
        if not has_open_sentinel:
            raise ValueError(
                f"回答文件 Q{code_question.qid} 的选项结构与原问卷填空题不一致"
            )
    elif has_open_sentinel:
        raise ValueError(
            f"回答文件 Q{code_question.qid} 的填空题标记与原问卷题型不一致"
        )
    if (
        question.role in {
            "single_choice", "multi_choice", "matrix_single", "matrix_multi",
        }
        and code_question.option_codes
        and set(code_question.option_codes) != set(question.option_codes)
    ):
        raise ValueError(
            f"原问卷 Q{question.qid} 与回答文件 Q{code_question.qid} 的选项编码不一致"
        )
    if (
        question.role.startswith("matrix_")
        and code_question.row_codes
        and set(code_question.row_codes) != set(question.row_codes)
    ):
        raise ValueError(
            f"原问卷 Q{question.qid} 与回答文件 Q{code_question.qid} 的矩阵行编码不一致"
        )


def _readable_question_pairs(
    source_questions: list[_BestedQuestion],
    code_questions: list[_BestedCodeQuestion],
) -> list[tuple[_BestedQuestion, _BestedCodeQuestion]]:
    source_by_qid = {question.qid: question for question in source_questions}
    questionnaire_is_chinese = _is_chinese_dominant("\n".join(
        question.title
        for question in source_questions
    ))
    pairs: list[tuple[_BestedQuestion, _BestedCodeQuestion]] = []
    has_cross_language_pair = False
    for code_question in code_questions:
        question = source_by_qid.get(code_question.qid)
        if question is None:
            raise ValueError(
                f"回答文件中的 Q{code_question.qid} 在原问卷中不存在"
            )
        if question.source_type == "描述题":
            continue
        _validate_code_label_codes(question, code_question)
        if _norm(question.title) != _norm(code_question.title):
            response_title_is_non_chinese = not _is_chinese_dominant(
                code_question.title
            )
            if not questionnaire_is_chinese or not response_title_is_non_chinese:
                raise ValueError(
                    f"Q{question.qid} 的题干在原问卷与回答文件中不一致"
                )
            has_cross_language_pair = True
            if question.role in {
                "single_choice", "multi_choice", "matrix_single", "matrix_multi",
            } and len(question.options) != len(code_question.options):
                raise ValueError(
                    f"Q{question.qid} 的中英文选项数量不一致"
                )
            if question.role.startswith("matrix_") \
                    and len(question.rows) != len(code_question.rows):
                raise ValueError(
                    f"Q{question.qid} 的中英文矩阵行数量不一致"
                )
        pairs.append((question, code_question))
    if has_cross_language_pair:
        source_qids = {
            question.qid
            for question in source_questions
            if question.source_type != "描述题"
        }
        response_qids = {
            question.qid
            for question in code_questions
            if not (
                question.qid in source_by_qid
                and source_by_qid[question.qid].source_type == "描述题"
            )
        }
        if source_qids != response_qids:
            raise ValueError("中英文问卷与回答文件的可回收题号集合不一致")
    return pairs


def _coded_question_pairs(
    source_questions: list[_BestedQuestion],
    code_questions: list[_BestedCodeQuestion],
) -> list[tuple[_BestedQuestion, _BestedCodeQuestion]]:
    source_by_qid = {question.qid: question for question in source_questions}
    analyzable = [
        question for question in source_questions
        if question.source_type != "描述题"
    ]
    response_questions = [
        question for question in code_questions
        if not (
            question.qid in source_by_qid
            and source_by_qid[question.qid].source_type == "描述题"
        )
    ]
    if len(analyzable) != len(response_questions):
        raise ValueError(
            "原问卷可回收题目数与回答文件 code 题目数不一致："
            f"{len(analyzable)} vs {len(response_questions)}"
        )

    source_qids = {question.qid for question in analyzable}
    response_qids = {question.qid for question in response_questions}
    mismatch_indexes = [
        index
        for index, (question, code_question) in enumerate(
            zip(analyzable, response_questions)
        )
        if question.qid != code_question.qid
    ]
    if len(mismatch_indexes) > 1:
        raise ValueError("原问卷与回答文件存在多处题号错位，无法唯一对应")
    if mismatch_indexes:
        mismatch_index = mismatch_indexes[0]
        has_two_neighbors = 0 < mismatch_index < len(analyzable) - 1
        previous_matches = (
            has_two_neighbors
            and analyzable[mismatch_index - 1].qid
            == response_questions[mismatch_index - 1].qid
        )
        next_matches = (
            has_two_neighbors
            and analyzable[mismatch_index + 1].qid
            == response_questions[mismatch_index + 1].qid
        )
        if not previous_matches or not next_matches:
            raise ValueError("原问卷与回答文件的题号错位缺少相邻题号校验")

    pairs: list[tuple[_BestedQuestion, _BestedCodeQuestion]] = []
    for question, code_question in zip(analyzable, response_questions):
        _validate_code_label_codes(question, code_question)
        if question.qid != code_question.qid:
            if (
                question.qid in response_qids
                or code_question.qid in source_qids
                or not question.options
                or len(question.options) != len(code_question.options)
            ):
                raise ValueError(
                    f"原问卷 Q{question.qid} 与回答文件 Q{code_question.qid} "
                    "无法按题序和选项数量唯一对应"
                )
        if (
            question.role in {
                "single_choice", "multi_choice", "matrix_single", "matrix_multi",
            }
            and len(question.options) != len(code_question.options)
        ):
            raise ValueError(
                f"原问卷 Q{question.qid} 与回答文件 Q{code_question.qid} "
                "的选项数量不一致"
            )
        pairs.append((question, code_question))
    return pairs


def _coded_question_label(
    question: _BestedQuestion,
    code_question: _BestedCodeQuestion,
) -> str:
    if question.qid == code_question.qid:
        return f"Q{question.qid}"
    return f"原问卷 Q{question.qid}（回答文件 Q{code_question.qid}）"


def _coded_split_indexes(
    headers: list[str],
    response_qid: int,
    expected_codes: list[int],
    used: set[int],
    label: str,
) -> list[int]:
    pattern = re.compile(rf"^Q{response_qid}__(\d+)$", re.IGNORECASE)
    by_code: dict[int, list[int]] = {}
    for index, header in enumerate(headers):
        if index in used:
            continue
        match = pattern.match(str(header or "").strip())
        if match:
            by_code.setdefault(int(match.group(1)), []).append(index)

    duplicates = [code for code, indexes in by_code.items() if len(indexes) != 1]
    expected = set(expected_codes)
    actual = set(by_code)
    if duplicates or actual != expected:
        missing = sorted(expected - actual)
        extra = sorted(actual - expected)
        details: list[str] = []
        if missing:
            details.append("缺少编码 " + "、".join(map(str, missing[:5])))
        if extra:
            details.append("多出编码 " + "、".join(map(str, extra[:5])))
        if duplicates:
            details.append("编码重复 " + "、".join(map(str, duplicates[:5])))
        raise ValueError(f"{label} 的拆分回答列无法完整匹配：{'；'.join(details)}")
    return [by_code[code][0] for code in expected_codes]


def _coded_answer_index(
    headers: list[str],
    response_qid: int,
    used: set[int],
    label: str,
    *,
    open_text: bool = False,
) -> int:
    names = [f"Q{response_qid}__open", f"Q{response_qid}"] \
        if open_text else [f"Q{response_qid}"]
    wanted = {_norm(name) for name in names}
    matches = [
        index for index, header in enumerate(headers)
        if index not in used and _norm(header) in wanted
    ]
    if len(matches) != 1:
        expected = " 或 ".join(names)
        raise ValueError(f"{label} 应有且仅有一个回答列：{expected}")
    return matches[0]


def _coded_binary_selected(value: str, label: str, row_number: int) -> bool:
    token = _norm(value)
    if token in {"", "0", "0.0", "false", "no", "否"}:
        return False
    if token in {"1", "1.0", "true", "yes", "是"}:
        return True
    raise ValueError(f"{label} 第 {row_number} 行存在非 0/1 的多选编码「{value}」")


def _decode_coded_choice(
    value: str,
    question: _BestedQuestion,
    code_question: _BestedCodeQuestion,
    row_number: int,
) -> str:
    text = str(value or "").strip()
    if not text:
        return ""

    source_by_value = {_norm(option): option for option in question.options}
    if _norm(text) in source_by_value:
        return source_by_value[_norm(text)]
    code_by_value = {
        _norm(option): code
        for option, code in zip(
            code_question.options, code_question.option_codes,
        )
    }
    if _norm(text) in code_by_value:
        option_code = code_by_value[_norm(text)]
        return question.options[question.option_codes.index(option_code)]

    if re.fullmatch(r"\d+(?:\.0+)?", text):
        option_code = int(float(text))
        if option_code in question.option_codes:
            option_index = question.option_codes.index(option_code)
            return question.options[option_index]

    label = _coded_question_label(question, code_question)
    raise ValueError(f"{label} 第 {row_number} 行存在未知选项编码「{text}」")


def _readable_labels(
    source: list[str],
    source_codes: list[int],
    response: list[str],
    response_codes: list[int],
) -> list[str]:
    if response and set(response_codes) == set(source_codes):
        response_by_code = dict(zip(response_codes, response))
        return [response_by_code[code] for code in source_codes]
    return source


def _decode_readable_choice(
    value: str,
    question: _BestedQuestion,
    code_question: _BestedCodeQuestion,
    row_number: int,
) -> str:
    text = str(value or "").strip()
    if not text:
        return ""

    source_by_value = {_norm(option): option for option in question.options}
    if _norm(text) in source_by_value:
        return source_by_value[_norm(text)]

    if len(code_question.options) == len(question.options):
        code_by_value = {
            _norm(option): code
            for option, code in zip(
                code_question.options, code_question.option_codes,
            )
        }
        option_code = code_by_value.get(_norm(text))
        if option_code is not None:
            return question.options[question.option_codes.index(option_code)]
        if any(
            _norm(source) != _norm(response)
            for source, response in zip(
                question.options, code_question.options,
            )
        ):
            raise ValueError(
                f"Q{question.qid} 第 {row_number} 行存在无法映射到原问卷的选项「{text}」"
            )
    raise ValueError(
        f"Q{question.qid} 第 {row_number} 行存在无法映射到原问卷的选项「{text}」"
    )


def _decode_readable_multi_choice(
    value: str,
    question: _BestedQuestion,
    code_question: _BestedCodeQuestion,
) -> list[str] | None:
    selected = _parse_same_column_multi_value(value, question.options)
    if selected is not None:
        return selected
    if len(code_question.options) != len(question.options):
        return None
    selected = _parse_same_column_multi_value(value, code_question.options)
    if selected is None:
        return None
    return [
        question.options[question.option_codes.index(
            code_question.option_codes[code_question.options.index(option)]
        )]
        for option in selected
    ]


def _parse_same_column_multi_value(
    value: str,
    options: list[str],
) -> list[str] | None:
    """按完整选项文本解析倍市得同列多选，避免误切选项内的逗号。"""
    text = str(value or "").strip()
    if not text:
        return []

    ordered_options = sorted(options, key=len, reverse=True)
    memo: dict[int, list[str] | None] = {}

    def walk(position: int) -> list[str] | None:
        if position == len(text):
            return []
        if position in memo:
            return memo[position]

        for option in ordered_options:
            if not text.startswith(option, position):
                continue
            end = position + len(option)
            if end == len(text):
                memo[position] = [option]
                return memo[position]
            if text[end] != ",":
                continue
            next_position = end + 1
            while next_position < len(text) and text[next_position].isspace():
                next_position += 1
            remaining = walk(next_position)
            if remaining is not None:
                memo[position] = [option, *remaining]
                return memo[position]

        memo[position] = None
        return None

    return walk(0)


def _translation_sources(questions: list[dict]) -> list[dict]:
    sources: list[dict] = []
    for question in questions:
        question_id = str(question.get("source_question_id") or "").strip()
        if not question_id:
            continue
        sources.append({
            "question_id": question_id,
            "name": str(question.get("name_zh") or "").strip(),
            "options": [
                str(value).strip() for value in question.get("options") or []
            ],
            "rows": [
                str(value).strip() for value in question.get("rows") or []
            ],
        })
    return sources


def build_questionnaire_translation_query(questions: list[dict]) -> str:
    """构造只包含可翻译文本的任务，不向模型开放题型与列结构。"""
    payload = {"questions": _translation_sources(questions)}
    return (
        "请逐题翻译以下 JSON 数据。question_id、数组长度和数组顺序必须原样保持。\n"
        "只输出一个 JSON 对象，不要解释，不要 Markdown。\n"
        "输出结构："
        '{"translations":[{"question_id":"Q1","name_zh":"中文题干",'
        '"options_zh":["中文选项1"],"rows_zh":["中文矩阵行1"]}]}\n'
        "待翻译数据：\n"
        + json.dumps(payload, ensure_ascii=False)
    )


def _translation_json(answer: str) -> dict:
    text = str(answer or "").strip()
    fenced = _JSON_FENCE_RE.search(text)
    if fenced:
        text = fenced.group(1).strip()
    try:
        parsed = json.loads(text)
    except json.JSONDecodeError as exc:
        raise ValueError(f"翻译结果不是有效 JSON：{exc}") from exc
    if not isinstance(parsed, dict):
        raise ValueError("翻译结果顶层必须是 JSON 对象")
    return parsed


def _translated_list(
    item: dict,
    key: str,
    source_values: list[str],
    question_id: str,
) -> list[str]:
    values = item.get(key)
    if not isinstance(values, list) or len(values) != len(source_values):
        raise ValueError(
            f"{question_id} 的 {key} 数量与原问卷不一致"
        )
    translated = [str(value or "").strip() for value in values]
    if any(not value for value in translated):
        raise ValueError(f"{question_id} 的 {key} 存在空翻译")
    normalized = [_norm(value) for value in translated]
    if len(normalized) != len(set(normalized)):
        raise ValueError(f"{question_id} 的 {key} 出现重复翻译")
    return translated


def parse_questionnaire_translations(
    answer: str,
    questions: list[dict],
) -> dict[str, dict]:
    """严格校验翻译覆盖率与数组长度，避免翻译结果改变问卷结构。"""
    sources = _translation_sources(questions)
    expected = {source["question_id"]: source for source in sources}
    parsed = _translation_json(answer)
    items = parsed.get("translations")
    if not isinstance(items, list):
        raise ValueError("翻译结果缺少 translations 数组")

    translations: dict[str, dict] = {}
    for item in items:
        if not isinstance(item, dict):
            raise ValueError("translations 中存在非对象元素")
        question_id = str(item.get("question_id") or "").strip()
        if question_id not in expected:
            raise ValueError(f"翻译结果包含未知题号：{question_id or '空'}")
        if question_id in translations:
            raise ValueError(f"翻译结果中 {question_id} 重复")
        source = expected[question_id]
        name_zh = str(item.get("name_zh") or "").strip()
        if not name_zh:
            raise ValueError(f"{question_id} 缺少中文题干")
        if _LATIN_RE.search(source["name"]) and not _CHINESE_RE.search(name_zh):
            raise ValueError(f"{question_id} 的题干未翻译为中文")
        translations[question_id] = {
            "name_zh": name_zh,
            "options_zh": _translated_list(
                item, "options_zh", source["options"], question_id,
            ),
            "rows_zh": _translated_list(
                item, "rows_zh", source["rows"], question_id,
            ),
        }

    missing = [question_id for question_id in expected if question_id not in translations]
    if missing:
        raise ValueError(f"翻译结果缺少题目：{'、'.join(missing[:5])}")
    return translations


def apply_questionnaire_translations(
    questions: list[dict],
    translations: dict[str, dict],
) -> list[dict]:
    """只替换展示文本，并为中文选项保留到原始回答值的确定性映射。"""
    translated_questions: list[dict] = []
    for question in questions:
        updated = dict(question)
        question_id = str(question.get("source_question_id") or "").strip()
        translated = translations.get(question_id)
        if not translated:
            translated_questions.append(updated)
            continue

        original_name = str(question.get("name_zh") or "").strip()
        updated["name_original"] = original_name
        updated["name_zh"] = translated["name_zh"]

        original_options = [
            str(value).strip() for value in question.get("options") or []
        ]
        if original_options:
            options_zh = list(translated["options_zh"])
            updated["options"] = options_zh
            updated["options_original"] = original_options
            aliases: dict[str, list[str]] = {}
            for canonical, original in zip(options_zh, original_options):
                if _norm(canonical) != _norm(original):
                    aliases[canonical] = [original]
            if aliases:
                updated["value_aliases"] = aliases
            else:
                updated.pop("value_aliases", None)

        original_rows = [
            str(value).strip() for value in question.get("rows") or []
        ]
        if original_rows:
            updated["rows_original"] = original_rows
            updated["rows"] = list(translated["rows_zh"])
        translated_questions.append(updated)
    return translated_questions


def parse_bested_qualitative_upload(
    response_content: bytes,
    questionnaire_content: bytes,
) -> dict:
    """返回标准化 rows 与确定性题型，供定性报告后续流程直接复用。"""
    source_questions, questionnaire_text = _parse_bested_questionnaire(questionnaire_content)
    data_rows, code_questions = _parse_response_workbook(response_content)
    headers = data_rows[0]
    body = data_rows[1:]
    coded_layout = _uses_coded_data_headers(headers, code_questions)
    question_pairs = (
        _coded_question_pairs(source_questions, code_questions)
        if coded_layout
        else _readable_question_pairs(source_questions, code_questions)
    )

    normalized_headers: list[str] = []
    normalized_columns: list[list[str]] = []
    detected_questions: list[dict] = []
    used_indexes: set[int] = set()
    cursor = 0

    for question, code_question in question_pairs:
        qid = question.qid
        code_title = code_question.title
        role = question.role
        if _is_contact_field(question.title):
            role = "ignore"

        if coded_layout and question.role == "multi_choice":
            label = _coded_question_label(question, code_question)
            source_indexes = _coded_split_indexes(
                headers,
                code_question.qid,
                question.option_codes,
                used_indexes,
                label,
            )
            combined: list[str] = []
            for row_number, row in enumerate(body, start=2):
                selected = [
                    option
                    for source_index, option in zip(
                        source_indexes, question.options,
                    )
                    if _coded_binary_selected(
                        _row_value(row, source_index), label, row_number,
                    )
                ]
                combined.append("\n".join(selected))

            target_index = len(normalized_headers)
            normalized_headers.append(question.title)
            normalized_columns.append(combined)
            detected_questions.append({
                "name_zh": question.title,
                "role": role,
                "column_indexes": [target_index],
                "delimiter": "\n",
                "options": list(question.options),
                "options_original": list(question.options),
                "source_question_id": f"Q{qid}",
            })
            used_indexes.update(source_indexes)
            cursor = max(cursor, max(source_indexes) + 1)
            continue

        if coded_layout and question.role in {
            "matrix_single", "matrix_multi", "matrix_scale",
        }:
            label = _coded_question_label(question, code_question)
            raise ValueError(f"{label} 暂不支持编码式矩阵回答列")

        if coded_layout:
            label = _coded_question_label(question, code_question)
            source_index = _coded_answer_index(
                headers,
                code_question.qid,
                used_indexes,
                label,
                open_text=question.role == "open_text",
            )
            target_index = len(normalized_headers)
            normalized_headers.append(question.title)
            if question.role == "single_choice":
                normalized_columns.append([
                    _decode_coded_choice(
                        _row_value(row, source_index),
                        question,
                        code_question,
                        row_number,
                    )
                    for row_number, row in enumerate(body, start=2)
                ])
            else:
                normalized_columns.append([
                    _row_value(row, source_index) for row in body
                ])
            detected = {
                "name_zh": question.title,
                "role": role,
                "column_indexes": [target_index],
                "source_question_id": f"Q{qid}",
            }
            if role == "single_choice":
                detected["options"] = list(question.options)
                detected["options_original"] = list(question.options)
            detected_questions.append(detected)
            used_indexes.add(source_index)
            cursor = source_index + 1
            continue

        if question.role == "multi_choice":
            split_prefix = f"{code_title}__"
            split_candidates = [
                index for index, header in enumerate(headers)
                if index not in used_indexes and header.startswith(split_prefix)
            ]
            same_column_index = _find_exact_header(
                headers, code_title, used_indexes, cursor,
            )
            if split_candidates and same_column_index is not None:
                raise ValueError(
                    f"题目「{code_title}」同时存在同列与拆列多选回答"
                )

            combined: list[str] = []
            if split_candidates:
                source_indexes = _group_headers(
                    headers,
                    code_title,
                    _readable_labels(
                        question.options,
                        question.option_codes,
                        code_question.options,
                        code_question.option_codes,
                    ),
                    used_indexes,
                )
                for row in body:
                    selected: list[str] = []
                    for source_index, option in zip(
                        source_indexes, question.options,
                    ):
                        value = _row_value(row, source_index).strip()
                        if value:
                            selected.append(option)
                    combined.append("\n".join(selected))
            elif same_column_index is not None:
                source_indexes = [same_column_index]
                for row_number, row in enumerate(body, start=2):
                    value = _row_value(row, same_column_index)
                    selected = _decode_readable_multi_choice(
                        value, question, code_question,
                    )
                    if selected is None:
                        raise ValueError(
                            f"题目「{code_title}」第 {row_number} 行的同列多选答案"
                            "无法按原问卷选项匹配"
                        )
                    combined.append("\n".join(selected))
            else:
                _group_headers(
                    headers,
                    code_title,
                    _readable_labels(
                        question.options,
                        question.option_codes,
                        code_question.options,
                        code_question.option_codes,
                    ),
                    used_indexes,
                )
                raise AssertionError("多选题列匹配未返回结果")

            target_index = len(normalized_headers)
            normalized_headers.append(question.title)
            normalized_columns.append(combined)
            detected_questions.append({
                "name_zh": question.title,
                "role": role,
                "column_indexes": [target_index],
                "delimiter": "\n",
                "options": list(question.options),
                "options_original": list(question.options),
                "source_question_id": f"Q{qid}",
            })
            used_indexes.update(source_indexes)
            cursor = max(cursor, max(source_indexes) + 1)
            continue

        if question.role in {"matrix_single", "matrix_multi", "matrix_scale"}:
            source_indexes = _group_headers(
                headers,
                code_title,
                _readable_labels(
                    question.rows,
                    question.row_codes,
                    code_question.rows,
                    code_question.row_codes,
                ),
                used_indexes,
            )
            target_indexes: list[int] = []
            for source_index, row_label in zip(source_indexes, question.rows):
                target_index = len(normalized_headers)
                target_indexes.append(target_index)
                normalized_headers.append(f"{question.title} [{row_label}]")
                values: list[str] = []
                for row_number, row in enumerate(body, start=2):
                    value = _row_value(row, source_index)
                    if question.role == "matrix_single":
                        value = _decode_readable_choice(
                            value, question, code_question, row_number,
                        )
                    elif question.role == "matrix_multi":
                        selected = _decode_readable_multi_choice(
                            value, question, code_question,
                        )
                        if selected is None:
                            raise ValueError(
                                f"题目「{code_title}」第 {row_number} 行的矩阵多选答案"
                                "无法按原问卷选项匹配"
                            )
                        value = "\n".join(selected)
                    values.append(value)
                normalized_columns.append(values)
            detected = {
                "name_zh": question.title,
                "role": role,
                "column_indexes": target_indexes,
                "rows": list(question.rows),
                "source_question_id": f"Q{qid}",
            }
            if role in {"matrix_single", "matrix_multi"}:
                detected["options"] = list(question.options)
                detected["options_original"] = list(question.options)
            if role == "matrix_multi":
                detected["delimiter"] = "\n"
            if role == "matrix_scale":
                detected["scale_min"] = 1
                detected["scale_max"] = max(5, len(question.options))
            detected_questions.append(detected)
            used_indexes.update(source_indexes)
            cursor = max(cursor, max(source_indexes) + 1)
            continue

        source_index = _find_exact_header(headers, code_title, used_indexes, cursor)
        if source_index is None:
            raise ValueError(f"未找到 Q{qid}「{code_title}」对应的回答列")
        target_index = len(normalized_headers)
        normalized_headers.append(question.title)
        if question.role == "single_choice":
            normalized_columns.append([
                _decode_readable_choice(
                    _row_value(row, source_index),
                    question,
                    code_question,
                    row_number,
                )
                for row_number, row in enumerate(body, start=2)
            ])
        else:
            normalized_columns.append([
                _row_value(row, source_index) for row in body
            ])
        detected = {
            "name_zh": question.title,
            "role": role,
            "column_indexes": [target_index],
            "source_question_id": f"Q{qid}",
        }
        if role == "single_choice":
            detected["options"] = list(question.options)
            detected["options_original"] = list(question.options)
        detected_questions.append(detected)
        used_indexes.add(source_index)
        cursor = source_index + 1

    for source_index, header in enumerate(headers):
        if source_index in used_indexes:
            continue
        low = header.casefold()
        if "role_id" in low or "roleid" in low:
            role = "id"
        else:
            role = "ignore"
        target_index = len(normalized_headers)
        normalized_headers.append(header)
        normalized_columns.append([
            _row_value(row, source_index) for row in body
        ])
        detected_questions.append({
            "name_zh": header,
            "role": role,
            "column_indexes": [target_index],
        })

    normalized_rows = [normalized_headers]
    for row_index in range(len(body)):
        normalized_rows.append([
            column[row_index] for column in normalized_columns
        ])

    return {
        "rows": normalized_rows,
        "questions": detected_questions,
        "questionnaire_text": questionnaire_text,
        "matched_questions": len(question_pairs),
    }
