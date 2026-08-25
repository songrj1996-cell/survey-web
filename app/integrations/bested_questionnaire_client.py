"""Bested questionnaire upload connector.

The verified Bested boundary is a user-supplied original-questionnaire Excel
workbook.  This module deliberately has no URL, HTTP, cookie, or persistence
path: online Bested acquisition remains unsupported.
"""

from __future__ import annotations

import hashlib
import io
import posixpath
import re
import stat
import warnings
import zipfile
from dataclasses import dataclass, field
from pathlib import PurePosixPath
from xml.etree import ElementTree

import openpyxl
from openpyxl.utils import get_column_letter


_QUESTION_RE = re.compile(r"^Q(\d+)\[([^\]]+)\]$")
_REMOTE_SOURCE_RE = re.compile(
    r"^[a-z][a-z0-9+.-]*://",
    re.IGNORECASE,
)
_WINDOWS_DRIVE_RE = re.compile(r"^[A-Za-z]:")
_CELL_REFERENCE_RE = re.compile(r"^\$?([A-Za-z]{1,3})\$?([1-9]\d*)$")
_DIMENSION_REFERENCE_RE = re.compile(
    r"^\$?([A-Za-z]{1,3})\$?([1-9]\d*)"
    r"(?::\$?([A-Za-z]{1,3})\$?([1-9]\d*))?$"
)
_RELATIONSHIP_NAMESPACE = (
    "http://schemas.openxmlformats.org/package/2006/relationships"
)
_SPREADSHEET_NAMESPACE = (
    "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
)
_OFFICE_RELATIONSHIP_NAMESPACE = (
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
)
_CONTENT_TYPES_NAMESPACE = (
    "http://schemas.openxmlformats.org/package/2006/content-types"
)
_DRAWING_RELATIONSHIP_TYPE = (
    _OFFICE_RELATIONSHIP_NAMESPACE + "/drawing"
)
_IMAGE_RELATIONSHIP_TYPE = (
    _OFFICE_RELATIONSHIP_NAMESPACE + "/image"
)
_WORKSHEET_RELATIONSHIP_TYPE = (
    _OFFICE_RELATIONSHIP_NAMESPACE + "/worksheet"
)
_WORKSHEET_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"
)
_DRAWING_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.drawing+xml"
)
_SPREADSHEET_DRAWING_NAMESPACE = (
    "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
)
_DRAWINGML_NAMESPACE = (
    "http://schemas.openxmlformats.org/drawingml/2006/main"
)
_OOXML_REQUIRED_MEMBERS = frozenset({
    "[Content_Types].xml",
    "_rels/.rels",
    "xl/workbook.xml",
    "xl/_rels/workbook.xml.rels",
})
_XLSX_MAX_ARCHIVE_BYTES = 50 * 1024 * 1024
_XLSX_MAX_MEMBERS = 4096
_XLSX_MAX_MEMBER_BYTES = 64 * 1024 * 1024
_XLSX_MAX_TOTAL_UNCOMPRESSED_BYTES = 256 * 1024 * 1024
_XLSX_MAX_COMPRESSION_RATIO = 200.0
_XLSX_MAX_IMAGES = 256
_XLSX_MAX_TOTAL_IMAGE_BYTES = 64 * 1024 * 1024
_XLSX_MAX_METADATA_XML_BYTES = 4 * 1024 * 1024
_XLSX_MAX_RELATIONSHIPS = 8192
_XLSX_MAX_CONTENT_TYPE_ENTRIES = 4096
_XLSX_MAX_DRAWING_ANCHORS = 1024
_XLSX_MAX_SHEETS = 64
_XLSX_MAX_ROWS_PER_SHEET = 100_000
_XLSX_MAX_COLUMNS_PER_SHEET = 1024
_XLSX_MAX_CELLS_PER_SHEET = 1_000_000
_XLSX_MAX_MERGED_RANGES_PER_SHEET = 10_000
_XLSX_MAX_MERGED_CELLS_PER_RANGE = 100_000
_XLSX_MAX_TOTAL_MERGED_CELLS_PER_SHEET = 1_000_000
_ZIP_READ_CHUNK_BYTES = 1024 * 1024
_BESTED_ROLE_MAP = {
    "单选题": "single_choice",
    "多选题": "multi_choice",
    "矩阵单选题": "matrix_single",
    "矩阵多选题": "matrix_multi",
    "矩阵打分题": "matrix_scale",
    "矩阵量表题": "matrix_scale",
    "量表题": "scale",
    "打分题": "scale",
    "填空题": "open_text",
}


@dataclass(frozen=True)
class BestedQuestionnaireQuestion:
    """One question preserved from the Bested provider workbook."""

    qid: int
    source_type: str
    role: str
    title: str
    options: tuple[str, ...] = ()
    rows: tuple[str, ...] = ()
    sheet_name: str = ""
    source_row: int = 0
    source_cell: str = ""
    raw_heading: str = ""
    raw_rows: tuple[tuple[str, ...], ...] = ()


@dataclass(frozen=True)
class BestedQuestionnaireImage:
    """Embedded workbook image with only file-local source evidence."""

    content: bytes = field(repr=False)
    mime_type: str
    sheet_name: str
    source_cell: str | None
    source_row: int | None
    coverage: str | None
    question_qid: int | None


@dataclass(frozen=True)
class BestedQuestionnaireMediaIssue:
    """Safe, file-local evidence for one recoverable media failure."""

    code: str
    sheet_name: str | None = None
    source_cell: str | None = None
    source_row: int | None = None


@dataclass(frozen=True)
class BestedQuestionnaireHyperlink:
    """Cell hyperlink preserved as a reference; the URL is never fetched here."""

    url: str
    display_text: str
    sheet_name: str
    source_cell: str
    source_row: int
    question_qid: int | None


@dataclass(frozen=True)
class BestedQuestionnaireParseResult:
    """Stable, file-local result of parsing a Bested questionnaire export."""

    content_sha256: str
    sheet_name: str
    provider_rows: tuple[tuple[str, ...], ...]
    questions: tuple[BestedQuestionnaireQuestion, ...]
    questionnaire_text: str
    images: tuple[BestedQuestionnaireImage, ...] = ()
    hyperlinks: tuple[BestedQuestionnaireHyperlink, ...] = ()
    media_issues: tuple[BestedQuestionnaireMediaIssue, ...] = ()


@dataclass
class _QuestionDraft:
    qid: int
    source_type: str
    role: str
    title: str
    sheet_name: str
    source_row: int
    source_cell: str
    raw_heading: str
    options: list[str] = field(default_factory=list)
    rows: list[str] = field(default_factory=list)
    raw_rows: list[tuple[str, ...]] = field(default_factory=list)

    def freeze(self) -> BestedQuestionnaireQuestion:
        return BestedQuestionnaireQuestion(
            qid=self.qid,
            source_type=self.source_type,
            role=self.role,
            title=self.title,
            options=tuple(self.options),
            rows=tuple(self.rows),
            sheet_name=self.sheet_name,
            source_row=self.source_row,
            source_cell=self.source_cell,
            raw_heading=self.raw_heading,
            raw_rows=tuple(self.raw_rows),
        )


def _cell_text(value: object) -> str:
    return "" if value is None else str(value).strip()


def _validate_zip_member(info: zipfile.ZipInfo) -> None:
    name = info.filename
    if (
        not name
        or "\\" in name
        or "\x00" in name
        or _WINDOWS_DRIVE_RE.match(name)
    ):
        raise ValueError("倍市得 .xlsx 包含非法成员路径")
    is_directory = info.is_dir()
    checked_name = name[:-1] if is_directory else name
    path = PurePosixPath(checked_name)
    if path.is_absolute() or any(
        part in {"", ".", ".."} for part in checked_name.split("/")
    ):
        raise ValueError(f"倍市得 .xlsx 包含不安全路径：{name}")
    if info.flag_bits & 0x1:
        raise ValueError(f"倍市得 .xlsx 不允许加密成员：{name}")
    mode = info.external_attr >> 16
    if stat.S_IFMT(mode) == stat.S_IFLNK:
        raise ValueError(f"倍市得 .xlsx 不允许符号链接：{name}")
    if info.compress_type not in {zipfile.ZIP_STORED, zipfile.ZIP_DEFLATED}:
        raise ValueError(f"倍市得 .xlsx 成员压缩算法不受支持：{name}")
    if is_directory:
        if info.file_size != 0:
            raise ValueError(f"倍市得 .xlsx 目录成员必须为空：{name}")
        return
    if (
        info.file_size < 0
        or info.file_size > _XLSX_MAX_MEMBER_BYTES
    ):
        raise ValueError(f"倍市得 .xlsx 单个成员解压超过安全上限：{name}")
    if info.compress_size < 0:
        raise ValueError(f"倍市得 .xlsx 成员压缩大小无效：{name}")
    if info.file_size and not info.compress_size:
        raise ValueError(f"倍市得 .xlsx 成员压缩率异常：{name}")
    if (
        info.compress_size
        and info.file_size / info.compress_size
        > _XLSX_MAX_COMPRESSION_RATIO
    ):
        raise ValueError(f"倍市得 .xlsx 成员压缩率超过安全上限：{name}")


def _read_member_for_validation(
    archive: zipfile.ZipFile,
    info: zipfile.ZipInfo,
) -> int:
    actual_size = 0
    try:
        with archive.open(info, "r") as source:
            while True:
                remaining = _XLSX_MAX_MEMBER_BYTES - actual_size
                chunk = source.read(min(_ZIP_READ_CHUNK_BYTES, remaining + 1))
                if not chunk:
                    break
                actual_size += len(chunk)
                if actual_size > _XLSX_MAX_MEMBER_BYTES:
                    raise ValueError(
                        "倍市得 .xlsx 单个成员实际解压超过安全上限："
                        f"{info.filename}"
                    )
    except ValueError:
        raise
    except (
        EOFError,
        NotImplementedError,
        OSError,
        RuntimeError,
        zipfile.BadZipFile,
    ) as exc:
        raise ValueError(
            f"倍市得 .xlsx 成员读取或校验失败：{info.filename}"
        ) from exc
    if actual_size != info.file_size:
        raise ValueError(f"倍市得 .xlsx 成员声明大小与实际不一致：{info.filename}")
    return actual_size


def _xml_root(
    archive: zipfile.ZipFile,
    info_by_name: dict[str, zipfile.ZipInfo],
    name: str,
    label: str,
) -> ElementTree.Element:
    info = info_by_name.get(name)
    if info is None:
        raise ValueError(f"倍市得 .xlsx 缺少{label}：{name}")
    if info.file_size > _XLSX_MAX_METADATA_XML_BYTES:
        raise ValueError(f"倍市得 .xlsx {label} XML 超过安全上限：{name}")
    try:
        with archive.open(info, "r") as source:
            return ElementTree.parse(source).getroot()
    except (ElementTree.ParseError, OSError, RuntimeError, zipfile.BadZipFile) as exc:
        raise ValueError(f"倍市得 .xlsx {label} XML 无效：{name}") from exc


def _part_relationships_name(part_name: str) -> str:
    path = PurePosixPath(part_name)
    return str(path.parent / "_rels" / f"{path.name}.rels")


def _resolve_ooxml_target(
    source_part: str,
    target: str,
    *,
    label: str,
) -> str:
    value = str(target or "").strip()
    if (
        not value
        or "\\" in value
        or "\x00" in value
        or "?" in value
        or "#" in value
        or _WINDOWS_DRIVE_RE.match(value)
        or _REMOTE_SOURCE_RE.match(value)
        or value.startswith("//")
    ):
        raise ValueError(f"倍市得 .xlsx {label}目标非法")
    if value.startswith("/"):
        combined = value.lstrip("/")
    else:
        combined = posixpath.join(
            str(PurePosixPath(source_part).parent),
            value,
        )
    normalized = posixpath.normpath(combined)
    if (
        not normalized
        or normalized == "."
        or normalized.startswith("../")
        or normalized == ".."
        or normalized.startswith("/")
        or any(part in {"", ".", ".."} for part in normalized.split("/"))
    ):
        raise ValueError(f"倍市得 .xlsx {label}目标越界")
    return normalized


def _relationships(
    archive: zipfile.ZipFile,
    info_by_name: dict[str, zipfile.ZipInfo],
    source_part: str,
    *,
    required: bool,
) -> dict[str, tuple[str, str, str]]:
    relationships_name = _part_relationships_name(source_part)
    info = info_by_name.get(relationships_name)
    if info is None:
        if required:
            raise ValueError(
                f"倍市得 .xlsx 缺少关系文件：{relationships_name}"
            )
        return {}
    root = _xml_root(
        archive,
        info_by_name,
        relationships_name,
        "关系",
    )
    expected_tag = f"{{{_RELATIONSHIP_NAMESPACE}}}Relationships"
    relationship_tag = f"{{{_RELATIONSHIP_NAMESPACE}}}Relationship"
    if root.tag != expected_tag:
        raise ValueError(f"倍市得 .xlsx 关系 XML 根节点无效：{relationships_name}")

    result: dict[str, tuple[str, str, str]] = {}
    for child in root:
        if child.tag != relationship_tag:
            raise ValueError(f"倍市得 .xlsx 关系 XML 包含未知节点：{relationships_name}")
        relationship_id = str(child.attrib.get("Id") or "").strip()
        relationship_type = str(child.attrib.get("Type") or "").strip()
        target = str(child.attrib.get("Target") or "").strip()
        target_mode = str(child.attrib.get("TargetMode") or "Internal").strip()
        if not relationship_id or relationship_id in result:
            raise ValueError(f"倍市得 .xlsx 关系 ID 缺失或重复：{relationships_name}")
        if len(result) >= _XLSX_MAX_RELATIONSHIPS:
            raise ValueError(f"倍市得 .xlsx 关系数量超过安全上限：{relationships_name}")
        result[relationship_id] = (
            relationship_type,
            target,
            target_mode,
        )
    return result


def _content_types(
    archive: zipfile.ZipFile,
    info_by_name: dict[str, zipfile.ZipInfo],
) -> tuple[dict[str, str], dict[str, str]]:
    name = "[Content_Types].xml"
    root = _xml_root(archive, info_by_name, name, "Content Types")
    if root.tag != f"{{{_CONTENT_TYPES_NAMESPACE}}}Types":
        raise ValueError("倍市得 .xlsx Content Types XML 根节点无效")
    default_tag = f"{{{_CONTENT_TYPES_NAMESPACE}}}Default"
    override_tag = f"{{{_CONTENT_TYPES_NAMESPACE}}}Override"
    defaults: dict[str, str] = {}
    overrides: dict[str, str] = {}
    for child in root:
        if len(defaults) + len(overrides) >= _XLSX_MAX_CONTENT_TYPE_ENTRIES:
            raise ValueError("倍市得 .xlsx Content Types 条目超过安全上限")
        if child.tag == default_tag:
            key = str(child.attrib.get("Extension") or "").strip().casefold()
            value = str(child.attrib.get("ContentType") or "").strip().casefold()
            if not key or not value or key in defaults:
                raise ValueError("倍市得 .xlsx Content Types 默认项缺失或重复")
            defaults[key] = value
            continue
        if child.tag == override_tag:
            raw_part = str(child.attrib.get("PartName") or "").strip()
            value = str(child.attrib.get("ContentType") or "").strip().casefold()
            part = _resolve_ooxml_target(
                "[Content_Types].xml",
                raw_part,
                label="Content Types",
            )
            if not value or part in overrides:
                raise ValueError("倍市得 .xlsx Content Types Override 缺失或重复")
            overrides[part] = value
            continue
        raise ValueError("倍市得 .xlsx Content Types XML 包含未知节点")
    return defaults, overrides


def _part_content_type(
    part: str,
    defaults: dict[str, str],
    overrides: dict[str, str],
) -> str | None:
    override = overrides.get(part)
    if override is not None:
        return override
    suffix = PurePosixPath(part).suffix
    return defaults.get(suffix.lstrip(".").casefold()) if suffix else None


def _column_index(letters: str) -> int:
    value = 0
    for character in letters.upper():
        value = value * 26 + ord(character) - ord("A") + 1
    return value


def _cell_coordinates(reference: str, *, label: str) -> tuple[int, int]:
    match = _CELL_REFERENCE_RE.fullmatch(str(reference or "").strip())
    if not match:
        raise ValueError(f"倍市得 .xlsx {label}单元格坐标无效")
    column = _column_index(match.group(1))
    row = int(match.group(2))
    if (
        row > _XLSX_MAX_ROWS_PER_SHEET
        or column > _XLSX_MAX_COLUMNS_PER_SHEET
    ):
        raise ValueError(f"倍市得 .xlsx {label}单元格坐标超过安全上限")
    return row, column


def _range_coordinates(
    reference: str,
    *,
    label: str,
) -> tuple[int, int, int, int]:
    match = _DIMENSION_REFERENCE_RE.fullmatch(str(reference or "").strip())
    if not match:
        raise ValueError(f"倍市得 .xlsx {label}范围无效")
    start_row, start_column = _cell_coordinates(
        f"{match.group(1)}{match.group(2)}",
        label=label,
    )
    end_row, end_column = _cell_coordinates(
        f"{match.group(3) or match.group(1)}"
        f"{match.group(4) or match.group(2)}",
        label=label,
    )
    if end_row < start_row or end_column < start_column:
        raise ValueError(f"倍市得 .xlsx {label}范围倒置")
    return start_row, start_column, end_row, end_column


def _validate_worksheet_xml(
    archive: zipfile.ZipFile,
    info_by_name: dict[str, zipfile.ZipInfo],
    worksheet_part: str,
) -> None:
    info = info_by_name.get(worksheet_part)
    if info is None:
        raise ValueError(f"倍市得 .xlsx 工作表目标不存在：{worksheet_part}")
    dimension_tag = f"{{{_SPREADSHEET_NAMESPACE}}}dimension"
    row_tag = f"{{{_SPREADSHEET_NAMESPACE}}}row"
    cell_tag = f"{{{_SPREADSHEET_NAMESPACE}}}c"
    merge_cell_tag = f"{{{_SPREADSHEET_NAMESPACE}}}mergeCell"
    row_nodes = 0
    cell_nodes = 0
    merge_nodes = 0
    total_merged_cells = 0
    previous_row = 0
    root_seen = False
    try:
        with archive.open(info, "r") as source:
            for event, element in ElementTree.iterparse(
                source,
                events=("start", "end"),
            ):
                if not root_seen:
                    if event != "start" or element.tag != (
                        f"{{{_SPREADSHEET_NAMESPACE}}}worksheet"
                    ):
                        raise ValueError(
                            "倍市得 .xlsx 工作表 XML 根节点无效："
                            f"{worksheet_part}"
                        )
                    root_seen = True
                    continue
                if event != "end":
                    continue
                if element.tag == dimension_tag:
                    reference = str(element.attrib.get("ref") or "").strip()
                    (
                        start_row,
                        start_column,
                        end_row,
                        end_column,
                    ) = _range_coordinates(
                        reference,
                        label="工作表 dimension ",
                    )
                    if (
                        (end_row - start_row + 1)
                        * (end_column - start_column + 1)
                        > _XLSX_MAX_CELLS_PER_SHEET
                    ):
                        raise ValueError("倍市得 .xlsx 工作表 dimension 范围超过安全上限")
                elif element.tag == row_tag:
                    row_nodes += 1
                    if row_nodes > _XLSX_MAX_ROWS_PER_SHEET:
                        raise ValueError("倍市得 .xlsx 工作表行节点超过安全上限")
                    raw_row = str(element.attrib.get("r") or "").strip()
                    if not raw_row.isdigit() or int(raw_row) < 1:
                        raise ValueError("倍市得 .xlsx 工作表行坐标无效")
                    row_number = int(raw_row)
                    if row_number > _XLSX_MAX_ROWS_PER_SHEET:
                        raise ValueError("倍市得 .xlsx 工作表行坐标超过安全上限")
                    if row_number <= previous_row:
                        raise ValueError("倍市得 .xlsx 工作表行坐标重复或无序")
                    previous_row = row_number
                elif element.tag == cell_tag:
                    cell_nodes += 1
                    if cell_nodes > _XLSX_MAX_CELLS_PER_SHEET:
                        raise ValueError("倍市得 .xlsx 工作表单元格节点超过安全上限")
                    _cell_coordinates(
                        str(element.attrib.get("r") or ""),
                        label="工作表 ",
                    )
                elif element.tag == merge_cell_tag:
                    merge_nodes += 1
                    if merge_nodes > _XLSX_MAX_MERGED_RANGES_PER_SHEET:
                        raise ValueError("倍市得 .xlsx 工作表合并区域数量超过安全上限")
                    (
                        start_row,
                        start_column,
                        end_row,
                        end_column,
                    ) = _range_coordinates(
                        str(element.attrib.get("ref") or ""),
                        label="工作表合并区域 ",
                    )
                    area = (
                        (end_row - start_row + 1)
                        * (end_column - start_column + 1)
                    )
                    if area > _XLSX_MAX_MERGED_CELLS_PER_RANGE:
                        raise ValueError("倍市得 .xlsx 工作表单个合并区域超过安全上限")
                    total_merged_cells += area
                    if (
                        total_merged_cells
                        > _XLSX_MAX_TOTAL_MERGED_CELLS_PER_SHEET
                    ):
                        raise ValueError("倍市得 .xlsx 工作表合并区域总面积超过安全上限")
                element.clear()
    except ValueError:
        raise
    except (ElementTree.ParseError, OSError, RuntimeError, zipfile.BadZipFile) as exc:
        raise ValueError(f"倍市得 .xlsx 工作表 XML 无效：{worksheet_part}") from exc


def _workbook_worksheet_parts(
    archive: zipfile.ZipFile,
    info_by_name: dict[str, zipfile.ZipInfo],
    defaults: dict[str, str],
    overrides: dict[str, str],
) -> tuple[str, ...]:
    workbook_part = "xl/workbook.xml"
    root = _xml_root(archive, info_by_name, workbook_part, "workbook")
    if root.tag != f"{{{_SPREADSHEET_NAMESPACE}}}workbook":
        raise ValueError("倍市得 .xlsx workbook XML 根节点无效")
    relationships = _relationships(
        archive,
        info_by_name,
        workbook_part,
        required=True,
    )
    sheet_tag = f"{{{_SPREADSHEET_NAMESPACE}}}sheet"
    relationship_id_attribute = f"{{{_OFFICE_RELATIONSHIP_NAMESPACE}}}id"
    sheet_parts: list[str] = []
    used_relationship_ids: set[str] = set()
    for sheet in root.iter(sheet_tag):
        relationship_id = str(sheet.attrib.get(relationship_id_attribute) or "").strip()
        relationship = relationships.get(relationship_id)
        if relationship is None or relationship_id in used_relationship_ids:
            raise ValueError("倍市得 .xlsx 工作表关系缺失或重复")
        used_relationship_ids.add(relationship_id)
        relationship_type, target, target_mode = relationship
        if (
            relationship_type != _WORKSHEET_RELATIONSHIP_TYPE
            or target_mode.casefold() != "internal"
        ):
            raise ValueError("倍市得 .xlsx 工作表关系类型或模式无效")
        part = _resolve_ooxml_target(
            workbook_part,
            target,
            label="工作表关系",
        )
        if part in sheet_parts:
            raise ValueError("倍市得 .xlsx 多个工作表关系指向同一目标")
        if info_by_name.get(part) is None:
            raise ValueError(f"倍市得 .xlsx 工作表目标不存在：{part}")
        if _part_content_type(part, defaults, overrides) != _WORKSHEET_CONTENT_TYPE:
            raise ValueError(f"倍市得 .xlsx 工作表 Content Type 无效：{part}")
        sheet_parts.append(part)
    if not sheet_parts or len(sheet_parts) > _XLSX_MAX_SHEETS:
        raise ValueError("倍市得 .xlsx 工作表数量无效或超过安全上限")
    for part in sheet_parts:
        _validate_worksheet_xml(archive, info_by_name, part)
    return tuple(sheet_parts)


def _drawing_parts(
    archive: zipfile.ZipFile,
    info_by_name: dict[str, zipfile.ZipInfo],
    sheet_parts: tuple[str, ...],
    defaults: dict[str, str],
    overrides: dict[str, str],
) -> tuple[str, ...]:
    result: list[str] = []
    for sheet_part in sheet_parts:
        for relationship_type, target, target_mode in _relationships(
            archive,
            info_by_name,
            sheet_part,
            required=False,
        ).values():
            if relationship_type != _DRAWING_RELATIONSHIP_TYPE:
                continue
            if target_mode.casefold() != "internal":
                raise ValueError("倍市得 .xlsx drawing 关系不允许外部目标")
            part = _resolve_ooxml_target(
                sheet_part,
                target,
                label="drawing 关系",
            )
            if part in result:
                raise ValueError("倍市得 .xlsx drawing 目标存在重复歧义")
            if info_by_name.get(part) is None:
                raise ValueError(f"倍市得 .xlsx drawing 目标不存在：{part}")
            if _part_content_type(part, defaults, overrides) != _DRAWING_CONTENT_TYPE:
                raise ValueError(f"倍市得 .xlsx drawing Content Type 无效：{part}")
            result.append(part)
    return tuple(result)


def _drawing_image_parts(
    archive: zipfile.ZipFile,
    info_by_name: dict[str, zipfile.ZipInfo],
    drawing_parts: tuple[str, ...],
    defaults: dict[str, str],
    overrides: dict[str, str],
) -> tuple[tuple[str, ...], int]:
    unique_parts: list[str] = []
    image_uses = 0
    anchor_count = 0
    anchor_tags = {
        f"{{{_SPREADSHEET_DRAWING_NAMESPACE}}}oneCellAnchor",
        f"{{{_SPREADSHEET_DRAWING_NAMESPACE}}}twoCellAnchor",
        f"{{{_SPREADSHEET_DRAWING_NAMESPACE}}}absoluteAnchor",
    }
    blip_tag = f"{{{_DRAWINGML_NAMESPACE}}}blip"
    embed_attribute = f"{{{_OFFICE_RELATIONSHIP_NAMESPACE}}}embed"
    link_attribute = f"{{{_OFFICE_RELATIONSHIP_NAMESPACE}}}link"
    for drawing_part in drawing_parts:
        relationships = _relationships(
            archive,
            info_by_name,
            drawing_part,
            required=True,
        )
        info = info_by_name[drawing_part]
        if info.file_size > _XLSX_MAX_METADATA_XML_BYTES:
            raise ValueError(f"倍市得 .xlsx drawing XML 超过安全上限：{drawing_part}")
        root_seen = False
        try:
            with archive.open(info, "r") as source:
                for event, element in ElementTree.iterparse(
                    source,
                    events=("start", "end"),
                ):
                    if not root_seen:
                        if event != "start" or element.tag != (
                            f"{{{_SPREADSHEET_DRAWING_NAMESPACE}}}wsDr"
                        ):
                            raise ValueError(
                                "倍市得 .xlsx drawing XML 根节点无效："
                                f"{drawing_part}"
                            )
                        root_seen = True
                        continue
                    if event != "end" or element.tag not in anchor_tags:
                        continue
                    anchor_count += 1
                    if anchor_count > _XLSX_MAX_DRAWING_ANCHORS:
                        raise ValueError("倍市得 .xlsx drawing anchor 数量超过安全上限")
                    for blip in element.iter(blip_tag):
                        image_uses += 1
                        if image_uses > _XLSX_MAX_IMAGES:
                            raise ValueError("倍市得 .xlsx drawing 图片使用数超过安全上限")
                        relationship_id = str(
                            blip.attrib.get(embed_attribute) or ""
                        ).strip()
                        if (
                            blip.attrib.get(link_attribute) is not None
                            or not relationship_id
                        ):
                            raise ValueError(
                                "倍市得 .xlsx drawing 图片不允许外部或空引用："
                                f"{drawing_part}"
                            )
                        relationship = relationships.get(relationship_id)
                        if relationship is None:
                            raise ValueError(
                                "倍市得 .xlsx drawing 图片关系不存在："
                                f"{relationship_id}"
                            )
                        relationship_type, target, target_mode = relationship
                        if (
                            relationship_type != _IMAGE_RELATIONSHIP_TYPE
                            or target_mode.casefold() != "internal"
                        ):
                            raise ValueError(
                                "倍市得 .xlsx drawing 图片关系类型或模式无效："
                                f"{relationship_id}"
                            )
                        part = _resolve_ooxml_target(
                            drawing_part,
                            target,
                            label="drawing 图片关系",
                        )
                        if info_by_name.get(part) is None:
                            raise ValueError(
                                "倍市得 .xlsx drawing 图片目标不存在："
                                f"{part}"
                            )
                        content_type = _part_content_type(
                            part,
                            defaults,
                            overrides,
                        )
                        if (
                            content_type is None
                            or not content_type.startswith("image/")
                        ):
                            raise ValueError(
                                "倍市得 .xlsx drawing 图片 Content Type 无效："
                                f"{part}"
                            )
                        if part not in unique_parts:
                            unique_parts.append(part)
                    element.clear()
        except ValueError:
            raise
        except (
            ElementTree.ParseError,
            OSError,
            RuntimeError,
            zipfile.BadZipFile,
        ) as exc:
            raise ValueError(f"倍市得 .xlsx drawing XML 无效：{drawing_part}") from exc
    return tuple(unique_parts), image_uses


def _validate_xlsx_package(content: bytes) -> None:
    if not isinstance(content, bytes):
        raise TypeError("倍市得原问卷内容必须是 bytes")
    if not content:
        raise ValueError("倍市得 .xlsx 内容为空")
    if len(content) > _XLSX_MAX_ARCHIVE_BYTES:
        raise ValueError("倍市得 .xlsx 压缩包超过安全上限")

    try:
        archive = zipfile.ZipFile(io.BytesIO(content), "r")
    except (OSError, zipfile.BadZipFile, zipfile.LargeZipFile) as exc:
        raise ValueError("倍市得原问卷不是有效的 .xlsx/ZIP 文件") from exc

    with archive:
        infos = archive.infolist()
        if not infos or len(infos) > _XLSX_MAX_MEMBERS:
            raise ValueError("倍市得 .xlsx 成员数量无效或超过安全上限")
        names = [info.filename for info in infos]
        if len(names) != len(set(names)):
            raise ValueError("倍市得 .xlsx 包含重复成员")

        total_declared = 0
        total_compressed = 0
        for info in infos:
            _validate_zip_member(info)
            total_declared += info.file_size
            total_compressed += info.compress_size
            if total_declared > _XLSX_MAX_TOTAL_UNCOMPRESSED_BYTES:
                raise ValueError("倍市得 .xlsx 声明的解压总量超过安全上限")

        if (
            total_declared
            and total_declared / max(1, total_compressed)
            > _XLSX_MAX_COMPRESSION_RATIO
        ):
            raise ValueError("倍市得 .xlsx 总压缩率超过安全上限")
        missing = _OOXML_REQUIRED_MEMBERS.difference(names)
        if missing:
            raise ValueError(
                "倍市得 .xlsx 缺少 OOXML 关键成员："
                + "、".join(sorted(missing))
            )

        info_by_name = {info.filename: info for info in infos}
        defaults, overrides = _content_types(archive, info_by_name)
        sheet_parts = _workbook_worksheet_parts(
            archive,
            info_by_name,
            defaults,
            overrides,
        )
        drawing_parts = _drawing_parts(
            archive,
            info_by_name,
            sheet_parts,
            defaults,
            overrides,
        )
        image_parts, image_uses = _drawing_image_parts(
            archive,
            info_by_name,
            drawing_parts,
            defaults,
            overrides,
        )
        if image_uses > _XLSX_MAX_IMAGES:
            raise ValueError("倍市得 .xlsx drawing 图片使用数超过安全上限")
        total_image_bytes = sum(
            info_by_name[part].file_size for part in image_parts
        )
        if total_image_bytes > _XLSX_MAX_TOTAL_IMAGE_BYTES:
            raise ValueError("倍市得 .xlsx drawing 图片总字节超过安全上限")

        total_actual = 0
        for info in infos:
            total_actual += _read_member_for_validation(archive, info)
            if total_actual > _XLSX_MAX_TOTAL_UNCOMPRESSED_BYTES:
                raise ValueError("倍市得 .xlsx 实际解压总量超过安全上限")


def worksheet_image_use_count(content: bytes, sheet_name: str) -> int:
    """Return the validated OOXML image-use count for one worksheet."""

    _validate_xlsx_package(content)
    return _validated_worksheet_image_use_count(content, sheet_name)


def _validated_worksheet_image_use_count(content: bytes, sheet_name: str) -> int:
    """Read one worksheet inventory after the package preflight has passed."""

    with zipfile.ZipFile(io.BytesIO(content), "r") as archive:
        info_by_name = {info.filename: info for info in archive.infolist()}
        defaults, overrides = _content_types(archive, info_by_name)
        workbook_part = "xl/workbook.xml"
        root = _xml_root(archive, info_by_name, workbook_part, "workbook")
        relationships = _relationships(
            archive,
            info_by_name,
            workbook_part,
            required=True,
        )
        sheet_tag = f"{{{_SPREADSHEET_NAMESPACE}}}sheet"
        relationship_attribute = f"{{{_OFFICE_RELATIONSHIP_NAMESPACE}}}id"
        worksheet_part: str | None = None
        for sheet in root.iter(sheet_tag):
            if sheet.attrib.get("name") != sheet_name:
                continue
            relationship_id = str(
                sheet.attrib.get(relationship_attribute) or ""
            ).strip()
            relationship = relationships.get(relationship_id)
            if relationship is None:
                raise ValueError("倍市得 .xlsx 问卷工作表关系缺失")
            relationship_type, target, target_mode = relationship
            if (
                relationship_type != _WORKSHEET_RELATIONSHIP_TYPE
                or target_mode.casefold() != "internal"
            ):
                raise ValueError("倍市得 .xlsx 问卷工作表关系无效")
            worksheet_part = _resolve_ooxml_target(
                workbook_part,
                target,
                label="问卷工作表关系",
            )
            break
        if worksheet_part is None:
            raise ValueError("倍市得 .xlsx 缺少已解析的问卷工作表")
        drawing_parts = _drawing_parts(
            archive,
            info_by_name,
            (worksheet_part,),
            defaults,
            overrides,
        )
        _, image_uses = _drawing_image_parts(
            archive,
            info_by_name,
            drawing_parts,
            defaults,
            overrides,
        )
        return image_uses


def _load_workbook(content: bytes, *, read_only: bool = True):
    _validate_xlsx_package(content)
    try:
        return openpyxl.load_workbook(
            io.BytesIO(content),
            read_only=read_only,
            data_only=True,
        )
    except Exception as exc:
        raise ValueError(
            "无法读取 Excel 内容，请确认文件来自倍市得且未损坏"
        ) from exc


def _worksheet_rows(worksheet) -> list[list[str]]:
    rows = [
        [_cell_text(value) for value in row]
        for row in worksheet.iter_rows(values_only=True)
    ]
    while rows and not any(cell for cell in rows[-1]):
        rows.pop()
    return rows


def _questionnaire_sheet(workbook):
    if "问卷内容" in workbook.sheetnames:
        return workbook["问卷内容"]
    return workbook[workbook.sheetnames[0]]


def _question_qid_for_row(
    questions: tuple[BestedQuestionnaireQuestion, ...],
    source_row: int | None,
) -> int | None:
    if source_row is None:
        return None
    for question in questions:
        end_row = question.source_row + len(question.raw_rows) - 1
        if question.source_row <= source_row <= end_row:
            return question.qid
    return None


def _anchor_cell(marker) -> tuple[str | None, int | None]:
    if marker is None:
        return None, None
    column = getattr(marker, "col", None)
    row = getattr(marker, "row", None)
    if not isinstance(column, int) or not isinstance(row, int):
        return None, None
    source_row = row + 1
    return f"{get_column_letter(column + 1)}{source_row}", source_row


def _image_location(
    image,
) -> tuple[str | None, int | None, int | None, str | None]:
    anchor = getattr(image, "anchor", None)
    if isinstance(anchor, str):
        match = re.fullmatch(r"([A-Za-z]+)([1-9]\d*)", anchor.strip())
        if not match:
            return None, None, None, None
        source_cell = f"{match.group(1).upper()}{match.group(2)}"
        source_row = int(match.group(2))
        return source_cell, source_row, source_row, source_cell

    source_cell, source_row = _anchor_cell(getattr(anchor, "_from", None))
    end_cell, end_row = _anchor_cell(getattr(anchor, "to", None))
    if source_cell is None:
        return None, None, None, None
    coverage = (
        f"{source_cell}:{end_cell}"
        if end_cell is not None and end_cell != source_cell
        else source_cell
    )
    return source_cell, source_row, end_row or source_row, coverage


def _image_mime_type(content: bytes) -> str:
    if content.startswith(b"\x89PNG\r\n\x1a\n"):
        return "image/png"
    if content.startswith(b"\xff\xd8\xff"):
        return "image/jpeg"
    if content.startswith((b"GIF87a", b"GIF89a")):
        return "image/gif"
    if content.startswith(b"RIFF") and content[8:12] == b"WEBP":
        return "image/webp"
    if content.startswith(b"BM"):
        return "image/bmp"
    return "application/octet-stream"


def _discover_questionnaire_media(
    content: bytes,
    sheet_name: str,
    questions: tuple[BestedQuestionnaireQuestion, ...],
) -> tuple[
    tuple[BestedQuestionnaireImage, ...],
    tuple[BestedQuestionnaireHyperlink, ...],
    tuple[BestedQuestionnaireMediaIssue, ...],
]:
    _validate_xlsx_package(content)
    expected_image_uses = _validated_worksheet_image_use_count(
        content,
        sheet_name,
    )
    try:
        with warnings.catch_warnings(record=True):
            warnings.simplefilter("always")
            workbook = openpyxl.load_workbook(
                io.BytesIO(content),
                read_only=False,
                data_only=True,
            )
    except (MemoryError, RecursionError):
        raise
    except Exception:
        return (), (), (BestedQuestionnaireMediaIssue(
            code="media_workbook_load_failed",
            sheet_name=sheet_name,
        ),)
    try:
        worksheet = workbook[sheet_name]
        if len(worksheet._images) > _XLSX_MAX_IMAGES:
            raise ValueError("倍市得 .xlsx 问卷工作表图片数量超过安全上限")
        if len(worksheet._images) > expected_image_uses:
            raise ValueError("倍市得 .xlsx 图片加载结果超过 OOXML 声明数量")
        images: list[BestedQuestionnaireImage] = []
        media_issues: list[BestedQuestionnaireMediaIssue] = [
            BestedQuestionnaireMediaIssue(
                code="image_loading_failed",
                sheet_name=sheet_name,
            )
            for _ in range(expected_image_uses - len(worksheet._images))
        ]
        total_image_bytes = 0
        seen_image_hashes: set[str] = set()
        for image in worksheet._images:
            source_cell, source_row, end_row, coverage = _image_location(image)
            try:
                image_content = image._data()
            except (MemoryError, RecursionError):
                raise
            except Exception:
                media_issues.append(BestedQuestionnaireMediaIssue(
                    code="image_extraction_failed",
                    sheet_name=sheet_name,
                    source_cell=source_cell,
                    source_row=source_row,
                ))
                continue
            image_hash = hashlib.sha256(image_content).hexdigest()
            if image_hash not in seen_image_hashes:
                seen_image_hashes.add(image_hash)
                total_image_bytes += len(image_content)
            if total_image_bytes > _XLSX_MAX_TOTAL_IMAGE_BYTES:
                raise ValueError("倍市得 .xlsx 问卷工作表图片总字节超过安全上限")
            start_qid = _question_qid_for_row(questions, source_row)
            end_qid = _question_qid_for_row(questions, end_row)
            images.append(BestedQuestionnaireImage(
                content=image_content,
                mime_type=_image_mime_type(image_content),
                sheet_name=sheet_name,
                source_cell=source_cell,
                source_row=source_row,
                coverage=coverage,
                question_qid=(
                    start_qid
                    if start_qid is not None and start_qid == end_qid
                    else None
                ),
            ))

        hyperlinks: list[BestedQuestionnaireHyperlink] = []
        for row in worksheet.iter_rows():
            for cell in row:
                hyperlink = cell.hyperlink
                if hyperlink is None:
                    continue
                url = str(hyperlink.target or hyperlink.location or "").strip()
                if not url:
                    continue
                hyperlinks.append(BestedQuestionnaireHyperlink(
                    url=url,
                    display_text=_cell_text(cell.value),
                    sheet_name=sheet_name,
                    source_cell=cell.coordinate,
                    source_row=cell.row,
                    question_qid=_question_qid_for_row(questions, cell.row),
                ))
        return tuple(images), tuple(hyperlinks), tuple(media_issues)
    finally:
        workbook.close()


def parse_bested_questionnaire(
    content: bytes,
    *,
    discover_media: bool = True,
) -> BestedQuestionnaireParseResult:
    """Parse one uploaded Bested questionnaire workbook without remote access.

    ``discover_media=False`` exists only for the legacy answer-alignment path,
    whose historical success/failure behavior must not depend on drawings.
    New connector callers should keep the default complete parse.
    """
    workbook = _load_workbook(content)
    try:
        worksheet = _questionnaire_sheet(workbook)
        sheet_name = worksheet.title
        provider_rows = _worksheet_rows(worksheet)
    finally:
        workbook.close()
    if len(provider_rows) <= 1:
        raise ValueError("调研问卷为空或缺少题目")

    question_drafts: list[_QuestionDraft] = []
    current: _QuestionDraft | None = None
    section = ""
    for source_row, row in enumerate(provider_rows[1:], start=2):
        first = row[0] if row else ""
        second = row[1] if len(row) > 1 else ""
        match = _QUESTION_RE.match(first)
        if match:
            raw_type = match.group(2).strip()
            role = _BESTED_ROLE_MAP.get(raw_type)
            if not role:
                raise ValueError(
                    f"暂不支持 Q{match.group(1)} 的题型「{raw_type}」"
                )
            if not second:
                raise ValueError(f"Q{match.group(1)} 缺少题干")
            current = _QuestionDraft(
                qid=int(match.group(1)),
                source_type=raw_type,
                role=role,
                title=second,
                sheet_name=sheet_name,
                source_row=source_row,
                source_cell=f"A{source_row}",
                raw_heading=first,
                raw_rows=[tuple(row)],
            )
            question_drafts.append(current)
            section = ""
            continue
        if not current:
            continue
        current.raw_rows.append(tuple(row))
        if first == "选项":
            section = "options"
            continue
        if first == "矩阵行":
            section = "rows"
            continue
        if first.isdigit() and second:
            if section == "options":
                current.options.append(second)
            elif section == "rows":
                current.rows.append(second)

    if not question_drafts:
        raise ValueError("未识别到 Q号[题型] 格式的题目")
    questions = tuple(question.freeze() for question in question_drafts)
    seen: set[int] = set()
    for question in questions:
        if question.qid in seen:
            raise ValueError(f"原问卷中 Q{question.qid} 重复")
        seen.add(question.qid)
        if (
            question.role
            in {"single_choice", "multi_choice", "matrix_single", "matrix_multi"}
            and not question.options
        ):
            raise ValueError(f"Q{question.qid} 缺少选项")
        if question.role.startswith("matrix_") and not question.rows:
            raise ValueError(f"Q{question.qid} 缺少矩阵行")

    questionnaire_text = "\n".join(
        " | ".join(cell for cell in row if cell)
        for row in provider_rows
        if any(row)
    )
    images: tuple[BestedQuestionnaireImage, ...] = ()
    hyperlinks: tuple[BestedQuestionnaireHyperlink, ...] = ()
    media_issues: tuple[BestedQuestionnaireMediaIssue, ...] = ()
    if discover_media:
        images, hyperlinks, media_issues = _discover_questionnaire_media(
            content,
            sheet_name,
            questions,
        )
    return BestedQuestionnaireParseResult(
        content_sha256=hashlib.sha256(content).hexdigest(),
        sheet_name=sheet_name,
        provider_rows=tuple(tuple(row) for row in provider_rows),
        questions=questions,
        questionnaire_text=questionnaire_text,
        images=images,
        hyperlinks=hyperlinks,
        media_issues=media_issues,
    )


def parse_bested_questionnaire_upload(
    filename: str,
    content: bytes,
) -> BestedQuestionnaireParseResult:
    """Validate a local upload and parse it; URLs are intentionally unsupported."""
    if not isinstance(filename, str) or not filename.strip():
        raise ValueError("倍市得原问卷上传必须提供文件名")
    normalized_filename = filename.strip()
    if (
        _REMOTE_SOURCE_RE.match(normalized_filename)
        or normalized_filename.startswith("//")
    ):
        raise ValueError(
            "倍市得在线 URL/页面抓取不受支持，请上传原问卷 Excel 文件"
        )
    upload_name = normalized_filename.replace("\\", "/")
    if PurePosixPath(upload_name).suffix.casefold() != ".xlsx":
        raise ValueError("倍市得原问卷连接器仅支持 .xlsx 上传文件")
    if not isinstance(content, bytes):
        raise TypeError("倍市得原问卷内容必须是 bytes")
    if not content:
        raise ValueError("倍市得原问卷上传内容为空")
    return parse_bested_questionnaire(content)
