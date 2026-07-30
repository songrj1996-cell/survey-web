"""访谈记录 Excel 解析：保留所有 Sheet 的非空单元格与可追溯引用。"""
from io import BytesIO
from pathlib import Path

from fastapi import HTTPException
from openpyxl import load_workbook


def _cell_text(value) -> str:
    if value is None:
        return ""
    text = str(value).replace("\x00", "").strip()
    return text


def parse_interview_workbook(filename: str, content: bytes) -> dict:
    """读取一个 xlsx 的全部 Sheet，不假设题目或玩家位于固定行列。"""
    if Path(filename).suffix.lower() != ".xlsx":
        raise HTTPException(status_code=400, detail="访谈报告目前仅支持 .xlsx 文件")
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"无法读取 Excel 文件：{exc}") from exc

    sheets: list[dict] = []
    total_cells = 0
    total_chars = 0
    try:
        for worksheet in workbook.worksheets:
            cells: list[dict] = []
            for row in worksheet.iter_rows():
                for cell in row:
                    text = _cell_text(cell.value)
                    if not text:
                        continue
                    cells.append({"cell": cell.coordinate, "value": text})
                    total_cells += 1
                    total_chars += len(text)
            sheets.append(
                {
                    "name": worksheet.title,
                    "max_row": worksheet.max_row,
                    "max_column": worksheet.max_column,
                    "nonempty_count": len(cells),
                    "cells": cells,
                }
            )
    finally:
        workbook.close()

    if not total_cells:
        raise HTTPException(status_code=400, detail="Excel 中没有可读取的访谈记录")
    return {
        "filename": filename,
        "sheets": sheets,
        "total_cells": total_cells,
        "total_chars": total_chars,
    }


def serialize_interview_workbook(workbook: dict) -> str:
    """转为给模型读取的稳定文本格式，每行都带 Sheet!Cell 引用。"""
    parts: list[str] = []
    for sheet in workbook.get("sheets", []):
        sheet_name = str(sheet.get("name") or "")
        parts.append(f"## Sheet: {sheet_name}")
        for cell in sheet.get("cells", []):
            parts.append(f"- [{sheet_name}!{cell['cell']}] {cell['value']}")
        parts.append("")
    return "\n".join(parts).strip()


def interview_source_refs(workbook: dict) -> set[str]:
    refs: set[str] = set()
    for sheet in workbook.get("sheets", []):
        name = str(sheet.get("name") or "")
        for cell in sheet.get("cells", []):
            refs.add(f"{name}!{cell['cell']}")
    return refs
