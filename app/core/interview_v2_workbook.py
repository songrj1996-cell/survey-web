"""Deterministic and side-effect-free physical parser for interview V2 workbooks.

This module deliberately stops at physical workbook facts and review candidates. It
does not perform HTTP work, persistence, participant binding, or model calls.
"""

from __future__ import annotations

from collections import Counter, defaultdict
from datetime import date, datetime, time
from decimal import Decimal
from hashlib import sha256
from io import BytesIO
import json
import math
from pathlib import Path
import posixpath
import re
import unicodedata
import zipfile
from xml.etree import ElementTree

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.formula import ArrayFormula, DataTableFormula

from app.core.config import (
    INTERVIEW_V2_MAX_COLUMNS_PER_SHEET,
    INTERVIEW_V2_MAX_COMPRESSION_RATIO,
    INTERVIEW_V2_MAX_FILE_BYTES,
    INTERVIEW_V2_MAX_NON_EMPTY_CELLS,
    INTERVIEW_V2_MAX_ROWS_PER_SHEET,
    INTERVIEW_V2_MAX_SHEETS,
    INTERVIEW_V2_MAX_TEXT_CHARS,
    INTERVIEW_V2_MAX_UNCOMPRESSED_BYTES,
    INTERVIEW_V2_MAX_ZIP_ENTRIES,
)


SCHEMA_VERSION = "interview-workbook-physical-snapshot/1.0"
PARSER_VERSION = "interview-v2-workbook-parser/1.0"

# These parts are materialized by openpyxl even when their definitions are not
# referenced by any worksheet cell. Keep a separate, deliberately conservative
# load-sensitive budget in addition to the package-wide ZIP limits.
MAX_STYLES_XML_BYTES = 8 * 1024 * 1024
MAX_SHARED_STRINGS_XML_BYTES = 16 * 1024 * 1024
MAX_SHARED_STRING_ITEMS = 32_768
MAX_STYLE_DEFINITIONS = {
    "font": 4_096,
    "fill": 4_096,
    "border": 4_096,
    "xf": 16_384,
    "dxf": 4_096,
    "numFmt": 4_096,
}
MAX_WORKSHEET_ROW_NODES = (
    INTERVIEW_V2_MAX_SHEETS * INTERVIEW_V2_MAX_ROWS_PER_SHEET
)
MAX_WORKSHEET_COLUMN_DEFINITION_NODES = (
    INTERVIEW_V2_MAX_SHEETS * INTERVIEW_V2_MAX_COLUMNS_PER_SHEET
)
MAX_WORKSHEET_AUXILIARY_NODES = INTERVIEW_V2_MAX_NON_EMPTY_CELLS
MAX_WORKSHEET_AUXILIARY_DEFINITIONS = {
    "conditionalFormatting": 32_768,
    "cfRule": 65_536,
    "dataValidation": 32_768,
    "hyperlink": 32_768,
    "tablePart": 4_096,
    "drawing": 4_096,
    "legacyDrawing": 4_096,
    "oleObject": 4_096,
    "objectPr": 4_096,
    "comment": 32_768,
    "table": 4_096,
    "twoCellAnchor": 16_384,
    "oneCellAnchor": 16_384,
    "absoluteAnchor": 16_384,
}
MAX_CONTENT_TYPES_XML_BYTES = 2 * 1024 * 1024
MAX_WORKBOOK_XML_BYTES = 2 * 1024 * 1024
MAX_RELATIONSHIPS_XML_BYTES = 2 * 1024 * 1024
MAX_COMMENTS_XML_BYTES = 4 * 1024 * 1024
MAX_WORKSHEET_XML_BYTES = 16 * 1024 * 1024
MAX_AUXILIARY_XML_BYTES = 8 * 1024 * 1024
MAX_CONTENT_TYPE_NODES = 8_192
MAX_RELATIONSHIP_NODES = 32_768
MAX_WORKBOOK_SHEET_NODES = INTERVIEW_V2_MAX_SHEETS
MAX_WORKSHEET_XML_NODES = INTERVIEW_V2_MAX_NON_EMPTY_CELLS * 8
MAX_AUXILIARY_XML_NODES = INTERVIEW_V2_MAX_NON_EMPTY_CELLS

_SHARED_STRINGS_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml"
)
_STYLES_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"
)
_OOXML_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"
)
_WORKBOOK_CONTENT_TYPES = {
    _OOXML_CONTENT_TYPE,
}
_MACRO_CONTENT_TYPE_MARKERS = (
    "macroenabled",
    "vba",
    "macrosheet",
    "dialogsheet",
)
_MACRO_RELATIONSHIP_SUFFIXES = (
    "/vbaproject",
    "/xlmacrosheet",
    "/intlmacrosheet",
    "/macrosheet",
    "/dialogsheet",
)
_AUXILIARY_XML_RELATIONSHIP_SUFFIXES = (
    "/comments",
    "/table",
    "/drawing",
    "/vmldrawing",
    "/chart",
    "/ctrlprop",
    "/threadedcomment",
    "/person",
)

_CFB_SIGNATURE = bytes.fromhex("D0CF11E0A1B11AE1")
_PARTICIPANT_HEADER_RE = re.compile(
    r"^(?:p(?:layer)?|participant|user|玩家|用户|受访者|访谈对象)"
    r"[\s_\-:#（）()]*[a-z0-9一二三四五六七八九十]+$",
    re.IGNORECASE,
)


class InterviewV2WorkbookError(Exception):
    """Stable workbook parsing error consumed by the import service."""

    def __init__(
        self,
        code: str,
        message: str,
        context: dict | None = None,
        retryable: bool = False,
        suggested_action: str | None = None,
    ) -> None:
        super().__init__(message)
        self.code = code
        self.message = message
        self.context = dict(context or {})
        self.retryable = bool(retryable)
        self.suggested_action = suggested_action

    def as_dict(self) -> dict:
        return {
            "code": self.code,
            "message": self.message,
            "context": self.context,
            "retryable": self.retryable,
            "suggested_action": self.suggested_action,
        }


def _error(
    code: str,
    message: str,
    *,
    context: dict | None = None,
    suggested_action: str,
) -> InterviewV2WorkbookError:
    return InterviewV2WorkbookError(
        code=code,
        message=message,
        context=context,
        retryable=False,
        suggested_action=suggested_action,
    )


def _limit_error(metric: str, actual: int | float, limit: int | float, **extra) -> None:
    context = {"metric": metric, "actual": actual, "limit": limit}
    context.update(extra)
    raise _error(
        "WORKBOOK_LIMIT_EXCEEDED",
        "工作簿超过当前安全上限，请按提示拆分或精简文件后重试。",
        context=context,
        suggested_action="simplify_workbook",
    )


def _local_name(tag: str) -> str:
    return tag.rsplit("}", 1)[-1]


def _relationship_target(base_part: str, target: str) -> str:
    target = target.replace("\\", "/")
    if target.startswith("/"):
        normalized = posixpath.normpath(target.lstrip("/"))
    else:
        normalized = posixpath.normpath(
            posixpath.join(posixpath.dirname(base_part), target)
        )
    if normalized == ".." or normalized.startswith("../"):
        raise ValueError("relationship target escapes package")
    return normalized


def _bounded_cell_range(reference: str) -> tuple[int, int, int, int]:
    min_col, min_row, max_col, max_row = range_boundaries(reference)
    if None in {min_col, min_row, max_col, max_row}:
        raise ValueError("cell reference must have bounded rows and columns")
    return int(min_col), int(min_row), int(max_col), int(max_row)


def _root_relationship_target(target: str) -> str:
    target = target.replace("\\", "/")
    normalized = posixpath.normpath(target.lstrip("/"))
    if normalized in {"", ".", ".."} or normalized.startswith("../"):
        raise ValueError("root relationship target escapes package")
    return normalized


def _relationships_part_path(source_part: str) -> str:
    return posixpath.join(
        posixpath.dirname(source_part),
        "_rels",
        posixpath.basename(source_part) + ".rels",
    )


def _parse_relationships_part(
    archive: zipfile.ZipFile,
    info_by_name: dict[str, zipfile.ZipInfo],
    rels_path: str,
    *,
    source_part: str,
    counters: dict[str, int],
) -> dict[str, dict]:
    info = info_by_name.get(rels_path)
    if info is None:
        return {}
    if info.file_size > MAX_RELATIONSHIPS_XML_BYTES:
        _limit_error(
            "relationships_xml_bytes",
            info.file_size,
            MAX_RELATIONSHIPS_XML_BYTES,
            zip_entry=rels_path,
        )
    relationships = {}
    with archive.open(info) as stream:
        for _event, node in ElementTree.iterparse(stream, events=("end",)):
            if _local_name(node.tag) != "Relationship":
                node.clear()
                continue
            counters["relationship_nodes"] += 1
            if counters["relationship_nodes"] > MAX_RELATIONSHIP_NODES:
                _limit_error(
                    "relationship_nodes",
                    counters["relationship_nodes"],
                    MAX_RELATIONSHIP_NODES,
                    zip_entry=rels_path,
                )
            rel_id = node.attrib.get("Id")
            target = node.attrib.get("Target")
            rel_type = node.attrib.get("Type", "")
            target_mode = node.attrib.get("TargetMode", "").lower()
            if not rel_id or not target or rel_id in relationships:
                raise ValueError("relationship is missing fields or duplicated")
            relationships[rel_id] = {
                "path": (
                    None
                    if target_mode == "external"
                    else (
                        _root_relationship_target(target)
                        if not source_part
                        else _relationship_target(source_part, target)
                    )
                ),
                "type": rel_type,
                "target_mode": target_mode,
            }
            node.clear()
    return relationships


def _workbook_sheet_parts(
    archive: zipfile.ZipFile,
    info_by_name: dict[str, zipfile.ZipInfo],
    *,
    workbook_path: str,
    workbook_relationships: dict[str, dict],
) -> list[dict]:
    workbook_info = info_by_name[workbook_path]
    if workbook_info.file_size > MAX_WORKBOOK_XML_BYTES:
        _limit_error(
            "workbook_xml_bytes",
            workbook_info.file_size,
            MAX_WORKBOOK_XML_BYTES,
            zip_entry=workbook_path,
        )
    parts = []
    relationship_id_key = (
        "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"
    )
    external_reference_count = 0
    with archive.open(workbook_info) as stream:
        for _event, node in ElementTree.iterparse(stream, events=("end",)):
            local_name = _local_name(node.tag)
            if local_name == "externalReference":
                external_reference_count += 1
            elif local_name == "sheet":
                if len(parts) >= MAX_WORKBOOK_SHEET_NODES:
                    _limit_error(
                        "workbook_sheet_nodes",
                        len(parts) + 1,
                        MAX_WORKBOOK_SHEET_NODES,
                        zip_entry=workbook_path,
                    )
                rel_id = node.attrib.get(relationship_id_key)
                relationship = workbook_relationships.get(rel_id or "")
                if not relationship or relationship["path"] is None:
                    raise ValueError("worksheet relationship is missing")
                rel_type = relationship["type"].rstrip("/").lower()
                if rel_type.endswith(_MACRO_RELATIONSHIP_SUFFIXES):
                    raise _error(
                        "WORKBOOK_MACROS_UNSUPPORTED",
                        "工作簿包含宏或宏工作表，请另存为不含宏的 .xlsx 后重试。",
                        context={"has_macros": True},
                        suggested_action="remove_macros",
                    )
                if not rel_type.endswith("/worksheet"):
                    raise ValueError("unsupported sheet relationship type")
                parts.append(
                    {
                        "name": node.attrib.get("name", ""),
                        "state": node.attrib.get("state", "visible"),
                        "path": relationship["path"],
                        "sheet_type": "worksheet",
                    }
                )
            node.clear()
    if external_reference_count:
        raise _error(
            "WORKBOOK_DEPENDS_ON_EXTERNAL_CONTENT",
            "工作簿包含外部链接，请将所需内容固化到工作簿内后重试。",
            context={"external_reference_count": external_reference_count},
            suggested_action="embed_external_content",
        )
    if not parts:
        raise ValueError("workbook contains no sheets")
    return parts


def _inspect_manifest(
    archive: zipfile.ZipFile,
    info_by_name: dict[str, zipfile.ZipInfo],
) -> dict:
    wanted_types = {
        _STYLES_CONTENT_TYPE: "styles",
        _SHARED_STRINGS_CONTENT_TYPE: "shared_strings",
    }
    matches: dict[str, list[str]] = {name: [] for name in wanted_types.values()}
    workbook_paths = []
    manifest_info = info_by_name["[Content_Types].xml"]
    if manifest_info.file_size > MAX_CONTENT_TYPES_XML_BYTES:
        _limit_error(
            "content_types_xml_bytes",
            manifest_info.file_size,
            MAX_CONTENT_TYPES_XML_BYTES,
            zip_entry=manifest_info.filename,
        )
    node_count = 0
    seen_overrides = set()
    with archive.open(manifest_info) as stream:
        for _event, node in ElementTree.iterparse(stream, events=("end",)):
            local_name = _local_name(node.tag)
            if local_name not in {"Override", "Default"}:
                node.clear()
                continue
            node_count += 1
            if node_count > MAX_CONTENT_TYPE_NODES:
                _limit_error(
                    "content_type_nodes",
                    node_count,
                    MAX_CONTENT_TYPE_NODES,
                    zip_entry=manifest_info.filename,
                )
            content_type = node.attrib.get("ContentType", "")
            if any(marker in content_type.lower() for marker in _MACRO_CONTENT_TYPE_MARKERS):
                raise _error(
                    "WORKBOOK_MACROS_UNSUPPORTED",
                    "工作簿包含宏或宏工作表，请另存为不含宏的 .xlsx 后重试。",
                    context={"has_macros": True},
                    suggested_action="remove_macros",
                )
            if local_name == "Override":
                raw_part_name = node.attrib.get("PartName", "")
                if not raw_part_name.startswith("/"):
                    raise ValueError("manifest override PartName must be absolute")
                normalized = posixpath.normpath(raw_part_name.lstrip("/"))
                if normalized in {"", ".", ".."} or normalized.startswith("../"):
                    raise ValueError("manifest override PartName escapes package")
                if normalized in seen_overrides or normalized not in info_by_name:
                    raise ValueError("manifest override is duplicated or missing")
                seen_overrides.add(normalized)
                key = wanted_types.get(content_type)
                if key is not None:
                    matches[key].append(normalized)
                if content_type in _WORKBOOK_CONTENT_TYPES:
                    workbook_paths.append(normalized)
            node.clear()

    if len(workbook_paths) != 1:
        raise ValueError("manifest must contain exactly one supported workbook part")

    resolved: dict[str, str | None] = {}
    for key, paths in matches.items():
        if len(paths) > 1 or len(paths) != len(set(paths)):
            raise ValueError(f"manifest contains duplicate {key} parts")
        resolved[key] = paths[0] if paths else None
    if resolved["styles"] not in {None, "xl/styles.xml"}:
        # The installed openpyxl loader reads ARC_STYLE (xl/styles.xml)
        # directly rather than following the manifest. Reject alternative
        # style PartNames so preflight and load cannot inspect different XML.
        raise ValueError("styles part must use the supported xl/styles.xml path")
    resolved["workbook"] = workbook_paths[0]
    resolved["content_type_nodes"] = node_count
    return resolved


def _inspect_load_sensitive_parts(
    archive: zipfile.ZipFile,
    info_by_name: dict[str, zipfile.ZipInfo],
    manifest_parts: dict[str, str | None],
) -> dict:
    style_counts = {name: 0 for name in MAX_STYLE_DEFINITIONS}
    styles_path = "xl/styles.xml" if "xl/styles.xml" in info_by_name else None
    styles_info = info_by_name.get(styles_path) if styles_path else None
    if styles_info is not None:
        if styles_info.file_size > MAX_STYLES_XML_BYTES:
            _limit_error(
                "styles_xml_bytes",
                styles_info.file_size,
                MAX_STYLES_XML_BYTES,
                zip_entry=styles_info.filename,
            )
        with archive.open(styles_info) as stream:
            for _event, node in ElementTree.iterparse(stream, events=("end",)):
                local_name = _local_name(node.tag)
                if local_name in style_counts:
                    style_counts[local_name] += 1
                    limit = MAX_STYLE_DEFINITIONS[local_name]
                    if style_counts[local_name] > limit:
                        _limit_error(
                            f"style_{local_name}_definitions",
                            style_counts[local_name],
                            limit,
                            zip_entry=styles_info.filename,
                        )
                node.clear()

    shared_string_items = 0
    shared_string_text_chars = 0
    shared_strings_path = manifest_parts["shared_strings"]
    shared_strings_info = (
        info_by_name.get(shared_strings_path) if shared_strings_path else None
    )
    if shared_strings_info is not None:
        if shared_strings_info.file_size > MAX_SHARED_STRINGS_XML_BYTES:
            _limit_error(
                "shared_strings_xml_bytes",
                shared_strings_info.file_size,
                MAX_SHARED_STRINGS_XML_BYTES,
                zip_entry=shared_strings_info.filename,
            )
        with archive.open(shared_strings_info) as stream:
            for _event, node in ElementTree.iterparse(stream, events=("end",)):
                local_name = _local_name(node.tag)
                if local_name == "si":
                    shared_string_items += 1
                    if shared_string_items > MAX_SHARED_STRING_ITEMS:
                        _limit_error(
                            "shared_string_items",
                            shared_string_items,
                            MAX_SHARED_STRING_ITEMS,
                            zip_entry=shared_strings_info.filename,
                        )
                elif local_name == "t" and node.text:
                    shared_string_text_chars += len(node.text)
                    if shared_string_text_chars > INTERVIEW_V2_MAX_TEXT_CHARS:
                        _limit_error(
                            "shared_string_text_chars",
                            shared_string_text_chars,
                            INTERVIEW_V2_MAX_TEXT_CHARS,
                            zip_entry=shared_strings_info.filename,
                        )
                node.clear()

    return {
        "styles_xml_bytes": styles_info.file_size if styles_info else 0,
        "styles_part_path": styles_path,
        "style_definition_counts": style_counts,
        "shared_strings_xml_bytes": (
            shared_strings_info.file_size if shared_strings_info else 0
        ),
        "shared_strings_part_path": shared_strings_path,
        "shared_string_items": shared_string_items,
        "shared_string_text_chars": shared_string_text_chars,
    }


def _inspect_worksheet_part(
    archive: zipfile.ZipFile,
    part_path: str,
    *,
    sheet_index: int,
    counters: dict[str, int],
) -> str | None:
    worksheet_info = archive.getinfo(part_path)
    if worksheet_info.file_size > MAX_WORKSHEET_XML_BYTES:
        _limit_error(
            "worksheet_xml_bytes",
            worksheet_info.file_size,
            MAX_WORKSHEET_XML_BYTES,
            zip_entry=part_path,
            sheet_index=sheet_index,
        )
    declared_range = None
    sheet_row_nodes = 0
    sheet_column_definition_nodes = 0
    last_row_index = 0
    hyperlink_materialized_cells = 0
    sheet_xml_nodes = 0
    with archive.open(part_path) as stream:
        for _event, node in ElementTree.iterparse(stream, events=("end",)):
            sheet_xml_nodes += 1
            counters["worksheet_xml_nodes"] += 1
            if sheet_xml_nodes > MAX_WORKSHEET_XML_NODES:
                _limit_error(
                    "sheet_xml_nodes",
                    sheet_xml_nodes,
                    MAX_WORKSHEET_XML_NODES,
                    zip_entry=part_path,
                    sheet_index=sheet_index,
                )
            if counters["worksheet_xml_nodes"] > MAX_WORKSHEET_XML_NODES:
                _limit_error(
                    "worksheet_xml_nodes",
                    counters["worksheet_xml_nodes"],
                    MAX_WORKSHEET_XML_NODES,
                    zip_entry=part_path,
                    sheet_index=sheet_index,
                )
            local_name = _local_name(node.tag)
            if local_name in MAX_WORKSHEET_AUXILIARY_DEFINITIONS:
                auxiliary_counts = counters["worksheet_auxiliary_node_counts"]
                auxiliary_counts[local_name] += 1
                counters["worksheet_auxiliary_nodes"] += 1
                definition_limit = MAX_WORKSHEET_AUXILIARY_DEFINITIONS[local_name]
                if auxiliary_counts[local_name] > definition_limit:
                    _limit_error(
                        f"worksheet_{local_name}_nodes",
                        auxiliary_counts[local_name],
                        definition_limit,
                        zip_entry=part_path,
                        sheet_index=sheet_index,
                    )
                if (
                    counters["worksheet_auxiliary_nodes"]
                    > MAX_WORKSHEET_AUXILIARY_NODES
                ):
                    _limit_error(
                        "worksheet_auxiliary_nodes",
                        counters["worksheet_auxiliary_nodes"],
                        MAX_WORKSHEET_AUXILIARY_NODES,
                        zip_entry=part_path,
                        sheet_index=sheet_index,
                    )
            if local_name == "dimension" and declared_range is None:
                dimension_ref = node.attrib.get("ref", "")
                min_col, min_row, max_col, max_row = _bounded_cell_range(
                    dimension_ref
                )
                if max_row > INTERVIEW_V2_MAX_ROWS_PER_SHEET:
                    _limit_error(
                        "declared_range_max_row",
                        max_row,
                        INTERVIEW_V2_MAX_ROWS_PER_SHEET,
                        sheet_index=sheet_index,
                        declared_range=dimension_ref,
                    )
                if max_col > INTERVIEW_V2_MAX_COLUMNS_PER_SHEET:
                    _limit_error(
                        "declared_range_max_column",
                        max_col,
                        INTERVIEW_V2_MAX_COLUMNS_PER_SHEET,
                        sheet_index=sheet_index,
                        declared_range=dimension_ref,
                    )
                declared_range = dimension_ref
            elif local_name == "c":
                counters["physical_cell_count"] += 1
                if counters["physical_cell_count"] > INTERVIEW_V2_MAX_NON_EMPTY_CELLS:
                    _limit_error(
                        "physical_cell_count",
                        counters["physical_cell_count"],
                        INTERVIEW_V2_MAX_NON_EMPTY_CELLS,
                        sheet_index=sheet_index,
                    )
                cell_ref = node.attrib.get("r", "")
                min_col, min_row, max_col, max_row = _bounded_cell_range(cell_ref)
                if min_col != max_col or min_row != max_row:
                    raise ValueError("physical cell reference must identify one cell")
                if max_row > INTERVIEW_V2_MAX_ROWS_PER_SHEET:
                    _limit_error(
                        "physical_cell_max_row",
                        max_row,
                        INTERVIEW_V2_MAX_ROWS_PER_SHEET,
                        sheet_index=sheet_index,
                        cell=cell_ref,
                    )
                if max_col > INTERVIEW_V2_MAX_COLUMNS_PER_SHEET:
                    _limit_error(
                        "physical_cell_max_column",
                        max_col,
                        INTERVIEW_V2_MAX_COLUMNS_PER_SHEET,
                        sheet_index=sheet_index,
                        cell=cell_ref,
                    )
            elif local_name == "mergeCell":
                merged_ref = node.attrib.get("ref")
                if not merged_ref:
                    raise ValueError("merged cell range is missing its reference")
                min_col, min_row, max_col, max_row = _bounded_cell_range(
                    merged_ref
                )
                if max_row > INTERVIEW_V2_MAX_ROWS_PER_SHEET:
                    _limit_error(
                        "merged_range_max_row",
                        max_row,
                        INTERVIEW_V2_MAX_ROWS_PER_SHEET,
                        sheet_index=sheet_index,
                        merged_range=merged_ref,
                    )
                if max_col > INTERVIEW_V2_MAX_COLUMNS_PER_SHEET:
                    _limit_error(
                        "merged_range_max_column",
                        max_col,
                        INTERVIEW_V2_MAX_COLUMNS_PER_SHEET,
                        sheet_index=sheet_index,
                        merged_range=merged_ref,
                    )
                counters["merged_cell_area"] += (max_row - min_row + 1) * (
                    max_col - min_col + 1
                )
                if counters["merged_cell_area"] > INTERVIEW_V2_MAX_NON_EMPTY_CELLS:
                    _limit_error(
                        "merged_cell_area",
                        counters["merged_cell_area"],
                        INTERVIEW_V2_MAX_NON_EMPTY_CELLS,
                        sheet_index=sheet_index,
                        merged_range=merged_ref,
                    )
            elif local_name == "row":
                sheet_row_nodes += 1
                counters["worksheet_row_nodes"] += 1
                if sheet_row_nodes > INTERVIEW_V2_MAX_ROWS_PER_SHEET:
                    _limit_error(
                        "sheet_row_nodes",
                        sheet_row_nodes,
                        INTERVIEW_V2_MAX_ROWS_PER_SHEET,
                        sheet_index=sheet_index,
                    )
                if counters["worksheet_row_nodes"] > MAX_WORKSHEET_ROW_NODES:
                    _limit_error(
                        "worksheet_row_nodes",
                        counters["worksheet_row_nodes"],
                        MAX_WORKSHEET_ROW_NODES,
                        sheet_index=sheet_index,
                    )
                row_reference = node.attrib.get("r")
                if row_reference is None:
                    row_index = last_row_index + 1
                else:
                    row_index = int(row_reference)
                    if row_index < 1:
                        raise ValueError("worksheet row index must be positive")
                last_row_index = row_index
                if row_index > INTERVIEW_V2_MAX_ROWS_PER_SHEET:
                    _limit_error(
                        "row_dimension_max_row",
                        row_index,
                        INTERVIEW_V2_MAX_ROWS_PER_SHEET,
                        sheet_index=sheet_index,
                    )
            elif local_name == "col":
                sheet_column_definition_nodes += 1
                counters["worksheet_column_definition_nodes"] += 1
                if (
                    sheet_column_definition_nodes
                    > INTERVIEW_V2_MAX_COLUMNS_PER_SHEET
                ):
                    _limit_error(
                        "sheet_column_dimension_nodes",
                        sheet_column_definition_nodes,
                        INTERVIEW_V2_MAX_COLUMNS_PER_SHEET,
                        sheet_index=sheet_index,
                    )
                if (
                    counters["worksheet_column_definition_nodes"]
                    > MAX_WORKSHEET_COLUMN_DEFINITION_NODES
                ):
                    _limit_error(
                        "worksheet_column_dimension_nodes",
                        counters["worksheet_column_definition_nodes"],
                        MAX_WORKSHEET_COLUMN_DEFINITION_NODES,
                        sheet_index=sheet_index,
                    )
                min_column = int(node.attrib.get("min", "0"))
                max_column = int(node.attrib.get("max", "0"))
                if min_column < 1 or max_column < min_column:
                    raise ValueError("worksheet column dimension is invalid")
                if min_column > INTERVIEW_V2_MAX_COLUMNS_PER_SHEET:
                    _limit_error(
                        "column_dimension_min_column",
                        min_column,
                        INTERVIEW_V2_MAX_COLUMNS_PER_SHEET,
                        sheet_index=sheet_index,
                    )
                if max_column > INTERVIEW_V2_MAX_COLUMNS_PER_SHEET:
                    _limit_error(
                        "column_dimension_max_column",
                        max_column,
                        INTERVIEW_V2_MAX_COLUMNS_PER_SHEET,
                        sheet_index=sheet_index,
                    )
            elif local_name == "hyperlink":
                hyperlink_ref = node.attrib.get("ref")
                if not hyperlink_ref:
                    raise ValueError("worksheet hyperlink is missing its reference")
                min_col, min_row, max_col, max_row = _bounded_cell_range(
                    hyperlink_ref
                )
                if max_row > INTERVIEW_V2_MAX_ROWS_PER_SHEET:
                    _limit_error(
                        "hyperlink_max_row",
                        max_row,
                        INTERVIEW_V2_MAX_ROWS_PER_SHEET,
                        sheet_index=sheet_index,
                        hyperlink_ref=hyperlink_ref,
                    )
                if max_col > INTERVIEW_V2_MAX_COLUMNS_PER_SHEET:
                    _limit_error(
                        "hyperlink_max_column",
                        max_col,
                        INTERVIEW_V2_MAX_COLUMNS_PER_SHEET,
                        sheet_index=sheet_index,
                        hyperlink_ref=hyperlink_ref,
                    )
                area = (max_row - min_row + 1) * (max_col - min_col + 1)
                hyperlink_materialized_cells += area
                counters["potential_materialized_cells"] += area
                if (
                    hyperlink_materialized_cells > INTERVIEW_V2_MAX_NON_EMPTY_CELLS
                    or counters["potential_materialized_cells"]
                    > INTERVIEW_V2_MAX_NON_EMPTY_CELLS
                ):
                    _limit_error(
                        "potential_materialized_cells",
                        counters["potential_materialized_cells"],
                        INTERVIEW_V2_MAX_NON_EMPTY_CELLS,
                        sheet_index=sheet_index,
                        hyperlink_ref=hyperlink_ref,
                    )
            node.clear()
    return declared_range


def _inspect_auxiliary_parts(
    archive: zipfile.ZipFile,
    info_by_name: dict[str, zipfile.ZipInfo],
    *,
    referenced_parts: set[str],
    counters: dict[str, object],
) -> dict:
    counts = counters["worksheet_auxiliary_node_counts"]
    auxiliary_xml_nodes = 0
    auxiliary_xml_bytes = 0
    auxiliary_xml_part_count = 0
    for name in sorted(referenced_parts):
        info = info_by_name.get(name)
        if info is None:
            raise ValueError("auxiliary relationship target is missing")
        if info.file_size > MAX_AUXILIARY_XML_BYTES:
            _limit_error(
                "auxiliary_xml_bytes",
                info.file_size,
                MAX_AUXILIARY_XML_BYTES,
                zip_entry=name,
            )
        auxiliary_xml_bytes += info.file_size
        auxiliary_xml_part_count += 1
        with archive.open(info) as stream:
            for _event, node in ElementTree.iterparse(stream, events=("end",)):
                auxiliary_xml_nodes += 1
                if auxiliary_xml_nodes > MAX_AUXILIARY_XML_NODES:
                    _limit_error(
                        "auxiliary_xml_nodes",
                        auxiliary_xml_nodes,
                        MAX_AUXILIARY_XML_NODES,
                        zip_entry=name,
                    )
                local_name = _local_name(node.tag)
                if local_name in counts:
                    counts[local_name] += 1
                    counters["worksheet_auxiliary_nodes"] += 1
                    limit = MAX_WORKSHEET_AUXILIARY_DEFINITIONS[local_name]
                    if counts[local_name] > limit:
                        _limit_error(
                            f"worksheet_{local_name}_nodes",
                            counts[local_name],
                            limit,
                            zip_entry=name,
                        )
                    if (
                        counters["worksheet_auxiliary_nodes"]
                        > MAX_WORKSHEET_AUXILIARY_NODES
                    ):
                        _limit_error(
                            "worksheet_auxiliary_nodes",
                            counters["worksheet_auxiliary_nodes"],
                            MAX_WORKSHEET_AUXILIARY_NODES,
                            zip_entry=name,
                        )
                node.clear()
    return {
        "worksheet_auxiliary_nodes": counters["worksheet_auxiliary_nodes"],
        "worksheet_auxiliary_node_counts": counts,
        "auxiliary_xml_part_count": auxiliary_xml_part_count,
        "auxiliary_xml_bytes": auxiliary_xml_bytes,
        "auxiliary_xml_nodes": auxiliary_xml_nodes,
    }


def _inspect_comment_parts(
    archive: zipfile.ZipFile,
    info_by_name: dict[str, zipfile.ZipInfo],
    *,
    comment_parts: set[str],
    counters: dict[str, int],
) -> dict:
    comment_count = 0
    for part_path in sorted(comment_parts):
        info = info_by_name.get(part_path)
        if info is None:
            raise ValueError("comment relationship target is missing")
        if info.file_size > MAX_COMMENTS_XML_BYTES:
            _limit_error(
                "comments_xml_bytes",
                info.file_size,
                MAX_COMMENTS_XML_BYTES,
                zip_entry=part_path,
            )
        with archive.open(info) as stream:
            for _event, node in ElementTree.iterparse(stream, events=("end",)):
                if _local_name(node.tag) == "comment":
                    comment_count += 1
                    if comment_count > MAX_WORKSHEET_AUXILIARY_DEFINITIONS["comment"]:
                        _limit_error(
                            "worksheet_comment_nodes",
                            comment_count,
                            MAX_WORKSHEET_AUXILIARY_DEFINITIONS["comment"],
                            zip_entry=part_path,
                        )
                    cell_ref = node.attrib.get("ref", "")
                    min_col, min_row, max_col, max_row = _bounded_cell_range(
                        cell_ref
                    )
                    if min_col != max_col or min_row != max_row:
                        raise ValueError("comment reference must identify one cell")
                    if max_row > INTERVIEW_V2_MAX_ROWS_PER_SHEET:
                        _limit_error(
                            "comment_max_row",
                            max_row,
                            INTERVIEW_V2_MAX_ROWS_PER_SHEET,
                            zip_entry=part_path,
                            comment_ref=cell_ref,
                        )
                    if max_col > INTERVIEW_V2_MAX_COLUMNS_PER_SHEET:
                        _limit_error(
                            "comment_max_column",
                            max_col,
                            INTERVIEW_V2_MAX_COLUMNS_PER_SHEET,
                            zip_entry=part_path,
                            comment_ref=cell_ref,
                        )
                    counters["potential_materialized_cells"] += 1
                    if (
                        counters["potential_materialized_cells"]
                        > INTERVIEW_V2_MAX_NON_EMPTY_CELLS
                    ):
                        _limit_error(
                            "potential_materialized_cells",
                            counters["potential_materialized_cells"],
                            INTERVIEW_V2_MAX_NON_EMPTY_CELLS,
                            zip_entry=part_path,
                            comment_ref=cell_ref,
                        )
                node.clear()
    return {"comment_count": comment_count}


def _inspect_zip(content: bytes) -> dict:
    try:
        with zipfile.ZipFile(BytesIO(content), mode="r") as archive:
            infos = archive.infolist()
            if len(infos) > INTERVIEW_V2_MAX_ZIP_ENTRIES:
                _limit_error(
                    "zip_entries", len(infos), INTERVIEW_V2_MAX_ZIP_ENTRIES
                )

            names = [info.filename for info in infos]
            if len(names) != len(set(names)):
                raise ValueError("duplicate ZIP entry")
            info_by_name = {info.filename: info for info in infos}

            total_uncompressed = 0
            maximum_ratio = 0.0
            for info in infos:
                normalized = info.filename.replace("\\", "/")
                if (
                    normalized.startswith("/")
                    or normalized == ".."
                    or normalized.startswith("../")
                    or "/../" in f"/{normalized}/"
                ):
                    raise ValueError("unsafe ZIP entry path")
                if info.flag_bits & 0x1:
                    raise _error(
                        "WORKBOOK_ENCRYPTED",
                        "工作簿已加密，请移除打开密码后重新上传。",
                        context={"container": "encrypted_zip"},
                        suggested_action="remove_password",
                    )
                total_uncompressed += info.file_size
                if total_uncompressed > INTERVIEW_V2_MAX_UNCOMPRESSED_BYTES:
                    _limit_error(
                        "zip_uncompressed_bytes",
                        total_uncompressed,
                        INTERVIEW_V2_MAX_UNCOMPRESSED_BYTES,
                    )
                if info.file_size:
                    ratio = (
                        math.inf
                        if info.compress_size == 0
                        else info.file_size / info.compress_size
                    )
                    maximum_ratio = max(maximum_ratio, ratio)
                    if ratio > INTERVIEW_V2_MAX_COMPRESSION_RATIO:
                        _limit_error(
                            "zip_entry_compression_ratio",
                            round(ratio, 4),
                            INTERVIEW_V2_MAX_COMPRESSION_RATIO,
                            zip_entry=info.filename,
                        )

            required_parts = {"[Content_Types].xml", "_rels/.rels"}
            if not required_parts.issubset(set(names)):
                raise ValueError("required OOXML parts are missing")

            lowered_names = {name.lower() for name in names}
            has_macros = (
                any(
                    name.endswith("/vbaproject.bin")
                    or "/macrosheets/" in name
                    or "/dialogsheets/" in name
                    for name in lowered_names
                )
            )
            if has_macros:
                raise _error(
                    "WORKBOOK_MACROS_UNSUPPORTED",
                    "工作簿包含宏或宏工作表，请另存为不含宏的 .xlsx 后重试。",
                    context={"has_macros": True},
                    suggested_action="remove_macros",
                )
            manifest_parts = _inspect_manifest(archive, info_by_name)
            load_sensitive_parts = _inspect_load_sensitive_parts(
                archive,
                info_by_name,
                manifest_parts,
            )
            relationship_counters = {"relationship_nodes": 0}
            parsed_relationship_parts = {}
            external_relationship_count = 0
            macro_relationship_count = 0
            comment_parts = set()
            for name in sorted(names):
                if not name.lower().endswith(".rels"):
                    continue
                if name == "_rels/.rels":
                    source_part = ""
                elif "/_rels/" in name and name.endswith(".rels"):
                    base, rel_name = name.rsplit("/_rels/", 1)
                    source_part = posixpath.join(base, rel_name[:-5])
                else:
                    raise ValueError("relationship part path is invalid")
                relationships = _parse_relationships_part(
                    archive,
                    info_by_name,
                    name,
                    source_part=source_part,
                    counters=relationship_counters,
                )
                parsed_relationship_parts[name] = relationships
                for relationship in relationships.values():
                    rel_type = relationship["type"].rstrip("/").lower()
                    if relationship["target_mode"] == "external" or rel_type.endswith(
                        "/externallink"
                    ):
                        external_relationship_count += 1
                    if rel_type.endswith(_MACRO_RELATIONSHIP_SUFFIXES):
                        macro_relationship_count += 1
                    if rel_type.endswith("/comments") and relationship["path"]:
                        comment_parts.add(relationship["path"])
            if macro_relationship_count:
                raise _error(
                    "WORKBOOK_MACROS_UNSUPPORTED",
                    "工作簿包含宏或宏工作表，请另存为不含宏的 .xlsx 后重试。",
                    context={"has_macros": True},
                    suggested_action="remove_macros",
                )
            if external_relationship_count:
                raise _error(
                    "WORKBOOK_DEPENDS_ON_EXTERNAL_CONTENT",
                    "工作簿包含外部链接，请将所需内容固化到工作簿内后重试。",
                    context={
                        "external_relationship_count": external_relationship_count
                    },
                    suggested_action="embed_external_content",
                )

            workbook_path = manifest_parts["workbook"]
            root_relationships = parsed_relationship_parts.get("_rels/.rels", {})
            office_document_targets = {
                relationship["path"]
                for relationship in root_relationships.values()
                if relationship["type"].rstrip("/").lower().endswith(
                    "/officedocument"
                )
                and relationship["path"] is not None
            }
            if office_document_targets != {workbook_path}:
                raise ValueError("root officeDocument relationship is ambiguous")
            workbook_rels_path = _relationships_part_path(workbook_path)
            workbook_relationships = parsed_relationship_parts.get(
                workbook_rels_path,
                {},
            )
            sheet_parts = _workbook_sheet_parts(
                archive,
                info_by_name,
                workbook_path=workbook_path,
                workbook_relationships=workbook_relationships,
            )
            if len(sheet_parts) > INTERVIEW_V2_MAX_SHEETS:
                _limit_error(
                    "sheet_count", len(sheet_parts), INTERVIEW_V2_MAX_SHEETS
                )

            declared_ranges: dict[str, str | None] = {}
            worksheet_counters = {
                "merged_cell_area": 0,
                "physical_cell_count": 0,
                "worksheet_row_nodes": 0,
                "worksheet_column_definition_nodes": 0,
                "potential_materialized_cells": 0,
                "worksheet_xml_nodes": 0,
                "worksheet_auxiliary_nodes": 0,
                "worksheet_auxiliary_node_counts": {
                    name: 0 for name in MAX_WORKSHEET_AUXILIARY_DEFINITIONS
                },
            }
            for sheet_index, part in enumerate(sheet_parts, start=1):
                part_path = part["path"]
                if part["sheet_type"] != "worksheet":
                    declared_ranges[part_path] = None
                    continue
                if part_path not in names:
                    raise ValueError("worksheet part is missing")
                declared_ranges[part_path] = _inspect_worksheet_part(
                    archive,
                    part_path,
                    sheet_index=sheet_index,
                    counters=worksheet_counters,
                )

            controlled_xml_parts = {
                workbook_path,
                manifest_parts["styles"],
                manifest_parts["shared_strings"],
                *(part["path"] for part in sheet_parts),
            }
            auxiliary_referenced_parts = {
                relationship["path"]
                for relationships in parsed_relationship_parts.values()
                for relationship in relationships.values()
                if relationship["path"] is not None
                and relationship["path"] not in controlled_xml_parts
                and (
                    relationship["type"].rstrip("/").lower().endswith(
                        _AUXILIARY_XML_RELATIONSHIP_SUFFIXES
                    )
                    or relationship["path"].lower().endswith((".xml", ".vml"))
                )
            }
            auxiliary_parts = _inspect_auxiliary_parts(
                archive,
                info_by_name,
                referenced_parts=auxiliary_referenced_parts,
                counters=worksheet_counters,
            )
            comment_parts_result = _inspect_comment_parts(
                archive,
                info_by_name,
                comment_parts=comment_parts,
                counters=worksheet_counters,
            )

            bad_member = archive.testzip()
            if bad_member:
                raise ValueError("ZIP CRC validation failed")

            return {
                "zip_entry_count": len(infos),
                "uncompressed_bytes": total_uncompressed,
                "max_compression_ratio": round(maximum_ratio, 4),
                "has_macros": False,
                "has_external_links": False,
                "merged_cell_area": worksheet_counters["merged_cell_area"],
                "physical_cell_count": worksheet_counters["physical_cell_count"],
                "worksheet_row_nodes": worksheet_counters["worksheet_row_nodes"],
                "worksheet_column_definition_nodes": worksheet_counters[
                    "worksheet_column_definition_nodes"
                ],
                "worksheet_xml_nodes": worksheet_counters["worksheet_xml_nodes"],
                "potential_materialized_cells": worksheet_counters[
                    "potential_materialized_cells"
                ],
                "content_type_nodes": manifest_parts["content_type_nodes"],
                "relationship_nodes": relationship_counters["relationship_nodes"],
                "workbook_part_path": workbook_path,
                **load_sensitive_parts,
                **auxiliary_parts,
                **comment_parts_result,
                "sheet_parts": sheet_parts,
                "declared_ranges": declared_ranges,
            }
    except InterviewV2WorkbookError:
        raise
    except (
        ElementTree.ParseError,
        zipfile.BadZipFile,
        KeyError,
        RuntimeError,
        UnicodeError,
        ValueError,
    ) as exc:
        raise _error(
            "WORKBOOK_CORRUPTED",
            "无法读取工作簿，请使用 Excel 修复或另存为新的 .xlsx 后重试。",
            context={"validation_stage": "ooxml_package"},
            suggested_action="repair_or_resave_workbook",
        ) from exc


def _untrimmed_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        text = "true" if value else "false"
    elif isinstance(value, (datetime, date, time)):
        text = value.isoformat()
    elif isinstance(value, float):
        text = str(value) if math.isfinite(value) else repr(value)
    else:
        text = str(value)
    return unicodedata.normalize(
        "NFC",
        text.replace("\x00", "").replace("\r\n", "\n").replace("\r", "\n"),
    )


def _normalized_text(value) -> str:
    return _untrimmed_text(value).strip()


def _cell_text_budget_chars(value) -> int:
    """Count one underlying textual representation, including outer whitespace.

    ``raw_value``, ``display_value`` and ``normalized_text`` are aliases of the
    same ordinary cell value and therefore count once. Formula text and its
    cached result are independent persisted values and each count once.
    """

    return len(_untrimmed_text(value))


def _json_value(value):
    if value is None or isinstance(value, (str, int, bool)):
        return value
    if isinstance(value, float):
        return value if math.isfinite(value) else repr(value)
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    return _normalized_text(value)


def _formula_details(value) -> tuple[str | None, str, str | None]:
    if isinstance(value, ArrayFormula):
        return value.text, "array", value.ref
    if isinstance(value, DataTableFormula):
        return None, "data_table", value.ref
    return _normalized_text(value), "normal", None


def _value_type(cell, *, formula: bool) -> str:
    if formula:
        return "formula"
    data_type = getattr(cell, "data_type", None)
    if data_type in {"s", "str", "inlineStr"}:
        return "string"
    if data_type == "b":
        return "boolean"
    if data_type == "e":
        return "error"
    if data_type == "d" or isinstance(cell.value, (datetime, date, time)):
        return "date_time"
    if data_type == "n":
        return "number"
    return "unknown"


def _color_snapshot(color) -> dict | None:
    if color is None or not getattr(color, "type", None):
        return None
    color_type = color.type
    color_value = getattr(color, color_type, None)
    if not isinstance(color_value, (str, int, bool, float)):
        color_value = None
    return {
        "type": color_type,
        "value": color_value,
        "tint": float(color.tint or 0.0),
    }


def _style_snapshot(cell) -> dict:
    """Keep deterministic layout cues without attempting Excel rendering."""

    return {
        "style_id": int(cell.style_id),
        "number_format": cell.number_format,
        "font": {
            "name": cell.font.name,
            "size": float(cell.font.sz) if cell.font.sz is not None else None,
            "bold": bool(cell.font.bold),
            "italic": bool(cell.font.italic),
            "underline": cell.font.underline,
            "color": _color_snapshot(cell.font.color),
        },
        "fill": {
            "fill_type": cell.fill.fill_type,
            "foreground_color": _color_snapshot(cell.fill.fgColor),
            "background_color": _color_snapshot(cell.fill.bgColor),
        },
        "alignment": {
            "horizontal": cell.alignment.horizontal,
            "vertical": cell.alignment.vertical,
            "wrap_text": cell.alignment.wrap_text,
            "indent": float(cell.alignment.indent or 0.0),
            "text_rotation": int(cell.alignment.textRotation or 0),
        },
    }


def _range_from_bounds(
    min_row: int | None,
    min_col: int | None,
    max_row: int | None,
    max_col: int | None,
) -> str | None:
    if None in {min_row, min_col, max_row, max_col}:
        return None
    return (
        f"{get_column_letter(min_col)}{min_row}:"
        f"{get_column_letter(max_col)}{max_row}"
    )


def _hidden_columns(worksheet) -> set[int]:
    hidden: set[int] = set()
    for dimension in worksheet.column_dimensions.values():
        if not dimension.hidden:
            continue
        start = dimension.min or 1
        end = dimension.max or start
        hidden.update(range(start, end + 1))
    return hidden


def _column_profiles(cells: list[dict], min_col: int, max_col: int) -> list[dict]:
    grouped: dict[int, list[dict]] = defaultdict(list)
    for cell in cells:
        grouped[cell["column"]].append(cell)

    profiles = []
    for column in range(min_col, max_col + 1):
        column_cells = sorted(grouped.get(column, []), key=lambda item: item["row"])
        type_counts = Counter(cell["value_type"] for cell in column_cells)
        first_cell = column_cells[0] if column_cells else None
        normalized_values = {
            cell["normalized_text"]
            for cell in column_cells
            if cell["normalized_text"]
        }
        profiles.append(
            {
                "column": column,
                "column_letter": get_column_letter(column),
                "non_empty_count": len(column_cells),
                "first_non_empty_row": first_cell["row"] if first_cell else None,
                "last_non_empty_row": column_cells[-1]["row"] if column_cells else None,
                "header_address": first_cell["address"] if first_cell else None,
                "header_value": first_cell["display_value"] if first_cell else None,
                "unique_normalized_value_count": len(normalized_values),
                "value_type_counts": dict(sorted(type_counts.items())),
                "formula_count": type_counts.get("formula", 0),
                "hidden": bool(first_cell and first_cell["column_hidden"]),
            }
        )
    return profiles


def _candidate_regions(profiles: list[dict]) -> tuple[dict | None, dict | None]:
    if not profiles:
        return None, None

    explicit = [
        profile
        for profile in profiles
        if profile["header_value"]
        and _PARTICIPANT_HEADER_RE.fullmatch(profile["header_value"].strip())
    ]
    basis = "participant_like_column_headers"
    confidence = "medium"
    selected = explicit

    if not selected:
        non_empty_blocks: list[list[dict]] = []
        current_block: list[dict] = []
        for profile in profiles:
            if profile["non_empty_count"]:
                current_block.append(profile)
            elif current_block:
                non_empty_blocks.append(current_block)
                current_block = []
        if current_block:
            non_empty_blocks.append(current_block)

        separated_candidates = []
        for block_index, block in enumerate(non_empty_blocks[1:], start=1):
            preceding_width = max(
                len(previous) for previous in non_empty_blocks[:block_index]
            )
            if len(block) >= 3 and len(block) >= preceding_width:
                separated_candidates.append(block)
        if separated_candidates:
            selected = max(
                separated_candidates,
                key=lambda block: (len(block), block[-1]["column"]),
            )
            basis = "right_side_block_after_empty_separator"
            confidence = "medium"

    if not selected:
        by_first_row: dict[int, list[dict]] = defaultdict(list)
        for profile in profiles:
            first_row = profile["first_non_empty_row"]
            if first_row is not None and profile["non_empty_count"] >= 2:
                by_first_row[first_row].append(profile)
        groups = [
            group
            for group in by_first_row.values()
            if len(group) >= 2 and min(item["column"] for item in group) > profiles[0]["column"]
        ]
        if groups:
            seed = max(
                groups,
                key=lambda group: (
                    len(group),
                    max(item["column"] for item in group),
                ),
            )
            seed_start = min(item["column"] for item in seed)
            seed_end = max(item["column"] for item in seed)
            selected = list(seed)
            for profile in profiles:
                if profile["column"] <= seed_end:
                    continue
                previous = profiles[profile["column"] - profiles[0]["column"] - 1]
                if not previous["non_empty_count"] or not profile["non_empty_count"]:
                    break
                selected.append(profile)
            selected = [
                profile for profile in selected if profile["column"] >= seed_start
            ]
            basis = "parallel_columns_with_shared_header_row_and_right_block_extension"
            confidence = "low"

    participant_region = None
    if selected:
        start = min(profile["column"] for profile in selected)
        end = max(profile["column"] for profile in selected)
        participant_region = {
            "status": "confirmation_required",
            "range": f"{get_column_letter(start)}:{get_column_letter(end)}",
            "start_column": start,
            "end_column": end,
            "candidate_columns": [
                profile["column_letter"]
                for profile in sorted(selected, key=lambda item: item["column"])
            ],
            "candidate_count": len(selected),
            "header_row": Counter(
                profile["first_non_empty_row"] for profile in selected
            ).most_common(1)[0][0],
            "basis": [basis],
            "confidence": confidence,
        }

        leading_profiles = [
            profile for profile in profiles if profile["column"] < start
        ]
        non_empty_leading = [
            profile for profile in leading_profiles if profile["non_empty_count"]
        ]
        if non_empty_leading:
            structure_end = non_empty_leading[-1]["column"]
            structure_profiles = [
                profile
                for profile in leading_profiles
                if profile["column"] <= structure_end
            ]
            structure_region = {
                "status": "confirmation_required",
                "range": (
                    f"{get_column_letter(profiles[0]['column'])}:"
                    f"{get_column_letter(structure_end)}"
                ),
                "start_column": profiles[0]["column"],
                "end_column": structure_end,
                "non_empty_columns": [
                    profile["column_letter"]
                    for profile in structure_profiles
                    if profile["non_empty_count"]
                ],
                "spacer_columns": [
                    profile["column_letter"]
                    for profile in leading_profiles
                    if not profile["non_empty_count"]
                ],
                "basis": ["columns_left_of_candidate_participant_region"],
            }
            return structure_region, participant_region
    return None, participant_region


def _parse_sheet(
    worksheet,
    cached_worksheet,
    *,
    sheet_index: int,
    declared_range: str | None,
    sheet_type: str,
    budget: dict[str, int],
) -> dict:
    sheet_id = f"sheet_{sheet_index:03d}"
    if sheet_type != "worksheet" or not hasattr(worksheet, "_cells"):
        return {
            "sheet_id": sheet_id,
            "index": sheet_index,
            "name": worksheet.title,
            "state": getattr(worksheet, "sheet_state", "visible"),
            "sheet_type": sheet_type,
            "declared_range": declared_range,
            "content_range": None,
            "dimensions": {
                "content_min_row": None,
                "content_max_row": None,
                "content_min_column": None,
                "content_max_column": None,
            },
            "hidden_rows": [],
            "hidden_columns": [],
            "merged_ranges": [],
            "non_empty_cell_count": 0,
            "text_char_count": 0,
            "formula_count": 0,
            "cells": [],
            "style_table": [],
            "column_profiles": [],
            "candidate_structure": None,
            "candidate_participant_region": None,
        }

    hidden_rows = {
        index
        for index, dimension in worksheet.row_dimensions.items()
        if dimension.hidden
    }
    hidden_columns = _hidden_columns(worksheet)
    merged_ranges_with_bounds = sorted(
        (
            (*range_boundaries(str(merged_range)), str(merged_range))
            for merged_range in worksheet.merged_cells.ranges
        ),
        key=lambda item: (item[1], item[0], item[3], item[2]),
    )
    merged_ranges = [item[4] for item in merged_ranges_with_bounds]
    merged_anchors = {
        (min_row, min_col): merged_ref
        for min_col, min_row, _max_col, _max_row, merged_ref in merged_ranges_with_bounds
    }

    cells: list[dict] = []
    min_row = min_col = max_row = max_col = None
    text_char_count = 0
    formula_count = 0
    unavailable_formula_addresses: list[str] = []
    style_table: dict[int, dict] = {}

    physical_cells = sorted(
        (
            cell
            for cell in worksheet._cells.values()
            if not isinstance(cell, MergedCell)
        ),
        key=lambda cell: (cell.row, cell.column),
    )
    for cell in physical_cells:
        formula = cell.data_type == "f" or isinstance(
            cell.value, (ArrayFormula, DataTableFormula)
        )
        formula_text = formula_kind = formula_ref = None
        cached_value = None
        formula_cache_status = "not_applicable"
        if formula:
            formula_text, formula_kind, formula_ref = _formula_details(cell.value)
            cached_cell = cached_worksheet._cells.get((cell.row, cell.column))
            cached_value = _json_value(cached_cell.value) if cached_cell else None
            formula_cache_status = (
                "available" if cached_value is not None else "unavailable"
            )
            raw_value = formula_text
            display_value = (
                _normalized_text(cached_value) if cached_value is not None else None
            )
            normalized_text = _normalized_text(formula_text)
            cell_text_chars = _cell_text_budget_chars(formula_text)
            if cached_cell is not None and cached_cell.value is not None:
                cell_text_chars += _cell_text_budget_chars(cached_cell.value)
            formula_count += 1
            if cached_value is None:
                unavailable_formula_addresses.append(cell.coordinate)
        else:
            raw_value = _json_value(cell.value)
            display_value = _normalized_text(cell.value)
            normalized_text = display_value
            cell_text_chars = _cell_text_budget_chars(cell.value)

        budget["text_chars"] += cell_text_chars
        if budget["text_chars"] > INTERVIEW_V2_MAX_TEXT_CHARS:
            _limit_error(
                "total_text_chars",
                budget["text_chars"],
                INTERVIEW_V2_MAX_TEXT_CHARS,
                sheet_id=sheet_id,
                cell=cell.coordinate,
            )
        if not formula and not normalized_text:
            continue
        budget["non_empty_cells"] += 1
        if budget["non_empty_cells"] > INTERVIEW_V2_MAX_NON_EMPTY_CELLS:
            _limit_error(
                "non_empty_cells",
                budget["non_empty_cells"],
                INTERVIEW_V2_MAX_NON_EMPTY_CELLS,
                sheet_id=sheet_id,
                cell=cell.coordinate,
            )
        value_type = _value_type(cell, formula=formula)
        min_row = cell.row if min_row is None else min(min_row, cell.row)
        max_row = cell.row if max_row is None else max(max_row, cell.row)
        min_col = cell.column if min_col is None else min(min_col, cell.column)
        max_col = cell.column if max_col is None else max(max_col, cell.column)
        text_char_count += cell_text_chars
        style_id = int(cell.style_id)
        if style_id not in style_table:
            style_table[style_id] = _style_snapshot(cell)
        cells.append(
            {
                "address": cell.coordinate,
                "row": cell.row,
                "column": cell.column,
                "raw_value": raw_value,
                "normalized_text": normalized_text,
                "display_value": display_value,
                "value_type": value_type,
                "formula_text": formula_text,
                "formula_kind": formula_kind,
                "formula_ref": formula_ref,
                "cached_value": cached_value,
                "formula_cache_status": formula_cache_status,
                "merged_range": merged_anchors.get((cell.row, cell.column)),
                "row_hidden": cell.row in hidden_rows,
                "column_hidden": cell.column in hidden_columns,
                "style_id": style_id,
                "value_sha256": sha256(normalized_text.encode("utf-8")).hexdigest(),
            }
        )

    if max_row is not None and max_row > INTERVIEW_V2_MAX_ROWS_PER_SHEET:
        _limit_error(
            "content_max_row",
            max_row,
            INTERVIEW_V2_MAX_ROWS_PER_SHEET,
            sheet_id=sheet_id,
        )
    if max_col is not None and max_col > INTERVIEW_V2_MAX_COLUMNS_PER_SHEET:
        _limit_error(
            "content_max_column",
            max_col,
            INTERVIEW_V2_MAX_COLUMNS_PER_SHEET,
            sheet_id=sheet_id,
        )

    profiles = (
        _column_profiles(cells, min_col, max_col)
        if min_col is not None and max_col is not None
        else []
    )
    candidate_structure, candidate_participant_region = _candidate_regions(profiles)
    return {
        "sheet_id": sheet_id,
        "index": sheet_index,
        "name": worksheet.title,
        "state": worksheet.sheet_state,
        "sheet_type": sheet_type,
        "declared_range": declared_range,
        "content_range": _range_from_bounds(min_row, min_col, max_row, max_col),
        "dimensions": {
            "content_min_row": min_row,
            "content_max_row": max_row,
            "content_min_column": min_col,
            "content_max_column": max_col,
        },
        "hidden_rows": sorted(hidden_rows),
        "hidden_columns": [
            get_column_letter(column) for column in sorted(hidden_columns)
        ],
        "merged_ranges": merged_ranges,
        "non_empty_cell_count": len(cells),
        "text_char_count": text_char_count,
        "formula_count": formula_count,
        "formula_cache_unavailable_addresses": unavailable_formula_addresses,
        "cells": cells,
        "style_table": [style_table[key] for key in sorted(style_table)],
        "column_profiles": profiles,
        "candidate_structure": candidate_structure,
        "candidate_participant_region": candidate_participant_region,
    }


def _load_openpyxl_workbooks(content: bytes):
    formula_workbook = cached_workbook = None
    try:
        formula_workbook = load_workbook(
            BytesIO(content),
            read_only=False,
            data_only=False,
            keep_links=False,
        )
        cached_workbook = load_workbook(
            BytesIO(content),
            read_only=False,
            data_only=True,
            keep_links=False,
        )
        return formula_workbook, cached_workbook
    except (
        InvalidFileException,
        zipfile.BadZipFile,
        ElementTree.ParseError,
        KeyError,
        OSError,
        OverflowError,
        RuntimeError,
        TypeError,
        ValueError,
    ) as exc:
        if formula_workbook is not None:
            formula_workbook.close()
        if cached_workbook is not None:
            cached_workbook.close()
        raise _error(
            "WORKBOOK_CORRUPTED",
            "无法读取工作簿，请使用 Excel 修复或另存为新的 .xlsx 后重试。",
            context={"validation_stage": "workbook_parse"},
            suggested_action="repair_or_resave_workbook",
        ) from exc


def _issue(code: str, message: str, *, context: dict, suggested_action: str) -> dict:
    return {
        "code": code,
        "message": message,
        "context": context,
        "retryable": False,
        "suggested_action": suggested_action,
    }


def _snapshot_sha256(snapshot: dict) -> str:
    hash_payload = {
        key: value
        for key, value in snapshot.items()
        if key not in {"original_filename", "snapshot_sha256"}
    }
    canonical = json.dumps(
        hash_payload,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
        allow_nan=False,
    ).encode("utf-8")
    return sha256(canonical).hexdigest()


def parse_interview_v2_workbook(filename: str, content: bytes) -> dict:
    """Validate and return a deterministic physical snapshot of one ``.xlsx``."""

    if Path(filename or "").suffix.lower() != ".xlsx":
        raise _error(
            "FILE_TYPE_UNSUPPORTED",
            "访谈报告 V2 目前仅支持未加密的 .xlsx 文件。",
            context={"received_extension": Path(filename or "").suffix.lower()},
            suggested_action="upload_xlsx",
        )
    if not isinstance(content, bytes):
        raise TypeError("content must be bytes")
    if not content:
        raise _error(
            "FILE_EMPTY",
            "上传文件为空，请重新选择有效的 .xlsx 文件。",
            context={"file_size": 0},
            suggested_action="upload_xlsx",
        )
    if len(content) > INTERVIEW_V2_MAX_FILE_BYTES:
        _limit_error("file_bytes", len(content), INTERVIEW_V2_MAX_FILE_BYTES)
    if content.startswith(_CFB_SIGNATURE):
        raise _error(
            "WORKBOOK_ENCRYPTED",
            "工作簿可能已加密，请移除打开密码并另存为 .xlsx 后重试。",
            context={"container": "ole_compound_file"},
            suggested_action="remove_password",
        )

    package = _inspect_zip(content)
    formula_workbook, cached_workbook = _load_openpyxl_workbooks(content)
    try:
        formula_sheets = list(formula_workbook._sheets)
        cached_by_title = {
            sheet.title: sheet for sheet in cached_workbook._sheets
        }
        sheet_parts = package["sheet_parts"]
        if len(formula_sheets) != len(sheet_parts):
            raise _error(
                "WORKBOOK_CORRUPTED",
                "工作簿的 Sheet 目录不一致，请另存为新的 .xlsx 后重试。",
                context={"validation_stage": "sheet_catalog"},
                suggested_action="repair_or_resave_workbook",
            )

        sheets = []
        budget = {"non_empty_cells": 0, "text_chars": 0}
        for sheet_index, (worksheet, part) in enumerate(
            zip(formula_sheets, sheet_parts), start=1
        ):
            cached_worksheet = cached_by_title.get(worksheet.title)
            if cached_worksheet is None:
                raise _error(
                    "WORKBOOK_CORRUPTED",
                    "工作簿的公式缓存视图不完整，请另存后重试。",
                    context={"sheet_index": sheet_index},
                    suggested_action="repair_or_resave_workbook",
                )
            sheets.append(
                _parse_sheet(
                    worksheet,
                    cached_worksheet,
                    sheet_index=sheet_index,
                    declared_range=package["declared_ranges"].get(part["path"]),
                    sheet_type=part["sheet_type"],
                    budget=budget,
                )
            )
    finally:
        formula_workbook.close()
        cached_workbook.close()

    total_cells = budget["non_empty_cells"]
    total_text_chars = budget["text_chars"]
    if total_cells == 0:
        raise _error(
            "WORKBOOK_STRUCTURE_MINIMUM_NOT_MET",
            "工作簿中没有可读取的访谈内容，请保留模块、问题和玩家记录后重试。",
            context={"non_empty_cell_count": 0},
            suggested_action="add_interview_structure",
        )
    has_minimum_structure = any(
        sheet.get("candidate_structure") is not None
        and sheet.get("candidate_participant_region") is not None
        and int(
            sheet["candidate_participant_region"].get("candidate_count") or 0
        )
        >= 1
        for sheet in sheets
    )
    if not has_minimum_structure:
        raise _error(
            "WORKBOOK_STRUCTURE_MINIMUM_NOT_MET",
            "工作簿无法形成模块、问题和玩家列的最低结构，请检查记录布局后重试。",
            context={
                "sheet_count": len(sheets),
                "sheets_with_participant_candidates": sum(
                    1
                    for sheet in sheets
                    if sheet.get("candidate_participant_region") is not None
                ),
            },
            suggested_action="add_interview_structure",
        )

    unavailable_formula_count = sum(
        len(sheet.get("formula_cache_unavailable_addresses", [])) for sheet in sheets
    )
    warnings = []
    if unavailable_formula_count:
        warnings.append(
            _issue(
                "FORMULA_CACHE_UNAVAILABLE",
                "部分公式没有可读缓存值；系统保留公式但不会执行公式。",
                context={"formula_count": unavailable_formula_count},
                suggested_action="open_and_save_in_excel",
            )
        )

    confirmation_required = []
    if len(sheets) > 1:
        confirmation_required.append(
            _issue(
                "GROUP_MAPPING_CONFIRMATION_REQUIRED",
                "请确认哪些 Sheet 属于同一访谈组；系统不会按 Sheet 顺序自动合并。",
                context={"sheet_ids": [sheet["sheet_id"] for sheet in sheets]},
                suggested_action="open_group_mapping",
            )
        )
    candidate_sheet_ids = [
        sheet["sheet_id"]
        for sheet in sheets
        if sheet["candidate_participant_region"] is not None
    ]
    if candidate_sheet_ids:
        confirmation_required.append(
            _issue(
                "PARTICIPANT_MAPPING_CONFIRMATION_REQUIRED",
                "请确认候选玩家列及同组玩家对应关系；候选结果不是最终玩家绑定。",
                context={"sheet_ids": candidate_sheet_ids},
                suggested_action="open_participant_mapping",
            )
        )

    snapshot = {
        "schema_version": SCHEMA_VERSION,
        "parser_version": PARSER_VERSION,
        "original_filename": Path(filename).name,
        "file_size": len(content),
        "content_sha256": sha256(content).hexdigest(),
        "snapshot_sha256": "",
        "preflight": {
            "ooxml_validated": True,
            "zip_entry_count": package["zip_entry_count"],
            "uncompressed_bytes": package["uncompressed_bytes"],
            "max_compression_ratio": package["max_compression_ratio"],
            "has_macros": package["has_macros"],
            "has_external_links": package["has_external_links"],
            "merged_cell_area": package["merged_cell_area"],
            "physical_cell_count": package["physical_cell_count"],
            "worksheet_row_nodes": package["worksheet_row_nodes"],
            "worksheet_column_definition_nodes": package[
                "worksheet_column_definition_nodes"
            ],
            "worksheet_xml_nodes": package["worksheet_xml_nodes"],
            "potential_materialized_cells": package[
                "potential_materialized_cells"
            ],
            "content_type_nodes": package["content_type_nodes"],
            "relationship_nodes": package["relationship_nodes"],
            "workbook_part_path": package["workbook_part_path"],
            "styles_xml_bytes": package["styles_xml_bytes"],
            "styles_part_path": package["styles_part_path"],
            "style_definition_counts": package["style_definition_counts"],
            "shared_strings_xml_bytes": package["shared_strings_xml_bytes"],
            "shared_strings_part_path": package["shared_strings_part_path"],
            "shared_string_items": package["shared_string_items"],
            "shared_string_text_chars": package["shared_string_text_chars"],
            "worksheet_auxiliary_nodes": package["worksheet_auxiliary_nodes"],
            "worksheet_auxiliary_node_counts": package[
                "worksheet_auxiliary_node_counts"
            ],
            "comment_count": package["comment_count"],
        },
        "summary": {
            "sheet_count": len(sheets),
            "non_empty_cell_count": total_cells,
            "total_text_chars": total_text_chars,
            "formula_count": sum(sheet["formula_count"] for sheet in sheets),
            "merged_range_count": sum(
                len(sheet["merged_ranges"]) for sheet in sheets
            ),
            "hidden_row_count": sum(len(sheet["hidden_rows"]) for sheet in sheets),
            "hidden_column_count": sum(
                len(sheet["hidden_columns"]) for sheet in sheets
            ),
        },
        "sheets": sheets,
        "warnings": warnings,
        "confirmation_required": confirmation_required,
    }
    snapshot["snapshot_sha256"] = _snapshot_sha256(snapshot)
    return snapshot


__all__ = [
    "InterviewV2WorkbookError",
    "PARSER_VERSION",
    "SCHEMA_VERSION",
    "parse_interview_v2_workbook",
]
