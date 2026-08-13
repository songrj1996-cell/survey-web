import io
import json
import unittest
from unittest.mock import patch
import zipfile

from openpyxl import Workbook
from openpyxl.styles import Font

from app.core import interview_v2_workbook as workbook_parser
from app.core.interview_v2_workbook import (
    InterviewV2WorkbookError,
    parse_interview_v2_workbook,
)


def _save(workbook: Workbook) -> bytes:
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _basic_workbook() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Recorder A"
    sheet.append(["Module", "Question", "P01", "P02"])
    sheet.append(["Profile", "Tell us about yourself", "Alpha", "Beta"])
    return _save(workbook)


def _rewrite_zip(content: bytes, replacements=None, additions=None) -> bytes:
    replacements = replacements or {}
    additions = additions or {}
    output = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(content), "r") as source:
        with zipfile.ZipFile(output, "w") as target:
            for info in source.infolist():
                payload = replacements.get(info.filename, source.read(info.filename))
                target.writestr(info, payload)
            for name, payload in additions.items():
                target.writestr(name, payload, compress_type=zipfile.ZIP_DEFLATED)
    return output.getvalue()


def _add_manifest_override(content: bytes, part_name: str, content_type: str) -> bytes:
    with zipfile.ZipFile(io.BytesIO(content), "r") as archive:
        manifest = archive.read("[Content_Types].xml")
    override = (
        f'<Override PartName="/{part_name}" ContentType="{content_type}"/>'
    ).encode("utf-8")
    manifest = manifest.replace(b"</Types>", override + b"</Types>")
    return _rewrite_zip(
        content,
        replacements={"[Content_Types].xml": manifest},
    )


def _replace_manifest_part(
    content: bytes,
    old_part_name: str,
    new_part_name: str,
    content_type: str,
) -> bytes:
    with zipfile.ZipFile(io.BytesIO(content), "r") as archive:
        manifest = archive.read("[Content_Types].xml")
    old_override = (
        f'<Override PartName="/{old_part_name}" ContentType="{content_type}"/>'
    ).encode("utf-8")
    new_override = (
        f'<Override PartName="/{new_part_name}" ContentType="{content_type}"/>'
    ).encode("utf-8")
    if old_override not in manifest:
        raise AssertionError(f"missing manifest override for {old_part_name}")
    return _rewrite_zip(
        content,
        replacements={"[Content_Types].xml": manifest.replace(old_override, new_override)},
    )


def _error_code(filename: str, content: bytes) -> InterviewV2WorkbookError:
    try:
        parse_interview_v2_workbook(filename, content)
    except InterviewV2WorkbookError as exc:
        return exc
    raise AssertionError("expected InterviewV2WorkbookError")


class InterviewV2WorkbookTests(unittest.TestCase):
    def test_physical_snapshot_preserves_order_visibility_cells_and_candidates(self):
        workbook = Workbook()
        first = workbook.active
        first.title = "Recorder A"
        first.append(["Module", "Question", "P01", "P02"])
        first["A1"].font = Font(bold=True)
        first.append(["Profile", "Background", "Alpha", "Beta"])
        first["B3"] = "=1+1"
        first.row_dimensions[2].hidden = True
        first.column_dimensions["D"].hidden = True
        first.merge_cells("A2:A3")

        second = workbook.create_sheet("Recorder B")
        second.append(["Module", "Question", "P01"])
        second.append(["Profile", "Background", "Gamma"])
        second.sheet_state = "hidden"
        content = _save(workbook)

        snapshot = parse_interview_v2_workbook("interview.xlsx", content)

        self.assertEqual(
            snapshot["schema_version"],
            "interview-workbook-physical-snapshot/1.0",
        )
        self.assertEqual(snapshot["parser_version"], "interview-v2-workbook-parser/1.0")
        self.assertRegex(snapshot["content_sha256"], r"^[0-9a-f]{64}$")
        self.assertRegex(snapshot["snapshot_sha256"], r"^[0-9a-f]{64}$")
        self.assertEqual(snapshot["summary"]["sheet_count"], 2)
        self.assertEqual(snapshot["summary"]["formula_count"], 1)
        self.assertEqual(snapshot["summary"]["merged_range_count"], 1)
        self.assertEqual(snapshot["summary"]["hidden_row_count"], 1)
        self.assertEqual(snapshot["summary"]["hidden_column_count"], 1)
        self.assertEqual(
            [(sheet["name"], sheet["state"]) for sheet in snapshot["sheets"]],
            [("Recorder A", "visible"), ("Recorder B", "hidden")],
        )

        sheet = snapshot["sheets"][0]
        self.assertEqual(sheet["declared_range"], "A1:D3")
        self.assertEqual(sheet["content_range"], "A1:D3")
        self.assertEqual(sheet["hidden_rows"], [2])
        self.assertEqual(sheet["hidden_columns"], ["D"])
        self.assertEqual(sheet["merged_ranges"], ["A2:A3"])
        self.assertEqual(sheet["candidate_structure"]["range"], "A:B")
        self.assertEqual(sheet["candidate_participant_region"]["range"], "C:D")
        self.assertEqual(
            sheet["candidate_participant_region"]["status"],
            "confirmation_required",
        )

        by_address = {cell["address"]: cell for cell in sheet["cells"]}
        self.assertEqual(by_address["A2"]["merged_range"], "A2:A3")
        styles = {style["style_id"]: style for style in sheet["style_table"]}
        self.assertTrue(styles[by_address["A1"]["style_id"]]["font"]["bold"])
        self.assertNotIn("style", by_address["A1"])
        self.assertTrue(by_address["A2"]["row_hidden"])
        self.assertTrue(by_address["D2"]["column_hidden"])
        self.assertEqual(by_address["B3"]["formula_text"], "=1+1")
        self.assertEqual(by_address["B3"]["value_type"], "formula")
        self.assertIsNone(by_address["B3"]["cached_value"])
        self.assertIsNone(by_address["B3"]["display_value"])
        self.assertEqual(by_address["B3"]["formula_cache_status"], "unavailable")
        self.assertEqual(snapshot["warnings"][0]["code"], "FORMULA_CACHE_UNAVAILABLE")
        self.assertEqual(
            {item["code"] for item in snapshot["confirmation_required"]},
            {
                "GROUP_MAPPING_CONFIRMATION_REQUIRED",
                "PARTICIPANT_MAPPING_CONFIRMATION_REQUIRED",
            },
        )

    def test_same_bytes_produce_same_content_and_snapshot_hash_after_rename(self):
        content = _basic_workbook()
        first = parse_interview_v2_workbook("first.xlsx", content)
        second = parse_interview_v2_workbook("renamed.XLSX", content)

        self.assertEqual(first["content_sha256"], second["content_sha256"])
        self.assertEqual(first["snapshot_sha256"], second["snapshot_sha256"])
        self.assertNotEqual(first["original_filename"], second["original_filename"])
        self.assertEqual(
            json.dumps(first["sheets"], sort_keys=True, ensure_ascii=False),
            json.dumps(second["sheets"], sort_keys=True, ensure_ascii=False),
        )

    def test_candidate_region_extends_to_late_sparse_columns_without_separator(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Module", "Question", None, None, None, None])
        for row in range(2, 8):
            sheet.cell(row, 1, f"M{row}")
            sheet.cell(row, 2, f"Q{row}")
        for column in range(3, 6):
            sheet.cell(4, column, f"column {column}")
            sheet.cell(6, column, "response")
        sheet.cell(6, 6, "late response")
        content = _save(workbook)

        snapshot = parse_interview_v2_workbook("late-columns.xlsx", content)
        sheet_snapshot = snapshot["sheets"][0]
        self.assertEqual(sheet_snapshot["candidate_structure"]["range"], "A:B")
        self.assertEqual(
            sheet_snapshot["candidate_participant_region"]["range"], "C:F"
        )
        self.assertEqual(
            sheet_snapshot["candidate_participant_region"]["status"],
            "confirmation_required",
        )

    def test_right_side_block_after_spacer_wins_over_left_shared_rows(self):
        workbook = Workbook()
        sheet = workbook.active
        for row in range(2, 7):
            sheet.cell(row, 1, f"M{row}")
            sheet.cell(row, 2, f"Q{row}")
            sheet.cell(row, 3, f"Type{row}")
        for column in range(5, 10):
            sheet.cell(1, column, f"candidate {column}")
            sheet.cell(4, column, "response")
        content = _save(workbook)

        snapshot = parse_interview_v2_workbook("spacer.xlsx", content)
        sheet_snapshot = snapshot["sheets"][0]
        self.assertEqual(sheet_snapshot["candidate_structure"]["range"], "A:C")
        self.assertEqual(sheet_snapshot["candidate_structure"]["spacer_columns"], ["D"])
        self.assertEqual(
            sheet_snapshot["candidate_participant_region"]["range"], "E:I"
        )

    def test_style_only_far_cell_changes_declared_not_content_range(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Module", "Question", "P01", "P02"])
        sheet.append(["Profile", "Background", "Alpha", "Beta"])
        sheet["A10"].font = Font(bold=True)
        content = _save(workbook)

        with patch.object(workbook_parser, "INTERVIEW_V2_MAX_ROWS_PER_SHEET", 10):
            snapshot = parse_interview_v2_workbook("styled.xlsx", content)

        sheet_snapshot = snapshot["sheets"][0]
        self.assertEqual(sheet_snapshot["declared_range"], "A1:D10")
        self.assertEqual(sheet_snapshot["content_range"], "A1:D2")
        self.assertEqual(sheet_snapshot["non_empty_cell_count"], 8)

    def test_values_keep_coordinates_types_and_normalized_text(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet["A1"] = "  text\r\nline  "
        sheet["B2"] = 2.5
        sheet["C3"] = False
        sheet["D1"] = "P01"
        sheet["E1"] = "P02"
        sheet["D2"] = "Alpha"
        sheet["E2"] = "Beta"
        content = _save(workbook)

        snapshot = parse_interview_v2_workbook("types.xlsx", content)
        cells = {
            cell["address"]: cell for cell in snapshot["sheets"][0]["cells"]
        }
        self.assertEqual(cells["A1"]["normalized_text"], "text\nline")
        self.assertEqual(cells["A1"]["value_type"], "string")
        self.assertEqual(cells["B2"]["raw_value"], 2.5)
        self.assertEqual(cells["B2"]["value_type"], "number")
        self.assertEqual(cells["C3"]["raw_value"], False)
        self.assertEqual(cells["C3"]["value_type"], "boolean")

    def test_non_xlsx_empty_corrupted_and_encrypted_are_stable_errors(self):
        cases = [
            ("notes.csv", b"x", "FILE_TYPE_UNSUPPORTED"),
            ("empty.xlsx", b"", "FILE_EMPTY"),
            ("broken.xlsx", b"not a zip", "WORKBOOK_CORRUPTED"),
            (
                "encrypted.xlsx",
                bytes.fromhex("D0CF11E0A1B11AE1") + b"encrypted package",
                "WORKBOOK_ENCRYPTED",
            ),
        ]
        for filename, content, expected_code in cases:
            with self.subTest(expected_code=expected_code):
                exc = _error_code(filename, content)
                self.assertEqual(exc.code, expected_code)
                self.assertTrue(exc.message)
                self.assertIsInstance(exc.context, dict)
                self.assertFalse(exc.retryable)
                self.assertTrue(exc.suggested_action)
                self.assertEqual(exc.as_dict()["code"], expected_code)

    def test_zip_that_is_not_real_ooxml_is_rejected(self):
        output = io.BytesIO()
        with zipfile.ZipFile(output, "w") as archive:
            archive.writestr("readme.txt", "not a workbook")
        exc = _error_code("fake.xlsx", output.getvalue())
        self.assertEqual(exc.code, "WORKBOOK_CORRUPTED")
        self.assertEqual(exc.context["validation_stage"], "ooxml_package")

    def test_corrupted_relationship_xml_is_wrapped_as_stable_error(self):
        content = _rewrite_zip(
            _basic_workbook(),
            replacements={"_rels/.rels": b"<Relationships><broken>"},
        )
        exc = _error_code("corrupted-xml.xlsx", content)
        self.assertEqual(exc.code, "WORKBOOK_CORRUPTED")
        self.assertEqual(exc.context["validation_stage"], "ooxml_package")

    def test_macro_and_external_relationships_are_rejected_without_execution(self):
        content = _basic_workbook()
        macro_content = _rewrite_zip(
            content,
            additions={"xl/vbaProject.bin": b"not executed"},
        )
        self.assertEqual(
            _error_code("macro.xlsx", macro_content).code,
            "WORKBOOK_MACROS_UNSUPPORTED",
        )

        with zipfile.ZipFile(io.BytesIO(content), "r") as archive:
            rels = archive.read("_rels/.rels")
        external_rel = (
            b'<Relationship Id="rExternal" '
            b'Type="http://schemas.openxmlformats.org/officeDocument/2006/'
            b'relationships/hyperlink" Target="https://example.invalid" '
            b'TargetMode="External"/>'
        )
        modified_rels = rels.replace(b"</Relationships>", external_rel + b"</Relationships>")
        external_content = _rewrite_zip(
            content,
            replacements={"_rels/.rels": modified_rels},
        )
        self.assertEqual(
            _error_code("external.xlsx", external_content).code,
            "WORKBOOK_DEPENDS_ON_EXTERNAL_CONTENT",
        )

    def test_zip_entry_resource_limits_are_enforced_before_workbook_load(self):
        content = _basic_workbook()
        with zipfile.ZipFile(io.BytesIO(content), "r") as archive:
            infos = archive.infolist()
            total_uncompressed = sum(info.file_size for info in infos)
            maximum_ratio = max(
                info.file_size / info.compress_size
                for info in infos
                if info.file_size and info.compress_size
            )

        patches = [
            ("INTERVIEW_V2_MAX_FILE_BYTES", len(content) - 1, "file_bytes"),
            ("INTERVIEW_V2_MAX_ZIP_ENTRIES", len(infos) - 1, "zip_entries"),
            (
                "INTERVIEW_V2_MAX_UNCOMPRESSED_BYTES",
                total_uncompressed - 1,
                "zip_uncompressed_bytes",
            ),
            (
                "INTERVIEW_V2_MAX_COMPRESSION_RATIO",
                maximum_ratio - 0.01,
                "zip_entry_compression_ratio",
            ),
        ]
        for setting, limit, metric in patches:
            with self.subTest(setting=setting):
                with patch.object(workbook_parser, setting, limit):
                    exc = _error_code("limits.xlsx", content)
                self.assertEqual(exc.code, "WORKBOOK_LIMIT_EXCEEDED")
                self.assertEqual(exc.context["metric"], metric)

    def test_sheet_row_column_cell_and_text_limits_are_enforced(self):
        two_sheets = Workbook()
        two_sheets.active["A1"] = "one"
        two_sheets.create_sheet()["A1"] = "two"
        two_sheet_content = _save(two_sheets)
        with patch.object(workbook_parser, "INTERVIEW_V2_MAX_SHEETS", 1):
            self.assertEqual(
                _error_code("sheets.xlsx", two_sheet_content).context["metric"],
                "sheet_count",
            )

        workbook = Workbook()
        sheet = workbook.active
        sheet["A1"] = "abcd"
        sheet["B2"] = "efgh"
        content = _save(workbook)
        settings = [
            ("INTERVIEW_V2_MAX_ROWS_PER_SHEET", 1, "declared_range_max_row"),
            (
                "INTERVIEW_V2_MAX_COLUMNS_PER_SHEET",
                1,
                "declared_range_max_column",
            ),
            ("INTERVIEW_V2_MAX_NON_EMPTY_CELLS", 1, "physical_cell_count"),
            ("INTERVIEW_V2_MAX_TEXT_CHARS", 7, "total_text_chars"),
        ]
        for setting, limit, metric in settings:
            with self.subTest(setting=setting):
                with patch.object(workbook_parser, setting, limit):
                    exc = _error_code("content-limits.xlsx", content)
                self.assertEqual(exc.code, "WORKBOOK_LIMIT_EXCEEDED")
                self.assertEqual(exc.context["metric"], metric)

    def test_formula_cache_and_untrimmed_whitespace_count_toward_text_budget(self):
        workbook = Workbook()
        workbook.active["A1"] = "=1+1"
        formula_content = _save(workbook)
        with zipfile.ZipFile(io.BytesIO(formula_content), "r") as archive:
            sheet_xml = archive.read("xl/worksheets/sheet1.xml")
        sheet_xml = sheet_xml.replace(b"<v></v>", b"<v>123456</v>")
        formula_content = _rewrite_zip(
            formula_content,
            replacements={"xl/worksheets/sheet1.xml": sheet_xml},
        )
        with patch.object(workbook_parser, "INTERVIEW_V2_MAX_TEXT_CHARS", 9):
            exc = _error_code("cached-formula.xlsx", formula_content)
        self.assertEqual(exc.code, "WORKBOOK_LIMIT_EXCEEDED")
        self.assertEqual(exc.context["metric"], "total_text_chars")

        whitespace_workbook = Workbook()
        whitespace_workbook.active["A1"] = " " * 20
        whitespace_content = _save(whitespace_workbook)
        with patch.object(workbook_parser, "INTERVIEW_V2_MAX_TEXT_CHARS", 10):
            exc = _error_code("whitespace.xlsx", whitespace_content)
        self.assertEqual(exc.code, "WORKBOOK_LIMIT_EXCEEDED")
        self.assertEqual(exc.context["metric"], "total_text_chars")

    def test_merged_area_and_physical_cells_fail_before_openpyxl_load(self):
        merged_workbook = Workbook()
        merged_workbook.active.merge_cells("A1:E5")
        merged_content = _save(merged_workbook)
        with (
            patch.object(workbook_parser, "INTERVIEW_V2_MAX_NON_EMPTY_CELLS", 24),
            patch.object(workbook_parser, "load_workbook") as mocked_load,
        ):
            exc = _error_code("merged-area.xlsx", merged_content)
        self.assertEqual(exc.context["metric"], "merged_cell_area")
        mocked_load.assert_not_called()

        styled_workbook = Workbook()
        sheet = styled_workbook.active
        for row in range(1, 6):
            sheet.cell(row, 1).font = Font(bold=True)
        styled_content = _save(styled_workbook)
        with (
            patch.object(workbook_parser, "INTERVIEW_V2_MAX_NON_EMPTY_CELLS", 4),
            patch.object(workbook_parser, "load_workbook") as mocked_load,
        ):
            exc = _error_code("styled-cells.xlsx", styled_content)
        self.assertEqual(exc.context["metric"], "physical_cell_count")
        mocked_load.assert_not_called()

    def test_unicode_namespace_prefix_cannot_bypass_worksheet_limits(self):
        namespace = (
            "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
        )
        unicode_cells_xml = (
            '<?xml version="1.0" encoding="UTF-8"?>'
            f'<表:worksheet xmlns:表="{namespace}">'
            '<表:dimension ref="A1:B1"/><表:sheetData><表:row r="1">'
            '<表:c r="A1"><表:v>1</表:v></表:c>'
            '<表:c r="B1"><表:v>2</表:v></表:c>'
            "</表:row></表:sheetData></表:worksheet>"
        ).encode("utf-8")
        content = _rewrite_zip(
            _basic_workbook(),
            replacements={"xl/worksheets/sheet1.xml": unicode_cells_xml},
        )
        package = workbook_parser._inspect_zip(content)
        self.assertEqual(
            package["declared_ranges"]["xl/worksheets/sheet1.xml"],
            "A1:B1",
        )
        self.assertEqual(package["physical_cell_count"], 2)

        with (
            patch.object(workbook_parser, "INTERVIEW_V2_MAX_NON_EMPTY_CELLS", 1),
            patch.object(workbook_parser, "load_workbook") as mocked_load,
        ):
            exc = _error_code("unicode-cells.xlsx", content)
        self.assertEqual(exc.context["metric"], "physical_cell_count")
        mocked_load.assert_not_called()

        unicode_merge_xml = (
            '<?xml version="1.0" encoding="UTF-8"?>'
            f'<表:worksheet xmlns:表="{namespace}">'
            '<表:dimension ref="A1:E5"/><表:sheetData/>'
            '<表:mergeCells count="1"><表:mergeCell ref="A1:E5"/>'
            "</表:mergeCells></表:worksheet>"
        ).encode("utf-8")
        content = _rewrite_zip(
            _basic_workbook(),
            replacements={"xl/worksheets/sheet1.xml": unicode_merge_xml},
        )
        with (
            patch.object(workbook_parser, "INTERVIEW_V2_MAX_NON_EMPTY_CELLS", 24),
            patch.object(workbook_parser, "load_workbook") as mocked_load,
        ):
            exc = _error_code("unicode-merge.xlsx", content)
        self.assertEqual(exc.context["metric"], "merged_cell_area")
        mocked_load.assert_not_called()

    def test_hidden_row_and_column_dimensions_fail_before_openpyxl_load(self):
        row_workbook = Workbook()
        row_workbook.active.row_dimensions[5001].hidden = True
        row_content = _save(row_workbook)
        with patch.object(workbook_parser, "load_workbook") as mocked_load:
            exc = _error_code("hidden-row.xlsx", row_content)
        self.assertEqual(exc.context["metric"], "row_dimension_max_row")
        self.assertEqual(exc.context["actual"], 5001)
        mocked_load.assert_not_called()

        column_workbook = Workbook()
        column_workbook.active.column_dimensions["IW"].hidden = True
        column_content = _save(column_workbook)
        with patch.object(workbook_parser, "load_workbook") as mocked_load:
            exc = _error_code("hidden-column.xlsx", column_content)
        self.assertEqual(exc.context["metric"], "column_dimension_min_column")
        self.assertEqual(exc.context["actual"], 257)
        mocked_load.assert_not_called()

    def test_many_empty_row_and_column_dimension_nodes_fail_before_load(self):
        namespace = (
            "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
        )
        row_nodes = "".join(
            f'<row r="{row}" hidden="1"/>' for row in range(1, 7)
        )
        row_xml = (
            f'<worksheet xmlns="{namespace}"><dimension ref="A1:A1"/>'
            f"<sheetData>{row_nodes}</sheetData></worksheet>"
        ).encode("utf-8")
        content = _rewrite_zip(
            _basic_workbook(),
            replacements={"xl/worksheets/sheet1.xml": row_xml},
        )
        with (
            patch.object(workbook_parser, "INTERVIEW_V2_MAX_ROWS_PER_SHEET", 5),
            patch.object(workbook_parser, "MAX_WORKSHEET_ROW_NODES", 100),
            patch.object(workbook_parser, "load_workbook") as mocked_load,
        ):
            exc = _error_code("many-empty-rows.xlsx", content)
        self.assertEqual(exc.context["metric"], "sheet_row_nodes")
        mocked_load.assert_not_called()

        column_nodes = "".join(
            f'<col min="{column}" max="{column}" hidden="1"/>'
            for column in range(1, 7)
        )
        column_xml = (
            f'<worksheet xmlns="{namespace}"><dimension ref="A1:A1"/>'
            f"<cols>{column_nodes}</cols><sheetData/></worksheet>"
        ).encode("utf-8")
        content = _rewrite_zip(
            _basic_workbook(),
            replacements={"xl/worksheets/sheet1.xml": column_xml},
        )
        with (
            patch.object(workbook_parser, "INTERVIEW_V2_MAX_COLUMNS_PER_SHEET", 5),
            patch.object(
                workbook_parser,
                "MAX_WORKSHEET_COLUMN_DEFINITION_NODES",
                100,
            ),
            patch.object(workbook_parser, "load_workbook") as mocked_load,
        ):
            exc = _error_code("many-column-definitions.xlsx", content)
        self.assertEqual(exc.context["metric"], "sheet_column_dimension_nodes")
        mocked_load.assert_not_called()

    def test_workbook_wide_row_and_column_node_budgets_fail_before_load(self):
        namespace = (
            "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
        )
        row_xml = (
            f'<worksheet xmlns="{namespace}"><dimension ref="A1:A1"/>'
            '<sheetData><row r="1"/><row r="2"/></sheetData></worksheet>'
        ).encode("utf-8")
        column_xml = (
            f'<worksheet xmlns="{namespace}"><dimension ref="A1:A1"/>'
            '<cols><col min="1" max="1"/><col min="2" max="2"/></cols>'
            "<sheetData/></worksheet>"
        ).encode("utf-8")

        workbook = Workbook()
        workbook.create_sheet()
        content = _save(workbook)
        content = _rewrite_zip(
            content,
            replacements={
                "xl/worksheets/sheet1.xml": row_xml,
                "xl/worksheets/sheet2.xml": row_xml,
            },
        )
        with (
            patch.object(workbook_parser, "MAX_WORKSHEET_ROW_NODES", 3),
            patch.object(workbook_parser, "load_workbook") as mocked_load,
        ):
            exc = _error_code("workbook-row-nodes.xlsx", content)
        self.assertEqual(exc.context["metric"], "worksheet_row_nodes")
        mocked_load.assert_not_called()

        content = _rewrite_zip(
            _save(workbook := Workbook()),
            replacements={"xl/worksheets/sheet1.xml": column_xml},
        )
        workbook.close()
        with (
            patch.object(
                workbook_parser,
                "MAX_WORKSHEET_COLUMN_DEFINITION_NODES",
                1,
            ),
            patch.object(workbook_parser, "load_workbook") as mocked_load,
        ):
            exc = _error_code("workbook-column-nodes.xlsx", content)
        self.assertEqual(
            exc.context["metric"],
            "worksheet_column_dimension_nodes",
        )
        mocked_load.assert_not_called()

    def test_unreferenced_style_and_shared_string_definitions_fail_before_load(self):
        namespace = (
            "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
        )
        style_definitions = "".join(
            f'<font><name val="unused-{index}"/></font>' for index in range(40_000)
        )
        styles_xml = (
            f'<styleSheet xmlns="{namespace}"><fonts count="40000">'
            f"{style_definitions}</fonts></styleSheet>"
        ).encode("utf-8")
        content = _rewrite_zip(
            _basic_workbook(),
            replacements={"xl/styles.xml": styles_xml},
        )
        with (
            patch.object(workbook_parser, "INTERVIEW_V2_MAX_COMPRESSION_RATIO", 1e9),
            patch.object(workbook_parser, "load_workbook") as mocked_load,
        ):
            exc = _error_code("unused-styles.xlsx", content)
        self.assertEqual(exc.context["metric"], "style_font_definitions")
        mocked_load.assert_not_called()

        shared_strings = "".join(
            f"<si><t>unused-{index}</t></si>" for index in range(40_000)
        )
        shared_strings_xml = (
            f'<sst xmlns="{namespace}" count="40000" uniqueCount="40000">'
            f"{shared_strings}</sst>"
        ).encode("utf-8")
        content = _rewrite_zip(
            _basic_workbook(),
            additions={"xl/sharedStrings.xml": shared_strings_xml},
        )
        content = _add_manifest_override(
            content,
            "xl/sharedStrings.xml",
            workbook_parser._SHARED_STRINGS_CONTENT_TYPE,
        )
        with (
            patch.object(workbook_parser, "INTERVIEW_V2_MAX_COMPRESSION_RATIO", 1e9),
            patch.object(workbook_parser, "load_workbook") as mocked_load,
        ):
            exc = _error_code("unused-shared-strings.xlsx", content)
        self.assertEqual(exc.context["metric"], "shared_string_items")
        mocked_load.assert_not_called()

    def test_load_sensitive_part_bytes_fail_before_openpyxl_load(self):
        content = _basic_workbook()
        with (
            patch.object(workbook_parser, "MAX_STYLES_XML_BYTES", 1),
            patch.object(workbook_parser, "load_workbook") as mocked_load,
        ):
            exc = _error_code("large-styles-part.xlsx", content)
        self.assertEqual(exc.context["metric"], "styles_xml_bytes")
        mocked_load.assert_not_called()

        shared_strings_xml = (
            '<sst xmlns="http://schemas.openxmlformats.org/'
            'spreadsheetml/2006/main"><si><t>x</t></si></sst>'
        ).encode("utf-8")
        content = _rewrite_zip(
            content,
            additions={"xl/sharedStrings.xml": shared_strings_xml},
        )
        content = _add_manifest_override(
            content,
            "xl/sharedStrings.xml",
            workbook_parser._SHARED_STRINGS_CONTENT_TYPE,
        )
        with (
            patch.object(workbook_parser, "MAX_SHARED_STRINGS_XML_BYTES", 1),
            patch.object(workbook_parser, "load_workbook") as mocked_load,
        ):
            exc = _error_code("large-shared-strings-part.xlsx", content)
        self.assertEqual(exc.context["metric"], "shared_strings_xml_bytes")
        mocked_load.assert_not_called()

    def test_runtime_nonempty_budget_is_cumulative_across_sheets(self):
        workbook = Workbook()
        workbook.active.append(["one", "two"])
        workbook.create_sheet().append(["three", "four"])
        content = _save(workbook)
        package = workbook_parser._inspect_zip(content)

        with (
            patch.object(workbook_parser, "INTERVIEW_V2_MAX_NON_EMPTY_CELLS", 3),
            patch.object(workbook_parser, "_inspect_zip", return_value=package),
        ):
            exc = _error_code("cross-sheet.xlsx", content)
        self.assertEqual(exc.context["metric"], "non_empty_cells")
        self.assertEqual(exc.context["sheet_id"], "sheet_002")

    def test_style_table_deduplicates_thousands_of_cell_style_payloads(self):
        workbook = Workbook()
        sheet = workbook.active
        for row in range(1, 101):
            for column in range(1, 51):
                cell = sheet.cell(row, column, "value")
                cell.font = Font(bold=True)
        sheet["B1"] = "P01"
        sheet["C1"] = "P02"
        content = _save(workbook)

        snapshot = parse_interview_v2_workbook("many-styled.xlsx", content)
        sheet_snapshot = snapshot["sheets"][0]
        self.assertEqual(sheet_snapshot["non_empty_cell_count"], 5000)
        self.assertEqual(len(sheet_snapshot["style_table"]), 1)
        self.assertEqual(
            {cell["style_id"] for cell in sheet_snapshot["cells"]},
            {sheet_snapshot["style_table"][0]["style_id"]},
        )
        self.assertTrue(
            all("style" not in cell for cell in sheet_snapshot["cells"])
        )

    def test_nonempty_workbook_without_minimum_interview_structure_is_rejected(self):
        workbook = Workbook()
        workbook.active["A1"] = "hello"
        content = _save(workbook)
        exc = _error_code("single-cell.xlsx", content)
        self.assertEqual(exc.code, "WORKBOOK_STRUCTURE_MINIMUM_NOT_MET")
        self.assertEqual(exc.context["sheets_with_participant_candidates"], 0)

    def test_auxiliary_nodes_fail_before_openpyxl_load(self):
        content = _basic_workbook()
        with zipfile.ZipFile(io.BytesIO(content), "r") as archive:
            sheet_xml = archive.read("xl/worksheets/sheet1.xml")
        nodes = (
            b'<conditionalFormatting sqref="A1"><cfRule type="expression" '
            b'priority="1"><formula>1</formula></cfRule></conditionalFormatting>'
            b'<conditionalFormatting sqref="A2"><cfRule type="expression" '
            b'priority="2"><formula>1</formula></cfRule></conditionalFormatting>'
        )
        sheet_xml = sheet_xml.replace(b"</worksheet>", nodes + b"</worksheet>")
        content = _rewrite_zip(
            content,
            replacements={"xl/worksheets/sheet1.xml": sheet_xml},
        )
        limits = dict(workbook_parser.MAX_WORKSHEET_AUXILIARY_DEFINITIONS)
        limits["conditionalFormatting"] = 1
        with (
            patch.object(
                workbook_parser,
                "MAX_WORKSHEET_AUXILIARY_DEFINITIONS",
                limits,
            ),
            patch.object(workbook_parser, "load_workbook") as mocked_load,
        ):
            exc = _error_code("many-cf.xlsx", content)
        self.assertEqual(
            exc.context["metric"],
            "worksheet_conditionalFormatting_nodes",
        )
        mocked_load.assert_not_called()

    def test_hyperlink_range_dimension_and_physical_cell_refs_are_bounded(self):
        content = _basic_workbook()
        with zipfile.ZipFile(io.BytesIO(content), "r") as archive:
            original = archive.read("xl/worksheets/sheet1.xml")
        hyperlink_xml = original.replace(
            b"</worksheet>",
            b'<hyperlinks><hyperlink ref="A1:IW5001" location="A1"/>'
            b"</hyperlinks></worksheet>",
        )
        malformed_dimension = original.replace(
            b'<dimension ref="A1:D2"/>',
            b'<dimension ref="THIS_IS_NOT_A_RANGE"/>',
        )
        far_cell = original.replace(b'r="D2"', b'r="IW5001"')
        cases = [
            (hyperlink_xml, "hyperlink_max_row"),
            (malformed_dimension, None),
            (far_cell, "physical_cell_max_row"),
        ]
        for sheet_xml, expected_metric in cases:
            with self.subTest(metric=expected_metric):
                modified = _rewrite_zip(
                    content,
                    replacements={"xl/worksheets/sheet1.xml": sheet_xml},
                )
                with patch.object(workbook_parser, "load_workbook") as mocked_load:
                    exc = _error_code("bad-ref.xlsx", modified)
                if expected_metric is None:
                    self.assertEqual(exc.code, "WORKBOOK_CORRUPTED")
                else:
                    self.assertEqual(exc.context["metric"], expected_metric)
                mocked_load.assert_not_called()

    def test_nonstandard_styles_and_shared_strings_parts_are_budgeted(self):
        content = _basic_workbook()
        with zipfile.ZipFile(io.BytesIO(content), "r") as archive:
            styles = archive.read("xl/styles.xml")
        content = _rewrite_zip(content, additions={"xl/altStyles.xml": styles})
        content = _replace_manifest_part(
            content,
            "xl/styles.xml",
            "xl/altStyles.xml",
            workbook_parser._STYLES_CONTENT_TYPE,
        )
        with patch.object(workbook_parser, "load_workbook") as mocked_load:
            exc = _error_code("alt-styles.xlsx", content)
        self.assertEqual(exc.code, "WORKBOOK_CORRUPTED")
        mocked_load.assert_not_called()

        strings = "".join(
            f"<si><t>unused-{index}</t></si>" for index in range(32_769)
        )
        strings_xml = (
            '<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
            f"{strings}</sst>"
        ).encode("utf-8")
        content = _rewrite_zip(
            _basic_workbook(),
            additions={"xl/altStrings.xml": strings_xml},
        )
        content = _add_manifest_override(
            content,
            "xl/altStrings.xml",
            workbook_parser._SHARED_STRINGS_CONTENT_TYPE,
        )
        with (
            patch.object(workbook_parser, "INTERVIEW_V2_MAX_COMPRESSION_RATIO", 1e9),
            patch.object(workbook_parser, "load_workbook") as mocked_load,
        ):
            exc = _error_code("alt-strings.xlsx", content)
        self.assertEqual(exc.context["metric"], "shared_string_items")
        mocked_load.assert_not_called()

    def test_alt_workbook_and_alt_worksheet_are_authoritative_for_preflight(self):
        content = _basic_workbook()
        with zipfile.ZipFile(io.BytesIO(content), "r") as archive:
            workbook_xml = archive.read("xl/workbook.xml")
            workbook_rels = archive.read("xl/_rels/workbook.xml.rels")
            root_rels = archive.read("_rels/.rels")
            sheet_xml = archive.read("xl/worksheets/sheet1.xml")
        alt_sheet = sheet_xml.replace(b'r="2"', b'r="5001"')
        alt_workbook_rels = workbook_rels.replace(
            b"/worksheets/sheet1.xml",
            b"/alt/sheet.xml",
        )
        content = _rewrite_zip(
            content,
            replacements={
                "_rels/.rels": root_rels.replace(
                    b"xl/workbook.xml",
                    b"alt/workbook.xml",
                )
            },
            additions={
                "alt/workbook.xml": workbook_xml,
                "alt/_rels/workbook.xml.rels": alt_workbook_rels,
                "xl/alt/sheet.xml": alt_sheet,
            },
        )
        content = _replace_manifest_part(
            content,
            "xl/workbook.xml",
            "alt/workbook.xml",
            workbook_parser._OOXML_CONTENT_TYPE,
        )
        with patch.object(workbook_parser, "load_workbook") as mocked_load:
            exc = _error_code("alt-workbook.xlsx", content)
        self.assertIn(
            exc.context["metric"],
            {"declared_range_max_row", "physical_cell_max_row", "row_dimension_max_row"},
        )
        mocked_load.assert_not_called()

        content = _basic_workbook()
        with zipfile.ZipFile(io.BytesIO(content), "r") as archive:
            workbook_rels = archive.read("xl/_rels/workbook.xml.rels")
            sheet_xml = archive.read("xl/worksheets/sheet1.xml")
        workbook_rels = workbook_rels.replace(
            b"/worksheets/sheet1.xml",
            b"/alt/sheet.xml",
        )
        content = _rewrite_zip(
            content,
            replacements={"xl/_rels/workbook.xml.rels": workbook_rels},
            additions={"xl/alt/sheet.xml": sheet_xml},
        )
        with patch.object(workbook_parser, "MAX_WORKSHEET_XML_BYTES", 1), patch.object(
            workbook_parser,
            "load_workbook",
        ) as mocked_load:
            exc = _error_code("alt-sheet.xlsx", content)
        self.assertEqual(exc.context["metric"], "worksheet_xml_bytes")
        mocked_load.assert_not_called()

    def test_empty_workbook_fails_minimum_structure_without_persisting_anything(self):
        content = _save(Workbook())
        exc = _error_code("blank.xlsx", content)
        self.assertEqual(exc.code, "WORKBOOK_STRUCTURE_MINIMUM_NOT_MET")
        self.assertEqual(exc.context, {"non_empty_cell_count": 0})


if __name__ == "__main__":
    unittest.main()
