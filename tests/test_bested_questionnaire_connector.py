"""Bested local-upload connector boundary tests."""

import base64
import hashlib
import io
import unittest
import warnings
import zipfile
from unittest.mock import patch

import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import (
    AnchorMarker,
    TwoCellAnchor,
)

from app.integrations.bested_questionnaire_client import (
    BestedQuestionnaireMediaIssue,
    BestedQuestionnaireParseResult,
    parse_bested_questionnaire,
    parse_bested_questionnaire_upload,
)
from app.integrations import bested_questionnaire_client as bested_client


def _questionnaire_bytes() -> bytes:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "问卷内容"
    for row in (
        ("题号", "题目"),
        ("Q1[单选题]", "是否满意"),
        ("选项", ""),
        ("1", "是"),
        ("2", "否"),
    ):
        worksheet.append(row)
    stream = io.BytesIO()
    workbook.save(stream)
    workbook.close()
    return stream.getvalue()


def _questionnaire_with_media_bytes() -> bytes:
    png = base64.b64decode(
        "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwC"
        "AAAAC0lEQVR42mNk+A8AAQUBAScY42YAAAAASUVORK5CYII="
    )
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "问卷内容"
    for row in (
        ("题号", "题目"),
        ("Q1[单选题]", "是否满意"),
        ("选项", ""),
        ("1", "是"),
        ("2", "否"),
        ("媒体链接", "演示视频"),
        ("Q2[填空题]", "其他建议"),
        ("外部文件", "演示文件"),
    ):
        worksheet.append(row)
    worksheet["B1"].hyperlink = "https://example.test/instructions"
    worksheet["B6"].hyperlink = "https://example.test/video"
    worksheet["B8"].hyperlink = "https://example.test/document"
    worksheet.add_image(Image(io.BytesIO(png)), "C3")
    worksheet.add_image(Image(io.BytesIO(png)), "C20")
    cross_question = Image(io.BytesIO(png))
    cross_question.anchor = TwoCellAnchor(
        _from=AnchorMarker(col=2, row=4),
        to=AnchorMarker(col=3, row=6),
    )
    worksheet.add_image(cross_question)
    stream = io.BytesIO()
    workbook.save(stream)
    workbook.close()
    return stream.getvalue()


def _rewrite_xlsx(
    content: bytes,
    *,
    omit: tuple[str, ...] = (),
    additions: tuple[tuple[str, bytes], ...] = (),
) -> bytes:
    output = io.BytesIO()
    with (
        zipfile.ZipFile(io.BytesIO(content), "r") as source,
        zipfile.ZipFile(output, "w", compression=zipfile.ZIP_DEFLATED) as target,
        warnings.catch_warnings(),
    ):
        warnings.simplefilter("ignore", UserWarning)
        for info in source.infolist():
            if info.filename not in omit:
                target.writestr(info.filename, source.read(info))
        for name, payload in additions:
            target.writestr(name, payload)
    return output.getvalue()


def _mark_first_central_member_encrypted(content: bytes) -> bytes:
    changed = bytearray(content)
    header = changed.find(b"PK\x01\x02")
    if header < 0:
        raise AssertionError("test fixture has no central-directory member")
    flag_offset = header + 8
    flags = int.from_bytes(changed[flag_offset:flag_offset + 2], "little")
    changed[flag_offset:flag_offset + 2] = (flags | 0x1).to_bytes(2, "little")
    return bytes(changed)


def _replace_xlsx_members(
    content: bytes,
    replacements: dict[str, bytes],
) -> bytes:
    output = io.BytesIO()
    with (
        zipfile.ZipFile(io.BytesIO(content), "r") as source,
        zipfile.ZipFile(output, "w", compression=zipfile.ZIP_DEFLATED) as target,
    ):
        for info in source.infolist():
            target.writestr(
                info.filename,
                replacements.get(info.filename, source.read(info)),
            )
    return output.getvalue()


def _move_drawing_images_to_nonstandard_directory(content: bytes) -> bytes:
    replacements: dict[str, bytes] = {}
    additions: list[tuple[str, bytes]] = []
    with zipfile.ZipFile(io.BytesIO(content), "r") as archive:
        rels_name = "xl/drawings/_rels/drawing1.xml.rels"
        rels = archive.read(rels_name).replace(
            b"/xl/media/",
            b"/xl/pictures/",
        )
        replacements[rels_name] = rels
        for name in archive.namelist():
            if name.startswith("xl/media/"):
                additions.append((
                    name.replace("xl/media/", "xl/pictures/"),
                    archive.read(name),
                ))
    return _rewrite_xlsx(
        _replace_xlsx_members(content, replacements),
        omit=tuple(
            f"xl/media/image{index}.png" for index in range(1, 4)
        ),
        additions=tuple(additions),
    )


def _worksheet_xml_replaced(content: bytes, old: bytes, new: bytes) -> bytes:
    name = "xl/worksheets/sheet1.xml"
    with zipfile.ZipFile(io.BytesIO(content), "r") as archive:
        worksheet = archive.read(name)
    if old not in worksheet:
        raise AssertionError("worksheet mutation target is missing")
    return _replace_xlsx_members(content, {name: worksheet.replace(old, new, 1)})


class BestedQuestionnaireConnectorTests(unittest.TestCase):
    def test_local_upload_calls_public_parser_and_returns_its_result(self):
        expected = BestedQuestionnaireParseResult(
            content_sha256="0" * 64,
            sheet_name="问卷内容",
            provider_rows=(),
            questions=(),
            questionnaire_text="",
        )
        content = b"uploaded workbook bytes"

        with patch(
            "app.integrations.bested_questionnaire_client.parse_bested_questionnaire",
            return_value=expected,
        ) as parser:
            actual = parse_bested_questionnaire_upload(
                "questionnaire.xlsx",
                content,
            )

        self.assertIs(actual, expected)
        parser.assert_called_once_with(content)

    def test_xlsx_filename_check_is_case_insensitive(self):
        content = _questionnaire_bytes()
        parsed = parse_bested_questionnaire_upload(
            "questionnaire.XLSX",
            content,
        )

        self.assertEqual(parsed.content_sha256, hashlib.sha256(content).hexdigest())
        self.assertEqual(parsed.questions[0].qid, 1)
        self.assertEqual(parsed.questions[0].options, ("是", "否"))

    def test_xls_is_rejected_before_the_parser_without_changing_byte_parser(self):
        content = _questionnaire_bytes()
        with patch(
            "app.integrations.bested_questionnaire_client."
            "parse_bested_questionnaire",
        ) as parser:
            with self.assertRaisesRegex(ValueError, "仅支持 .xlsx"):
                parse_bested_questionnaire_upload("questionnaire.xls", content)
        parser.assert_not_called()

        parsed = parse_bested_questionnaire(content, discover_media=False)
        self.assertEqual(parsed.questions[0].qid, 1)
        self.assertEqual(parsed.images, ())
        self.assertEqual(parsed.hyperlinks, ())
        self.assertEqual(parsed.media_issues, ())

    def test_embedded_images_and_cell_links_keep_local_source_evidence(self):
        parsed = parse_bested_questionnaire_upload(
            "questionnaire.xlsx",
            _questionnaire_with_media_bytes(),
        )

        self.assertEqual(len(parsed.images), 3)
        images = {image.source_cell: image for image in parsed.images}
        image = images["C3"]
        self.assertTrue(image.content.startswith(b"\x89PNG\r\n\x1a\n"))
        self.assertEqual(image.mime_type, "image/png")
        self.assertEqual(image.sheet_name, "问卷内容")
        self.assertEqual(image.source_cell, "C3")
        self.assertEqual(image.source_row, 3)
        self.assertEqual(image.coverage, "C3")
        self.assertEqual(image.question_qid, 1)
        self.assertEqual(images["C20"].question_qid, None)
        self.assertEqual(images["C5"].coverage, "C5:D7")
        self.assertEqual(images["C5"].question_qid, None)

        links = {link.source_cell: link for link in parsed.hyperlinks}
        self.assertEqual(set(links), {"B1", "B6", "B8"})
        self.assertEqual(links["B1"].question_qid, None)
        self.assertEqual(links["B6"].question_qid, 1)
        self.assertEqual(links["B8"].question_qid, 2)
        self.assertEqual(links["B6"].display_text, "演示视频")
        self.assertEqual(links["B6"].url, "https://example.test/video")
        self.assertEqual(parsed.media_issues, ())

    def test_one_image_extraction_failure_keeps_structure_and_other_media(self):
        content = _questionnaire_with_media_bytes()
        original_image_data = Image._data

        def extract_or_fail(image):
            source_cell, _, _, _ = bested_client._image_location(image)
            if source_cell == "C3":
                raise RuntimeError("secret image decoder detail")
            return original_image_data(image)

        with patch.object(Image, "_data", autospec=True, side_effect=extract_or_fail):
            parsed = parse_bested_questionnaire_upload(
                "questionnaire.xlsx",
                content,
            )

        self.assertEqual([question.qid for question in parsed.questions], [1, 2])
        self.assertIn("是否满意", parsed.questionnaire_text)
        self.assertEqual(
            {image.source_cell for image in parsed.images},
            {"C5", "C20"},
        )
        self.assertEqual(parsed.media_issues, (
            BestedQuestionnaireMediaIssue(
                code="image_extraction_failed",
                sheet_name="问卷内容",
                source_cell="C3",
                source_row=3,
            ),
        ))
        self.assertTrue(BestedQuestionnaireMediaIssue.__dataclass_params__.frozen)
        self.assertNotIn("secret image decoder detail", repr(parsed.media_issues))

    def test_corrupt_declared_image_is_reported_instead_of_silently_dropped(self):
        content = _questionnaire_with_media_bytes()
        with zipfile.ZipFile(io.BytesIO(content), "r") as archive:
            image_name = next(
                name for name in archive.namelist()
                if name.startswith("xl/media/")
            )
        corrupted = _replace_xlsx_members(
            content,
            {image_name: b"not a valid image"},
        )

        with warnings.catch_warnings(record=True) as emitted:
            parsed = parse_bested_questionnaire_upload(
                "questionnaire.xlsx",
                corrupted,
            )

        self.assertEqual([question.qid for question in parsed.questions], [1, 2])
        self.assertEqual(len(parsed.images), 2)
        self.assertEqual(
            tuple(issue.code for issue in parsed.media_issues),
            ("image_loading_failed",),
        )
        self.assertEqual(emitted, [])

    def test_media_workbook_load_failure_becomes_one_safe_issue(self):
        content = _questionnaire_with_media_bytes()
        original_loader = openpyxl.load_workbook
        load_count = 0

        def load_structure_only(*args, **kwargs):
            nonlocal load_count
            load_count += 1
            if load_count == 2:
                raise RuntimeError("secret media workbook detail")
            return original_loader(*args, **kwargs)

        with patch.object(openpyxl, "load_workbook", side_effect=load_structure_only):
            parsed = parse_bested_questionnaire_upload(
                "questionnaire.xlsx",
                content,
            )

        self.assertEqual([question.qid for question in parsed.questions], [1, 2])
        self.assertIn("其他建议", parsed.questionnaire_text)
        self.assertEqual(parsed.images, ())
        self.assertEqual(parsed.hyperlinks, ())
        self.assertEqual(parsed.media_issues, (
            BestedQuestionnaireMediaIssue(
                code="media_workbook_load_failed",
                sheet_name="问卷内容",
            ),
        ))
        self.assertNotIn("secret media workbook detail", repr(parsed.media_issues))

    def test_resource_exhaustion_during_media_load_remains_a_hard_failure(self):
        content = _questionnaire_with_media_bytes()
        original_loader = openpyxl.load_workbook
        load_count = 0

        def exhaust_on_media_load(*args, **kwargs):
            nonlocal load_count
            load_count += 1
            if load_count == 2:
                raise MemoryError("synthetic resource exhaustion")
            return original_loader(*args, **kwargs)

        with (
            patch.object(
                openpyxl,
                "load_workbook",
                side_effect=exhaust_on_media_load,
            ),
            self.assertRaises(MemoryError),
        ):
            parse_bested_questionnaire_upload(
                "questionnaire.xlsx",
                content,
            )

    def test_second_media_preflight_failure_remains_a_hard_failure(self):
        content = _questionnaire_with_media_bytes()
        validator = bested_client._validate_xlsx_package
        validation_count = 0

        def validate_then_fail(payload):
            nonlocal validation_count
            validation_count += 1
            if validation_count == 2:
                raise ValueError("blocked media preflight")
            return validator(payload)

        with (
            patch.object(
                bested_client,
                "_validate_xlsx_package",
                side_effect=validate_then_fail,
            ),
            self.assertRaisesRegex(ValueError, "blocked media preflight"),
        ):
            parse_bested_questionnaire_upload("questionnaire.xlsx", content)

    def test_online_urls_are_explicitly_rejected(self):
        for source in (
            "https://example.test/questionnaire.xlsx",
            "HTTP://example.test/questionnaire.xls",
            "//example.test/questionnaire.xlsx",
            "file:///tmp/questionnaire.xlsx",
        ):
            with self.subTest(source=source):
                with self.assertRaisesRegex(ValueError, "URL/页面抓取不受支持"):
                    parse_bested_questionnaire_upload(source, b"not fetched")

    def test_upload_boundary_rejects_wrong_extension_and_non_bytes(self):
        with self.assertRaisesRegex(ValueError, "仅支持 .xlsx"):
            parse_bested_questionnaire_upload("questionnaire.csv", b"content")
        with self.assertRaisesRegex(TypeError, "必须是 bytes"):
            parse_bested_questionnaire_upload("questionnaire.xlsx", "content")
        with self.assertRaisesRegex(ValueError, "上传内容为空"):
            parse_bested_questionnaire_upload("questionnaire.xlsx", b"")

    def test_security_preflight_runs_before_each_openpyxl_load(self):
        with patch.object(openpyxl, "load_workbook") as loader:
            with self.assertRaisesRegex(ValueError, "不是有效"):
                parse_bested_questionnaire_upload(
                    "questionnaire.xlsx",
                    b"not a zip",
                )
        loader.assert_not_called()

        validator = bested_client._validate_xlsx_package
        with patch.object(
            bested_client,
            "_validate_xlsx_package",
            wraps=validator,
        ) as checked:
            parse_bested_questionnaire_upload(
                "questionnaire.xlsx",
                _questionnaire_with_media_bytes(),
            )
        self.assertEqual(checked.call_count, 2)

    def test_security_preflight_rejects_declared_zip_limits(self):
        content = _questionnaire_bytes()
        with zipfile.ZipFile(io.BytesIO(content), "r") as archive:
            infos = archive.infolist()
            cases = (
                (
                    "_XLSX_MAX_ARCHIVE_BYTES",
                    len(content) - 1,
                    "压缩包超过",
                ),
                (
                    "_XLSX_MAX_MEMBERS",
                    len(infos) - 1,
                    "成员数量",
                ),
                (
                    "_XLSX_MAX_MEMBER_BYTES",
                    max(info.file_size for info in infos) - 1,
                    "单个成员解压",
                ),
                (
                    "_XLSX_MAX_TOTAL_UNCOMPRESSED_BYTES",
                    sum(info.file_size for info in infos) - 1,
                    "解压总量",
                ),
            )

        for constant, limit, message in cases:
            with self.subTest(constant=constant):
                with (
                    patch.object(bested_client, constant, limit),
                    self.assertRaisesRegex(ValueError, message),
                ):
                    parse_bested_questionnaire_upload(
                        "questionnaire.xlsx",
                        content,
                    )

    def test_security_preflight_rejects_zip_bomb_ratio(self):
        bomb = _rewrite_xlsx(
            _questionnaire_bytes(),
            additions=(("xl/worksheets/bomb.xml", b"A" * 1024 * 1024),),
        )

        with self.assertRaisesRegex(ValueError, "压缩率"):
            parse_bested_questionnaire_upload("questionnaire.xlsx", bomb)

    def test_security_preflight_rejects_duplicate_and_unsafe_paths(self):
        content = _questionnaire_bytes()
        with zipfile.ZipFile(io.BytesIO(content), "r") as archive:
            duplicate_payload = archive.read("[Content_Types].xml")
        cases = (
            (
                _rewrite_xlsx(
                    content,
                    additions=(("[Content_Types].xml", duplicate_payload),),
                ),
                "重复成员",
            ),
            (
                _rewrite_xlsx(
                    content,
                    additions=(("../escape.xml", b"escape"),),
                ),
                "不安全路径",
            ),
        )
        for package, message in cases:
            with self.subTest(message=message):
                with self.assertRaisesRegex(ValueError, message):
                    parse_bested_questionnaire_upload(
                        "questionnaire.xlsx",
                        package,
                    )

    def test_security_preflight_allows_empty_safe_directory_members(self):
        package = _rewrite_xlsx(
            _questionnaire_bytes(),
            additions=(
                ("_rels/", b""),
                ("docProps/", b""),
                ("xl/", b""),
                ("xl/worksheets/", b""),
            ),
        )

        parsed = parse_bested_questionnaire_upload(
            "questionnaire.xlsx",
            package,
        )

        self.assertEqual(len(parsed.questions), 1)

    def test_security_preflight_rejects_unsafe_directory_members(self):
        cases = (
            ("../", b"", "不安全路径"),
            ("xl/../", b"", "不安全路径"),
            ("xl/suspicious/", b"content", "目录成员必须为空"),
        )
        for name, payload, message in cases:
            with self.subTest(name=name):
                package = _rewrite_xlsx(
                    _questionnaire_bytes(),
                    additions=((name, payload),),
                )
                with self.assertRaisesRegex(ValueError, message):
                    parse_bested_questionnaire_upload(
                        "questionnaire.xlsx",
                        package,
                    )

    def test_security_preflight_rejects_encryption_and_missing_ooxml(self):
        cases = (
            (
                _mark_first_central_member_encrypted(
                    _questionnaire_bytes()
                ),
                "不允许加密成员",
            ),
            (
                _rewrite_xlsx(
                    _questionnaire_bytes(),
                    omit=("xl/workbook.xml",),
                ),
                "缺少 OOXML 关键成员",
            ),
        )
        for package, message in cases:
            with self.subTest(message=message):
                with self.assertRaisesRegex(ValueError, message):
                    parse_bested_questionnaire_upload(
                        "questionnaire.xlsx",
                        package,
                    )

    def test_security_preflight_limits_image_count_and_total_bytes(self):
        content = _questionnaire_with_media_bytes()
        cases = (
            ("_XLSX_MAX_IMAGES", 2, "图片使用数"),
            ("_XLSX_MAX_TOTAL_IMAGE_BYTES", 100, "图片总字节"),
        )
        for constant, limit, message in cases:
            with self.subTest(constant=constant):
                with (
                    patch.object(bested_client, constant, limit),
                    self.assertRaisesRegex(ValueError, message),
                ):
                    parse_bested_questionnaire_upload(
                        "questionnaire.xlsx",
                        content,
                    )

    def test_media_discovery_rechecks_actual_image_limits(self):
        content = _questionnaire_with_media_bytes()
        cases = (
            ("_XLSX_MAX_IMAGES", 2, "工作表图片数量"),
            ("_XLSX_MAX_TOTAL_IMAGE_BYTES", 60, "工作表图片总字节"),
        )
        for constant, limit, message in cases:
            with self.subTest(constant=constant):
                with (
                    patch.object(
                        bested_client,
                        "_validate_xlsx_package",
                        return_value=None,
                    ),
                    patch.object(
                        bested_client,
                        "_validated_worksheet_image_use_count",
                        return_value=3,
                    ),
                    patch.object(bested_client, constant, limit),
                    self.assertRaisesRegex(ValueError, message),
                ):
                    parse_bested_questionnaire_upload(
                        "questionnaire.xlsx",
                        content,
                    )

    def test_nonstandard_drawing_image_targets_are_limited_before_openpyxl(self):
        content = _move_drawing_images_to_nonstandard_directory(
            _questionnaire_with_media_bytes()
        )
        cases = (
            ("_XLSX_MAX_IMAGES", 2, "drawing 图片使用数"),
            ("_XLSX_MAX_TOTAL_IMAGE_BYTES", 100, "drawing 图片总字节"),
        )
        for constant, limit, message in cases:
            with self.subTest(constant=constant):
                with (
                    patch.object(bested_client, constant, limit),
                    patch.object(openpyxl, "load_workbook") as loader,
                    self.assertRaisesRegex(ValueError, message),
                ):
                    parse_bested_questionnaire_upload(
                        "questionnaire.xlsx",
                        content,
                    )
                loader.assert_not_called()

    def test_reused_image_target_counts_anchors_but_deduplicates_bytes(self):
        content = _questionnaire_with_media_bytes()
        rels_name = "xl/drawings/_rels/drawing1.xml.rels"
        with zipfile.ZipFile(io.BytesIO(content), "r") as archive:
            rels = archive.read(rels_name)
        rels = rels.replace(
            b"/xl/media/image2.png",
            b"/xl/media/image1.png",
        ).replace(
            b"/xl/media/image3.png",
            b"/xl/media/image1.png",
        )
        reused = _replace_xlsx_members(content, {rels_name: rels})

        with patch.object(bested_client, "_XLSX_MAX_TOTAL_IMAGE_BYTES", 68):
            parsed = parse_bested_questionnaire_upload(
                "questionnaire.xlsx",
                reused,
            )
        self.assertEqual(len(parsed.images), 3)

        with (
            patch.object(bested_client, "_XLSX_MAX_IMAGES", 2),
            patch.object(openpyxl, "load_workbook") as loader,
            self.assertRaisesRegex(ValueError, "图片使用数"),
        ):
            parse_bested_questionnaire_upload("questionnaire.xlsx", reused)
        loader.assert_not_called()

    def test_non_image_drawing_anchors_are_limited_before_openpyxl(self):
        content = _questionnaire_with_media_bytes()
        drawing_name = "xl/drawings/drawing1.xml"
        with zipfile.ZipFile(io.BytesIO(content), "r") as archive:
            drawing = archive.read(drawing_name)
        extra_anchor = (
            b"<oneCellAnchor><from><col>0</col><colOff>0</colOff>"
            b"<row>0</row><rowOff>0</rowOff></from><ext cx=\"1\" cy=\"1\"/>"
            b"<clientData/></oneCellAnchor>"
        )
        drawing = drawing.replace(b"</wsDr>", extra_anchor + b"</wsDr>")
        content = _replace_xlsx_members(content, {drawing_name: drawing})

        with (
            patch.object(bested_client, "_XLSX_MAX_DRAWING_ANCHORS", 3),
            patch.object(openpyxl, "load_workbook") as loader,
            self.assertRaisesRegex(ValueError, "drawing anchor 数量"),
        ):
            parse_bested_questionnaire_upload("questionnaire.xlsx", content)
        loader.assert_not_called()

    def test_invalid_drawing_image_relationships_fail_before_openpyxl(self):
        content = _questionnaire_with_media_bytes()
        rels_name = "xl/drawings/_rels/drawing1.xml.rels"
        with zipfile.ZipFile(io.BytesIO(content), "r") as archive:
            rels = archive.read(rels_name)
        cases = (
            (
                rels.replace(
                    b'Target="/xl/media/image1.png"',
                    b'Target="../../../escape.png"',
                ),
                "目标越界",
            ),
            (
                rels.replace(
                    b'Target="/xl/media/image1.png"',
                    b'Target="/xl/pictures/missing.png"',
                ),
                "目标不存在",
            ),
            (
                rels.replace(
                    b'Id="rId2"',
                    b'Id="rId1"',
                ),
                "ID 缺失或重复",
            ),
        )
        for replacement, message in cases:
            with self.subTest(message=message):
                malicious = _replace_xlsx_members(
                    content,
                    {rels_name: replacement},
                )
                with (
                    patch.object(openpyxl, "load_workbook") as loader,
                    self.assertRaisesRegex(ValueError, message),
                ):
                    parse_bested_questionnaire_upload(
                        "questionnaire.xlsx",
                        malicious,
                    )
                loader.assert_not_called()

    def test_oversized_sheet_dimension_is_rejected_before_openpyxl(self):
        content = _worksheet_xml_replaced(
            _questionnaire_bytes(),
            b'<dimension ref="A1:B5"/>',
            b'<dimension ref="A1:XFD1048576"/>',
        )

        with (
            patch.object(openpyxl, "load_workbook") as loader,
            self.assertRaisesRegex(ValueError, "dimension .*\u5b89\u5168\u4e0a\u9650"),
        ):
            parse_bested_questionnaire_upload("questionnaire.xlsx", content)
        loader.assert_not_called()

    def test_oversized_sheet_cell_coordinate_is_rejected_before_openpyxl(self):
        content = _worksheet_xml_replaced(
            _questionnaire_bytes(),
            b'r="B5"',
            b'r="XFD1048576"',
        )

        with (
            patch.object(openpyxl, "load_workbook") as loader,
            self.assertRaisesRegex(ValueError, "单元格坐标超过安全上限"),
        ):
            parse_bested_questionnaire_upload("questionnaire.xlsx", content)
        loader.assert_not_called()

    def test_oversized_sheet_row_coordinate_is_rejected_before_openpyxl(self):
        content = _worksheet_xml_replaced(
            _questionnaire_bytes(),
            b'<row r="5">',
            b'<row r="100001">',
        )

        with (
            patch.object(openpyxl, "load_workbook") as loader,
            self.assertRaisesRegex(ValueError, "行坐标超过安全上限"),
        ):
            parse_bested_questionnaire_upload("questionnaire.xlsx", content)
        loader.assert_not_called()

    def test_oversized_merged_range_is_rejected_before_openpyxl(self):
        content = _worksheet_xml_replaced(
            _questionnaire_bytes(),
            b"</sheetData>",
            (
                b'</sheetData><mergeCells count="1">'
                b'<mergeCell ref="A1:ZZ100000"/></mergeCells>'
            ),
        )

        with (
            patch.object(openpyxl, "load_workbook") as loader,
            self.assertRaisesRegex(ValueError, "合并区域.*安全上限"),
        ):
            parse_bested_questionnaire_upload("questionnaire.xlsx", content)
        loader.assert_not_called()


if __name__ == "__main__":
    unittest.main()
